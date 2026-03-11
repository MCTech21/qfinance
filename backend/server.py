from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo.errors import DuplicateKeyError
from pymongo import ASCENDING, ReturnDocument
import os
import logging
import json
import re
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
from decimal import Decimal, InvalidOperation
import uuid
from datetime import datetime, timezone, date
import zlib
import unicodedata
import asyncio
import jwt
import bcrypt
from enum import Enum
import csv
import io
import xmlrpc.client
from PIL import Image
from openpyxl import Workbook, load_workbook
from dateutil import parser as date_parser
import pytz

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Environment variables
MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
JWT_SECRET = os.environ.get('JWT_SECRET', 'finrealty-secret-key-2024')
TIMEZONE = pytz.timezone('America/Tijuana')

# MongoDB connection
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(
    title="FinRealty API",
    version="1.0.0",
    description="Sistema de Control Financiero para Desarrollos Inmobiliarios",
    docs_url="/docs",
    redoc_url="/redoc"
)
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Root health check
@app.get("/")
async def root():
    return {"status": "ok", "service": "FinRealty API", "version": "1.0.0"}

@app.get("/api/health")
async def api_health():
    return {"status": "ok", "api": "up"}

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ========================= ENUMS =========================
class UserRole(str, Enum):
    ADMIN = "admin"
    FINANZAS = "finanzas"
    DIRECTOR = "director"
    AUTORIZADOR = "autorizador"
    SOLO_LECTURA = "solo_lectura"
    CAPTURA = "captura"
    CAPTURA_INGRESOS = "captura_ingresos"

class Currency(str, Enum):
    MXN = "MXN"
    USD = "USD"

class AuthorizationStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"


class BudgetApprovalStatus(str, Enum):
    NOT_REQUIRED = "not_required"
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"


class ApprovalType(str, Enum):
    MOVEMENT = "movement"
    BUDGET_DEFINITION = "budget_definition"
    OVERBUDGET_EXCEPTION = "overbudget_exception"
    PURCHASE_ORDER_WORKFLOW = "purchase_order_workflow"

class MovementStatus(str, Enum):
    POSTED = "posted"                    # Contabilizado (aprobado o no requirió autorización)
    PENDING_APPROVAL = "pending_approval" # Pendiente de autorización (no contabiliza)
    REJECTED = "rejected"                 # Rechazado (no contabiliza)


class PurchaseOrderStatus(str, Enum):
    DRAFT = "draft"
    PENDING_APPROVAL = "pending_approval"
    APPROVED_FOR_PAYMENT = "approved_for_payment"
    REJECTED = "rejected"
    CANCELLED = "cancelled"


class BudgetGateStatus(str, Enum):
    NOT_CHECKED = "not_checked"
    OK = "ok"
    EXCEPTION_PENDING = "exception_pending"


class PostingStatus(str, Enum):
    NOT_POSTED = "not_posted"
    POSTED = "posted"
    POST_FAILED = "post_failed"

class PartidaGrupo(str, Enum):
    OBRA = "obra"
    GYA = "gya"
    FINANCIEROS = "financieros"
    INGRESOS = "ingresos"

# ========================= MODELS =========================

# Empresa (Multiempresa)
class EmpresaBase(BaseModel):
    nombre: str
    is_active: bool = True

class Empresa(EmpresaBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: UserRole
    empresa_id: Optional[str] = None
    empresa_ids: List[str] = Field(default_factory=list)

class UserCreate(UserBase):
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True
    must_change_password: bool = False

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

class ForceChangePasswordRequest(BaseModel):
    new_password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User
    must_change_password: bool = False

class ProjectBase(BaseModel):
    code: str
    name: str
    empresa_id: str  # OBLIGATORIO - vincula a empresa
    description: Optional[str] = None
    client_id: Optional[str] = None
    is_active: bool = True

class Project(ProjectBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Catálogo de Partidas Presupuestales (source of truth)
class CatalogoPartidaBase(BaseModel):
    codigo: str  # 100, 101, 200, etc.
    nombre: str
    grupo: PartidaGrupo
    is_active: bool = True

class CatalogoPartida(CatalogoPartidaBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Legacy Partida - mantener para compatibilidad, pero usar catalogo_partidas
class PartidaBase(BaseModel):
    code: str
    name: str
    description: Optional[str] = None
    is_active: bool = True

class Partida(PartidaBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProviderBase(BaseModel):
    code: str
    name: str
    rfc: Optional[str] = None
    is_active: bool = True

class Provider(ProviderBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BudgetBase(BaseModel):
    project_id: str
    partida_codigo: str  # Código del catálogo (100, 101, etc.)
    year: int
    month: int
    amount_mxn: Decimal
    notes: Optional[str] = None


class BudgetPlanInput(BaseModel):
    project_id: str
    partida_codigo: str
    total: Decimal
    annual_breakdown: Dict[str, Decimal] = Field(default_factory=dict)
    monthly_breakdown: Dict[str, Decimal] = Field(default_factory=dict)
    notes: Optional[str] = None


class BudgetWriteInput(BaseModel):
    project_id: str
    partida_codigo: str
    total_amount: Optional[Any] = Decimal("0")
    annual_breakdown: Optional[Any] = Field(default_factory=dict)
    monthly_breakdown: Optional[Any] = Field(default_factory=dict)
    notes: Optional[str] = None


class ApprovalDecisionInput(BaseModel):
    comment: Optional[str] = None


class Budget(BudgetBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str


class BudgetRequestBase(BudgetBase):
    reason: Optional[str] = None


class BudgetRequest(BudgetRequestBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: AuthorizationStatus = AuthorizationStatus.PENDING
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    requested_by: str
    resolved_at: Optional[datetime] = None
    resolved_by: Optional[str] = None

class MovementBase(BaseModel):
    project_id: str
    partida_codigo: str  # Código del catálogo (100, 101, etc.)
    provider_id: Optional[str] = None
    client_id: Optional[str] = None
    customer_name: Optional[str] = None
    date: datetime
    currency: Currency
    amount_original: float
    exchange_rate: float
    amount_mxn: float
    reference: str
    description: Optional[str] = None

class Movement(MovementBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    status: MovementStatus = MovementStatus.POSTED
    authorization_id: Optional[str] = None
    reversal_of_id: Optional[str] = None

class MovementCreate(BaseModel):
    project_id: str
    partida_codigo: str  # Código del catálogo
    provider_id: Optional[str] = None
    customer_name: Optional[str] = None
    date: str
    currency: Currency
    amount_original: float
    exchange_rate: float
    reference: str
    description: Optional[str] = None
    client_id: Optional[str] = None


class PurchaseOrderLineInput(BaseModel):
    line_no: int
    partida_codigo: str
    sku: Optional[str] = None
    description: str
    qty: Decimal
    uom: Optional[str] = None
    price_unit: Decimal
    discount_pct: Optional[Decimal] = Decimal("0")
    iva_rate: Decimal = Decimal("16")
    apply_isr_withholding: bool = False
    isr_withholding_rate: Optional[Decimal] = Decimal("0")


class PurchaseOrderCreate(BaseModel):
    external_id: Optional[str] = None
    invoice_folio: Optional[str] = None
    supplier_invoice_folio: Optional[str] = None
    project_id: str
    vendor_name: str
    vendor_rfc: Optional[str] = None
    vendor_email: Optional[str] = None
    vendor_phone: Optional[str] = None
    vendor_address: Optional[str] = None
    currency: Currency = Currency.MXN
    exchange_rate: Optional[Decimal] = Decimal("1")
    apply_iva_withholding: bool = False
    iva_withholding_rate: Optional[Decimal] = Decimal("0")
    order_date: str
    planned_date: Optional[str] = None
    notes: Optional[str] = None
    payment_terms: Optional[str] = None
    fob: Optional[str] = None
    lines: List[PurchaseOrderLineInput]


class PurchaseOrderRejectInput(BaseModel):
    reason: str


class InvoiceCreate(BaseModel):
    empresa_id: str
    project_id: str
    provider_id: Optional[str] = None
    provider_name: Optional[str] = None
    invoice_folio: str
    currency: Currency = Currency.MXN
    exchange_rate: Optional[Decimal] = Decimal("1")
    invoice_total_original: Decimal
    status: Optional[str] = "OPEN"


class InvoicePayInput(BaseModel):
    mode: str
    advance_pct: Optional[Decimal] = None
    amount_original: Optional[Decimal] = None
    date: str
    reference: str
    description: Optional[str] = None
    partida_codigo: str


class OCBudgetPreviewInput(BaseModel):
    project_id: str
    order_date: str
    currency: Currency = Currency.MXN
    exchange_rate: Optional[Decimal] = Decimal("1")
    lines: List[Dict[str, Any]]


class InventoryItemBase(BaseModel):
    company_id: str
    project_id: str
    m2_superficie: Decimal
    m2_construccion: Optional[Decimal] = Decimal("0")
    lote_edificio: str
    manzana_departamento: str
    precio_m2_superficie: Decimal
    precio_m2_construccion: Optional[Decimal] = Decimal("0")
    descuento_bonificacion: Decimal = Decimal("0")


class InventoryItem(InventoryItemBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    precio_venta: Decimal
    precio_total: Decimal
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ClientBase(BaseModel):
    company_id: str
    project_id: str
    nombre: str
    telefono: Optional[str] = None
    domicilio: Optional[str] = None
    inventory_item_id: Optional[str] = None


class ClientCreateRequest(BaseModel):
    company_id: Optional[str] = None
    project_id: Optional[str] = None
    nombre: Optional[str] = None
    telefono: Optional[str] = None
    domicilio: Optional[str] = None
    inventory_item_id: Optional[str] = None

    # Legacy/UI aliases
    empresa: Optional[str] = None
    proyecto: Optional[str] = None
    inventario: Optional[str] = None


class ClientUpdate(BaseModel):
    nombre: Optional[str] = None
    telefono: Optional[str] = None
    domicilio: Optional[str] = None
    inventory_item_id: Optional[str] = None


class InventoryItemUpdate(BaseModel):
    project_id: Optional[str] = None
    m2_superficie: Optional[Decimal] = None
    m2_construccion: Optional[Decimal] = None
    lote_edificio: Optional[str] = None
    manzana_departamento: Optional[str] = None
    precio_m2_superficie: Optional[Decimal] = None
    precio_m2_construccion: Optional[Decimal] = None
    descuento_bonificacion: Optional[Decimal] = None
class Client(ClientBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    precio_venta_snapshot: Decimal = Decimal("0")
    saldo_restante: Decimal = Decimal("0")
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class MovementAdminUpdate(BaseModel):
    project_id: Optional[str] = None
    partida_codigo: Optional[str] = None
    provider_id: Optional[str] = None
    customer_name: Optional[str] = None
    date: Optional[str] = None
    currency: Optional[Currency] = None
    amount_original: Optional[float] = None
    exchange_rate: Optional[float] = None
    reference: Optional[str] = None
    description: Optional[str] = None
    reason: str


class MovementAdminAction(BaseModel):
    reason: str


class MovementPurgeRequest(BaseModel):
    reason: str
    dry_run: bool = False
    force: bool = False
    hard: bool = False
    project_id: Optional[str] = None
    year: Optional[int] = None
    month: Optional[int] = None

class AuthorizationBase(BaseModel):
    movement_id: Optional[str] = None
    reason: str
    requested_by: str

class Authorization(AuthorizationBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    status: AuthorizationStatus = AuthorizationStatus.PENDING
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    resolved_at: Optional[datetime] = None
    resolved_by: Optional[str] = None
    notes: Optional[str] = None

class AuthorizationResolve(BaseModel):
    status: AuthorizationStatus
    notes: Optional[str] = None
    partial_amount: Optional[Decimal] = None

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    user_role: str
    action: str
    entity: str
    entity_id: str
    changes: Dict[str, Any]
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExchangeRate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    rate: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConfigSetting(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    key: str
    value: Any
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: str

class OdooIntegrationConfigInput(BaseModel):
    odoo_mode: str = "stub"
    odoo_url: Optional[str] = ""
    odoo_db: Optional[str] = ""
    odoo_username: Optional[str] = ""
    odoo_api_key: Optional[str] = ""
    default_model: Optional[str] = "purchase.order"


class CSVImportResult(BaseModel):
    total_filas: int
    insertadas: int
    rechazadas: int
    duplicadas_omitidas: int
    errores: List[Dict[str, Any]]
    movements_created: List[str]
    authorizations_required: List[str]

class ExportAuditEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    action: str  # "IMPORT" or "EXPORT"
    timestamp_inicio: datetime
    timestamp_fin: Optional[datetime] = None
    filtros: Optional[Dict[str, Any]] = None
    conteos: Optional[Dict[str, Any]] = None
    errores_resumen: Optional[List[str]] = None

# ========================= HELPERS =========================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def validate_password_policy(password: str):
    if len(password) < 8:
        raise HTTPException(status_code=422, detail="La nueva contraseña debe tener al menos 8 caracteres")
    if not any(ch.isalpha() for ch in password) or not any(ch.isdigit() for ch in password):
        raise HTTPException(status_code=422, detail="La nueva contraseña debe contener letras y números")


def create_token(user_id: str, email: str, role: str, must_change_password: bool = False, empresa_id: Optional[str] = None, empresa_ids: Optional[List[str]] = None) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "empresa_id": empresa_id,
        "empresa_ids": empresa_ids or [],
        "must_change_password": must_change_password,
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


async def get_current_user(request: Request, credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=["HS256"])
        if payload.get("must_change_password"):
            allowed_paths = {"/api/auth/force-change-password", "/api/auth/logout", "/api/auth/me", "/api/auth/permissions"}
            if request.url.path not in allowed_paths:
                raise HTTPException(status_code=403, detail="Debes cambiar tu contraseña para continuar")
        payload["role"] = normalize_role_input(payload.get("role")) or payload.get("role")
        payload["empresa_id"] = payload.get("empresa_id") or payload.get("company_id") or payload.get("company") or payload.get("empresa")
        payload["empresa_ids"] = payload.get("empresa_ids") or []
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

def require_roles(*roles: UserRole):
    async def role_checker(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in [r.value for r in roles]:
            raise HTTPException(status_code=403, detail="Permisos insuficientes para esta acción")
        return current_user
    return role_checker

# ========================= RBAC PERMISSIONS MATRIX =========================
# Define granular permissions per action type
class Permission(str, Enum):
    # Dashboard/Reports
    VIEW_DASHBOARD = "view_dashboard"
    VIEW_REPORTS = "view_reports"
    EXPORT_DATA = "export_data"
    
    # Movements
    CREATE_MOVEMENT = "create_movement"
    VIEW_MOVEMENTS = "view_movements"
    IMPORT_MOVEMENTS = "import_movements"
    
    # Authorizations
    VIEW_AUTHORIZATIONS = "view_authorizations"
    APPROVE_REJECT = "approve_reject"
    
    # Catalogs (empresas, proyectos, partidas, proveedores)
    VIEW_CATALOGS = "view_catalogs"
    MANAGE_CATALOGS = "manage_catalogs"
    
    # Budgets
    VIEW_BUDGETS = "view_budgets"
    MANAGE_BUDGETS = "manage_budgets"
    REQUEST_BUDGETS = "request_budgets"
    
    # Users
    VIEW_USERS = "view_users"
    MANAGE_USERS = "manage_users"
    
    # Audit
    VIEW_AUDIT = "view_audit"
    EXPORT_AUDIT = "export_audit"

# Role -> Permissions mapping
ROLE_PERMISSIONS = {
    UserRole.ADMIN.value: [p.value for p in Permission],  # Admin: ALL
    
    UserRole.FINANZAS.value: [
        Permission.VIEW_DASHBOARD.value,
        Permission.VIEW_REPORTS.value,
        Permission.EXPORT_DATA.value,
        Permission.CREATE_MOVEMENT.value,
        Permission.VIEW_MOVEMENTS.value,
        Permission.IMPORT_MOVEMENTS.value,
        Permission.VIEW_AUTHORIZATIONS.value,
        Permission.VIEW_CATALOGS.value,
        Permission.VIEW_BUDGETS.value,
        Permission.REQUEST_BUDGETS.value,
        Permission.MANAGE_CATALOGS.value,
    ],

    UserRole.DIRECTOR.value: [
        Permission.VIEW_DASHBOARD.value,
        Permission.VIEW_REPORTS.value,
        Permission.VIEW_MOVEMENTS.value,
        Permission.VIEW_AUTHORIZATIONS.value,
        Permission.APPROVE_REJECT.value,
        Permission.VIEW_CATALOGS.value,
        Permission.VIEW_BUDGETS.value,
    ],
    
    UserRole.AUTORIZADOR.value: [
        Permission.VIEW_DASHBOARD.value,
        Permission.VIEW_REPORTS.value,
        Permission.EXPORT_DATA.value,
        Permission.VIEW_MOVEMENTS.value,
        Permission.VIEW_AUTHORIZATIONS.value,
        Permission.APPROVE_REJECT.value,
        Permission.VIEW_CATALOGS.value,
        Permission.VIEW_BUDGETS.value,
        Permission.VIEW_AUDIT.value,
    ],
    
    UserRole.CAPTURA_INGRESOS.value: [
        Permission.VIEW_DASHBOARD.value,
        Permission.VIEW_REPORTS.value,
        Permission.CREATE_MOVEMENT.value,
        Permission.VIEW_MOVEMENTS.value,
        Permission.VIEW_CATALOGS.value,
        Permission.VIEW_BUDGETS.value,
    ],

    UserRole.CAPTURA.value: [
        Permission.VIEW_DASHBOARD.value,
        Permission.VIEW_REPORTS.value,
        Permission.CREATE_MOVEMENT.value,
        Permission.VIEW_MOVEMENTS.value,
        Permission.VIEW_CATALOGS.value,
        Permission.VIEW_BUDGETS.value,
    ],

    UserRole.SOLO_LECTURA.value: [
        Permission.VIEW_DASHBOARD.value,
        Permission.VIEW_REPORTS.value,
        Permission.VIEW_MOVEMENTS.value,
        Permission.VIEW_AUTHORIZATIONS.value,
        Permission.VIEW_CATALOGS.value,
        Permission.VIEW_BUDGETS.value,
    ],
}

def require_permission(*permissions: Permission):
    """Check if user has at least one of the required permissions"""
    async def permission_checker(current_user: dict = Depends(get_current_user)):
        user_role = normalize_role_input(current_user.get("role")) or current_user.get("role")
        user_permissions = ROLE_PERMISSIONS.get(user_role, [])
        
        if not any(p.value in user_permissions for p in permissions):
            raise HTTPException(
                status_code=403, 
                detail=f"Permisos insuficientes. Se requiere: {', '.join([p.value for p in permissions])}"
            )
        return current_user
    return permission_checker

async def log_audit(user: dict, action: str, entity_type: str, entity_id: str, changes: dict, ip_address: str = None):
    """Enhanced audit logging with IP and standardized format"""
    audit = {
        "id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "user_email": user["email"],
        "user_role": user["role"],
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "changes": to_json_safe(changes),
        "ip_address": ip_address,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(audit)

async def log_admin_action(
    request: Request,
    user: dict,
    action: str,
    entity: str,
    entity_id: str,
    success: bool,
    before: Optional[dict] = None,
    after: Optional[dict] = None,
    message: Optional[str] = None,
):
    payload = {
        "result": "success" if success else "failure",
        "before": before,
        "after": after,
        "message": message,
        "user_agent": request.headers.get("user-agent"),
    }
    await log_audit(user, action, entity, entity_id, payload, ip_address=request.client.host if request.client else None)


def ensure_admin(current_user: dict) -> dict:
    if current_user.get("role") != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    return current_user


def active_query(include_inactive: bool = False, extra: Optional[dict] = None) -> dict:
    query = extra.copy() if extra else {}
    if not include_inactive:
        query["is_active"] = {"$ne": False}
    return query


def movement_active_query(include_deleted: bool = False, extra: Optional[dict] = None) -> dict:
    query = extra.copy() if extra else {}
    if not include_deleted:
        query["is_deleted"] = {"$ne": True}
    return query


async def assert_no_references(entity: str, entity_id: str):
    checks = {
        "empresas": [("projects", {"empresa_id": entity_id}, "proyectos")],
        "projects": [
            ("budgets", {"project_id": entity_id}, "presupuestos"),
            ("movements", {"project_id": entity_id}, "movimientos"),
        ],
        "providers": [("movements", {"provider_id": entity_id}, "movimientos")],
        "users": [
            ("movements", {"created_by": entity_id}, "movimientos"),
            ("budgets", {"created_by": entity_id}, "presupuestos"),
            ("authorizations", {"requested_by": entity_id}, "autorizaciones"),
        ],
    }
    for collection, query, label in checks.get(entity, []):
        if await db[collection].find_one(query, {"_id": 0}):
            raise HTTPException(status_code=409, detail=f"No se puede eliminar físicamente: tiene referencias en {label}")

def get_traffic_light(percentage: float, threshold_yellow: float = 90.0, threshold_red: float = 100.0) -> str:
    if percentage < threshold_yellow:
        return "green"
    if percentage < threshold_red:
        return "yellow"
    return "red"


def build_budget_signal(budget_total: Decimal, executed_total: Decimal, threshold_yellow: Decimal, threshold_red: Decimal) -> dict:
    budget = money_dec(budget_total).quantize(TWO_DECIMALS)
    executed = money_dec(executed_total).quantize(TWO_DECIMALS)
    variation = (budget - executed).quantize(TWO_DECIMALS)
    if budget == 0 and executed == 0:
        return {
            "traffic_light": "green",
            "status_label": "SIN PRESUPUESTO (sin gasto)",
            "porcentaje": Decimal("0.00"),
            "porcentaje_label": "0.00",
            "variation_color": "neutral",
            "variacion": variation,
        }
    if budget == 0 and executed > 0:
        return {
            "traffic_light": "yellow",
            "status_label": "SIN PRESUPUESTO (con gasto)",
            "porcentaje": None,
            "porcentaje_label": "N/A",
            "variation_color": "yellow",
            "variacion": variation,
        }
    porcentaje = ((executed / budget) * Decimal("100")).quantize(TWO_DECIMALS)
    return {
        "traffic_light": get_traffic_light(float(porcentaje), float(threshold_yellow), float(threshold_red)),
        "status_label": "OK" if executed <= budget else "EXCEDIDO",
        "porcentaje": porcentaje,
        "porcentaje_label": str(porcentaje),
        "variation_color": "green" if variation >= 0 else "red",
        "variacion": variation,
    }


async def get_dashboard_thresholds() -> tuple[Decimal, Decimal]:
    rows = await db.config.find({"key": {"$in": ["threshold_yellow", "threshold_red"]}}, {"_id": 0}).to_list(20)
    values = {r.get("key"): r.get("value") for r in rows}
    yellow = decimal_from_value(values.get("threshold_yellow", "90"), "threshold_yellow")
    red = decimal_from_value(values.get("threshold_red", "100"), "threshold_red")
    if red < yellow:
        yellow, red = red, yellow
    return yellow, red

def parse_date_tijuana(date_str: str) -> datetime:
    parsed = date_parser.parse(date_str)
    if parsed.tzinfo is None:
        parsed = TIMEZONE.localize(parsed)
    return parsed.astimezone(pytz.UTC)

def to_tijuana(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        dt = pytz.UTC.localize(dt)
    return dt.astimezone(TIMEZONE)

def get_year_range() -> tuple[int, int]:
    current_year = to_tijuana(datetime.now(timezone.utc)).year
    from_year = min(2025, current_year)
    to_year = max(current_year + 10, 2031)
    return from_year, to_year

def validate_year_in_range(year: int):
    from_year, to_year = get_year_range()
    if year < from_year or year > to_year:
        raise HTTPException(status_code=422, detail=f"Año fuera de rango permitido ({from_year}-{to_year})")

def validate_date_in_range(dt: datetime):
    validate_year_in_range(to_tijuana(dt).year)

def is_ingresos_partida(codigo: str) -> bool:
    return str(codigo).startswith("4")


def require_client_write_access():
    async def checker(current_user: dict = Depends(get_current_user)):
        role = normalize_role_input(current_user.get("role")) or current_user.get("role")
        if role in {UserRole.ADMIN.value, UserRole.FINANZAS.value, UserRole.CAPTURA.value, UserRole.CAPTURA_INGRESOS.value}:
            current_user["role"] = role
            return current_user
        raise HTTPException(status_code=403, detail="Permisos insuficientes para gestionar clientes")
    return checker


CAPTURA_ALLOWED_BUDGET_CODES = {"103", "203", "206", "402", "403"}
NO_PROVIDER_BUDGET_CODES = {"402", "403"}


def _is_ingresos_code(code: str) -> bool:
    normalized = str(code or "").strip()
    return normalized.startswith("4")


def sanitize_mongo_document(doc: dict) -> dict:
    if not doc:
        return doc
    clean = dict(doc)
    mongo_id = clean.pop("_id", None)
    if mongo_id is not None and not clean.get("id"):
        clean["id"] = str(mongo_id)
    return clean


def to_json_safe(value: Any):
    if isinstance(value, dict):
        return {k: to_json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [to_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return [to_json_safe(v) for v in value]
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def normalize_role_input(role: Optional[str]) -> Optional[str]:
    if role is None:
        return None
    role_normalized = str(role).strip().lower().replace("-", "_").replace(" ", "_")
    compact = role_normalized.replace("_", "")
    if compact in {"capturaingresos", "captura_ingresos".replace("_", "")}: 
        return UserRole.CAPTURA_INGRESOS.value
    if compact == "captura":
        return UserRole.CAPTURA.value
    return role_normalized


def is_capture_role(role: Optional[str]) -> bool:
    return role in {"captura", UserRole.CAPTURA_INGRESOS.value}


def is_operational_role(role: Optional[str]) -> bool:
    normalized = normalize_role_input(role) or role
    return normalized in {UserRole.CAPTURA.value, UserRole.CAPTURA_INGRESOS.value, UserRole.FINANZAS.value}


def get_user_company_id(current_user: dict) -> Optional[str]:
    return (
        current_user.get("empresa_id")
        or current_user.get("company_id")
        or current_user.get("company")
        or current_user.get("empresa")
    )


def enforce_capture_budget_scope(current_user: dict, budget_code: str):
    normalized = str(budget_code)
    role = current_user.get("role")
    if role == UserRole.CAPTURA_INGRESOS.value:
        if not _is_ingresos_code(normalized):
            raise HTTPException(status_code=403, detail="Rol captura_ingresos solo puede operar partidas de ingresos (400)")
        return
    if role in {UserRole.CAPTURA.value, "captura"} and normalized not in CAPTURA_ALLOWED_BUDGET_CODES:
        raise HTTPException(
            status_code=403,
            detail=f"Rol captura solo puede operar partidas: {', '.join(sorted(CAPTURA_ALLOWED_BUDGET_CODES))}",
        )


def requires_budget(budget_code: str) -> bool:
    try:
        code = int(str(budget_code))
    except (TypeError, ValueError):
        return False
    return (100 <= code <= 199) or (200 <= code <= 299)




def decimal_from_value(value: Any, field_name: str = "value") -> Decimal:
    try:
        dec = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        raise HTTPException(status_code=422, detail={"code": "invalid_decimal", "message": f"{field_name} debe ser decimal válido"})
    return dec


def ensure_non_negative(value: Decimal, field_name: str):
    if value < Decimal("0"):
        raise HTTPException(status_code=422, detail={"code": "negative_not_allowed", "message": f"{field_name} no puede ser negativo"})


def user_company_scope_query(current_user: dict, company_field: str = "company_id") -> dict:
    role = current_user.get("role")
    if role in {UserRole.ADMIN.value, UserRole.FINANZAS.value}:
        return {}
    company_id = get_user_company_id(current_user)
    if not company_id:
        raise HTTPException(status_code=422, detail={"code": "empresa_not_selected", "message": "Selecciona la empresa para operar"})
    return {company_field: company_id}


def enforce_company_access(current_user: dict, company_id: Optional[str]):
    role = current_user.get("role")
    if role in {UserRole.ADMIN.value, UserRole.FINANZAS.value}:
        return
    user_company_id = get_user_company_id(current_user)
    if not company_id or user_company_id != company_id:
        raise HTTPException(status_code=403, detail="Acceso restringido a la empresa del usuario")


def has_company_access(current_user: dict, company_id: Optional[str]) -> bool:
    try:
        enforce_company_access(current_user, company_id)
        return True
    except HTTPException:
        return False


def get_budget_approval_mode() -> str:
    return os.getenv("BUDGET_APPROVAL_MODE", "soft").strip().lower() or "soft"


def get_overbudget_approval_mode() -> str:
    return os.getenv("OVERBUDGET_APPROVAL_MODE", "reject_and_request").strip().lower() or "reject_and_request"


def get_overbudget_admin_bypass_enabled() -> bool:
    raw = os.getenv("OVERBUDGET_ADMIN_BYPASS", "true").strip().lower()
    return raw not in {"0", "false", "no", "off"}


def is_admin_or_bypass(current_user: dict) -> bool:
    role = current_user.get("role")
    return role in {UserRole.ADMIN.value}


def enforce_oc_for_finanzas_egress_enabled() -> bool:
    raw = os.getenv("QF_ENFORCE_OC_FOR_EGRESS", "false").strip().lower()
    return raw in {"1", "true", "yes", "on"}


def decimal_map_to_strings(values: Dict[str, Decimal]) -> Dict[str, str]:
    return {k: str(v) for k, v in values.items()}


def normalize_for_sort(value: Optional[str]) -> str:
    raw = (value or "").strip().lower()
    normalized = unicodedata.normalize("NFD", raw)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


TWO_DECIMALS = Decimal("0.01")


def money_dec(value: Any) -> Decimal:
    return decimal_from_value(value or 0, "money").quantize(TWO_DECIMALS)


def parse_amount_like(value: Any, field_name: str) -> Decimal:
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    return money_dec(value)


def _validate_iva_rate(rate: Decimal):
    if rate not in {Decimal("0"), Decimal("8"), Decimal("16")}:
        raise HTTPException(status_code=422, detail={"code": "invalid_iva_rate", "message": "IVA inválido. Solo 0, 8, 16", "meta": {"iva_rate": str(rate)}})


def calculate_oc_line(line: PurchaseOrderLineInput) -> dict:
    qty = parse_amount_like(line.qty, "qty")
    unit = parse_amount_like(line.price_unit, "price_unit")
    if qty <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_amount", "message": "qty debe ser mayor a 0", "meta": {"line_no": line.line_no}})
    if unit < 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_amount", "message": "price_unit no puede ser negativo", "meta": {"line_no": line.line_no}})

    discount_pct = parse_amount_like(line.discount_pct or 0, "discount_pct")
    if discount_pct < 0 or discount_pct > 100:
        raise HTTPException(status_code=422, detail={"code": "invalid_discount_pct", "message": "discount_pct debe estar entre 0 y 100", "meta": {"line_no": line.line_no}})

    iva_rate = parse_amount_like(line.iva_rate, "iva_rate")
    _validate_iva_rate(iva_rate)
    isr_rate = parse_amount_like(line.isr_withholding_rate or 0, "isr_withholding_rate") if line.apply_isr_withholding else Decimal("0")
    if isr_rate < 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_isr_withholding_rate", "message": "Tasa ISR inválida", "meta": {"line_no": line.line_no}})

    subtotal_before_discount = (qty * unit).quantize(TWO_DECIMALS)
    discount_amount = (subtotal_before_discount * (discount_pct / Decimal("100"))).quantize(TWO_DECIMALS)
    taxable_base = (subtotal_before_discount - discount_amount).quantize(TWO_DECIMALS)
    if taxable_base < 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_tax_calculation", "message": "Base gravable inválida", "meta": {"line_no": line.line_no}})
    iva_amount = (taxable_base * (iva_rate / Decimal("100"))).quantize(TWO_DECIMALS)
    isr_amount = (taxable_base * (isr_rate / Decimal("100"))).quantize(TWO_DECIMALS) if line.apply_isr_withholding else Decimal("0")
    line_total = (taxable_base + iva_amount - isr_amount).quantize(TWO_DECIMALS)
    if line_total < 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_tax_calculation", "message": "Total de línea inválido", "meta": {"line_no": line.line_no}})

    return {
        "id": str(uuid.uuid4()),
        "line_no": line.line_no,
        "partida_codigo": str(line.partida_codigo),
        "sku": line.sku,
        "description": line.description,
        "qty": str(qty),
        "uom": line.uom,
        "price_unit": str(unit),
        "discount_pct": str(discount_pct),
        "discount_amount": str(discount_amount),
        "subtotal_before_discount": str(subtotal_before_discount),
        "taxable_base": str(taxable_base),
        "iva_rate": str(iva_rate),
        "iva_amount": str(iva_amount),
        "apply_isr_withholding": line.apply_isr_withholding,
        "isr_withholding_rate": str(isr_rate),
        "isr_withholding_amount": str(isr_amount),
        "line_total": str(line_total),
    }


def summarize_oc_lines(lines: List[dict], apply_iva_withholding: bool = False, iva_withholding_rate: Decimal = Decimal("0"), currency: str = "MXN", exchange_rate: Decimal = Decimal("1")) -> dict:
    subtotal_tax_base = sum((money_dec(l.get("taxable_base", 0)) for l in lines), Decimal("0"))
    tax_total = sum((money_dec(l.get("iva_amount", 0)) for l in lines), Decimal("0"))
    withholding_isr_total = sum((money_dec(l.get("isr_withholding_amount", 0)) for l in lines), Decimal("0"))
    iva_withheld = Decimal("0")
    if apply_iva_withholding:
        if iva_withholding_rate <= 0:
            raise HTTPException(status_code=422, detail={"code": "invalid_iva_withholding_rate", "message": "iva_withholding_rate debe ser mayor a 0 cuando aplica retención IVA"})
        iva_withheld = (tax_total * (iva_withholding_rate / Decimal("100"))).quantize(TWO_DECIMALS)
    total = (subtotal_tax_base + tax_total - iva_withheld - withholding_isr_total).quantize(TWO_DECIMALS)
    total_mxn = total if currency == Currency.MXN.value else (total * exchange_rate).quantize(TWO_DECIMALS)
    return {
        "subtotal_tax_base": str(subtotal_tax_base.quantize(TWO_DECIMALS)),
        "tax_total": str(tax_total.quantize(TWO_DECIMALS)),
        "iva_withholding_total": str(iva_withheld.quantize(TWO_DECIMALS)),
        "withholding_isr_total": str(withholding_isr_total.quantize(TWO_DECIMALS)),
        "total": str(total),
        "total_mxn": str(total_mxn),
    }


async def evaluate_oc_budget_gate(po: dict) -> Dict[str, Any]:
    order_date = date_parser.parse(po.get("order_date")) if isinstance(po.get("order_date"), str) else po.get("order_date")
    currency = (po.get("currency") or Currency.MXN.value).upper()
    exchange_rate = money_dec(po.get("exchange_rate", 1))
    if currency != Currency.MXN.value and exchange_rate <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_exchange_rate", "message": "exchange_rate debe ser mayor a 0 para moneda distinta a MXN"})
    exceeded = []
    lines_meta = []
    for line in po.get("lines", []):
        partida = str(line.get("partida_codigo"))
        if partida in {"400", "401", "402", "403", "404"}:
            continue
        requested_original = money_dec(line.get("line_total", 0))
        requested = requested_original if currency == Currency.MXN.value else (requested_original * exchange_rate).quantize(TWO_DECIMALS)
        if requested <= 0:
            continue
        overbudget = await evaluate_overbudget(po.get("project_id"), partida, order_date, requested)
        if overbudget:
            meta = overbudget.get("metadata", {})
            meta.update({"partida_codigo": partida, "requested": str(requested), "requested_original": str(requested_original)})
            exceeded.append(meta)
        lines_meta.append({"partida_codigo": partida, "requested": str(requested), "requested_original": str(requested_original)})
    return {"ok": len(exceeded) == 0, "exceeded": exceeded, "checked": lines_meta}



async def ensure_pending_approval(
    *,
    approval_type: ApprovalType,
    company_id: str,
    project_id: str,
    requested_by: str,
    budget_id: Optional[str] = None,
    movement_id: Optional[str] = None,
    purchase_order_id: Optional[str] = None,
    dedupe_key: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
):
    query = {
        "approval_type": approval_type.value,
        "status": AuthorizationStatus.PENDING.value,
        "company_id": company_id,
        "project_id": project_id,
    }
    if budget_id:
        query["budget_id"] = budget_id
    if movement_id:
        query["movement_id"] = movement_id
    if purchase_order_id:
        query["purchase_order_id"] = purchase_order_id
    if dedupe_key:
        query["dedupe_key"] = dedupe_key

    existing = await db.authorizations.find_one(query, {"_id": 0})
    if existing:
        return existing

    doc = {
        "id": str(uuid.uuid4()),
        "approval_type": approval_type.value,
        "status": AuthorizationStatus.PENDING.value,
        "company_id": company_id,
        "project_id": project_id,
        "budget_id": budget_id,
        "movement_id": movement_id,
        "purchase_order_id": purchase_order_id,
        "reason": "",
        "requested_by": requested_by,
        "requested_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "metadata": to_json_safe(metadata or {}),
        "dedupe_key": dedupe_key,
    }
    await db.authorizations.insert_one(doc)
    return doc


async def get_budget_plan_with_scope(budget_id: str, current_user: dict) -> dict:
    budget = await db.budget_plans.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    enforce_company_access(current_user, budget.get("company_id"))
    return budget


def normalize_plan_response(doc: dict) -> dict:
    normalized = sanitize_mongo_document(doc)
    for fld in ("total_amount",):
        if fld in normalized and normalized[fld] is not None:
            normalized[fld] = str(normalized[fld])
    for fld in ("annual_breakdown", "monthly_breakdown"):
        if fld in normalized and isinstance(normalized[fld], dict):
            normalized[fld] = {k: str(v) for k, v in normalized[fld].items()}
    return normalized


def normalize_decimal_document(doc: dict) -> dict:
    return sanitize_mongo_document(to_json_safe(doc))


async def evaluate_overbudget(project_id: str, partida_codigo: str, movement_date: datetime, movement_amount: Decimal):
    plan = await db.budget_plans.find_one({"project_id": project_id, "partida_codigo": partida_codigo}, {"_id": 0})
    if not plan:
        return None

    year = str(movement_date.year)
    ym = f"{movement_date.year:04d}-{movement_date.month:02d}"

    posted_movements = await db.movements.find(movement_active_query(extra={
        "project_id": project_id,
        "partida_codigo": partida_codigo,
        "status": MovementStatus.POSTED.value,
    }), {"_id": 0}).to_list(5000)

    total_executed = Decimal("0")
    annual_executed = Decimal("0")
    monthly_executed = Decimal("0")
    for mov in posted_movements:
        amount = decimal_from_value(mov.get("amount_mxn", 0), "amount_mxn")
        mov_date = date_parser.parse(mov["date"]) if isinstance(mov.get("date"), str) else mov.get("date")
        total_executed += amount
        if mov_date.year == movement_date.year:
            annual_executed += amount
        if mov_date.year == movement_date.year and mov_date.month == movement_date.month:
            monthly_executed += amount

    total_budget = decimal_from_value(plan.get("total_amount", "0"), "total_amount") if plan.get("total_amount") is not None else Decimal("0")
    annual_breakdown = plan.get("annual_breakdown") or {}
    monthly_breakdown = plan.get("monthly_breakdown") or {}
    annual_budget = decimal_from_value(annual_breakdown.get(year, "0"), f"annual_breakdown.{year}") if annual_breakdown.get(year) is not None else None
    monthly_budget = decimal_from_value(monthly_breakdown.get(ym, "0"), f"monthly_breakdown.{ym}") if monthly_breakdown.get(ym) is not None else None

    has_detail = bool(annual_breakdown) or bool(monthly_breakdown)
    effective_scope = "total"
    scope_budget = total_budget
    scope_executed = total_executed
    if monthly_budget is not None:
        effective_scope = "monthly"
        scope_budget = monthly_budget
        scope_executed = monthly_executed
    elif annual_budget is not None:
        effective_scope = "annual"
        scope_budget = annual_budget
        scope_executed = annual_executed

    total_available = total_budget - total_executed
    scope_available = scope_budget - scope_executed
    total_exceeded = total_budget > Decimal("0") and (total_executed + movement_amount) > total_budget
    scope_exceeded = (scope_executed + movement_amount) > scope_budget

    if not total_exceeded and not scope_exceeded:
        return None

    triggered = []
    if scope_exceeded:
        triggered.append(effective_scope)
    if total_exceeded and "total" not in triggered:
        triggered.append("total")

    metadata = {
        "budget_id": plan.get("id"),
        "triggered_buckets": triggered,
        "project_id": project_id,
        "partida_codigo": partida_codigo,
        "scope": "total" if total_exceeded else effective_scope,
        "year": movement_date.year,
        "month": movement_date.month,
        "requested": str(movement_amount),
        "available": str(total_available if total_exceeded else scope_available),
        "effective_scope": effective_scope,
        "detail_present": has_detail,
        "total": {
            "budget": str(total_budget),
            "executed": str(total_executed),
            "available": str(total_available),
        },
        "scoped": {
            "scope": effective_scope,
            "budget": str(scope_budget),
            "executed": str(scope_executed),
            "available": str(scope_available),
        },
    }
    return {
        "plan": plan,
        "metadata": metadata,
        "scope_exceeded": scope_exceeded,
        "total_exceeded": total_exceeded,
    }


async def compute_budget_availability(project_id: str, partida_codigo: str, movement_date: datetime) -> Dict[str, Any]:
    year = movement_date.year
    month = movement_date.month
    ym = f"{year:04d}-{month:02d}"

    posted_movements = await db.movements.find(movement_active_query(extra={
        "project_id": project_id,
        "partida_codigo": partida_codigo,
        "status": MovementStatus.POSTED.value,
    }), {"_id": 0}).to_list(5000)

    executed_total = Decimal("0.00")
    executed_annual = Decimal("0.00")
    executed_monthly = Decimal("0.00")
    for mov in posted_movements:
        amt = money_dec(mov.get("amount_mxn", 0))
        mov_date = date_parser.parse(mov["date"]) if isinstance(mov.get("date"), str) else mov.get("date")
        executed_total += amt
        if mov_date.year == year:
            executed_annual += amt
        if mov_date.year == year and mov_date.month == month:
            executed_monthly += amt

    plan = await db.budget_plans.find_one({"project_id": project_id, "partida_codigo": partida_codigo}, {"_id": 0})
    if plan:
        annual_breakdown = plan.get("annual_breakdown") or {}
        monthly_breakdown = plan.get("monthly_breakdown") or {}
        total_budget = money_dec(plan.get("total_amount", 0))
        annual_budget = money_dec(annual_breakdown.get(str(year), 0)) if str(year) in annual_breakdown else None
        monthly_budget = money_dec(monthly_breakdown.get(ym, 0)) if ym in monthly_breakdown else None

        effective_scope = "monthly" if monthly_budget is not None else ("annual" if annual_budget is not None else "total")
        effective_budget = monthly_budget if monthly_budget is not None else (annual_budget if annual_budget is not None else total_budget)
        effective_executed = executed_monthly if monthly_budget is not None else (executed_annual if annual_budget is not None else executed_total)
        effective_remaining = (effective_budget - effective_executed).quantize(TWO_DECIMALS)
        total_remaining = (total_budget - executed_total).quantize(TWO_DECIMALS)

        return {
            "has_budget": True,
            "source": "plan",
            "budget_id": plan.get("id"),
            "effective_scope": effective_scope,
            "budget_validation_applies": True,
            "zero_is_allowed": True,
            "can_post_if_exact_zero": True,
            "budget_total_amount": str(total_budget),
            "executed_total": str(executed_total.quantize(TWO_DECIMALS)),
            "remaining_total": str(total_remaining),
            "usage_pct_total": float(((executed_total / total_budget) * Decimal("100")).quantize(TWO_DECIMALS)) if total_budget > 0 else 0.0,
            "annual_budget": str(annual_budget) if annual_budget is not None else None,
            "executed_annual": str(executed_annual.quantize(TWO_DECIMALS)),
            "remaining_annual": str((annual_budget - executed_annual).quantize(TWO_DECIMALS)) if annual_budget is not None else None,
            "usage_pct_annual": float(((executed_annual / annual_budget) * Decimal("100")).quantize(TWO_DECIMALS)) if annual_budget and annual_budget > 0 else None,
            "monthly_budget": str(monthly_budget) if monthly_budget is not None else None,
            "executed_monthly": str(executed_monthly.quantize(TWO_DECIMALS)),
            "remaining_monthly": str((monthly_budget - executed_monthly).quantize(TWO_DECIMALS)) if monthly_budget is not None else None,
            "usage_pct_monthly": float(((executed_monthly / monthly_budget) * Decimal("100")).quantize(TWO_DECIMALS)) if monthly_budget and monthly_budget > 0 else None,
            "effective_budget": str(effective_budget.quantize(TWO_DECIMALS)),
            "effective_executed": str(effective_executed.quantize(TWO_DECIMALS)),
            "effective_remaining": str(effective_remaining),
            "rules": ["monthly_if_present", "annual_if_month_missing", "total_fallback", "total_hard_cap", "0_is_ok"],
        }

    legacy_rows = await db.budgets.find({
        "project_id": project_id,
        "partida_codigo": partida_codigo,
    }, {"_id": 0}).to_list(5000)

    monthly_budget = Decimal("0.00")
    annual_budget = Decimal("0.00")
    total_budget = Decimal("0.00")
    for row in legacy_rows:
        amount = money_dec(row.get("amount_mxn", 0))
        total_budget += amount
        if row.get("year") == year:
            annual_budget += amount
            if row.get("month") == month:
                monthly_budget += amount

    if legacy_rows:
        effective_scope = "monthly" if monthly_budget > 0 else ("annual" if annual_budget > 0 else "total")
        effective_budget = monthly_budget if monthly_budget > 0 else (annual_budget if annual_budget > 0 else total_budget)
        effective_executed = executed_monthly if monthly_budget > 0 else (executed_annual if annual_budget > 0 else executed_total)
        effective_remaining = (effective_budget - effective_executed).quantize(TWO_DECIMALS)
        total_remaining = (total_budget - executed_total).quantize(TWO_DECIMALS)
        return {
            "has_budget": True,
            "source": "legacy",
            "effective_scope": effective_scope,
            "budget_validation_applies": True,
            "zero_is_allowed": True,
            "can_post_if_exact_zero": True,
            "budget_total_amount": str(total_budget.quantize(TWO_DECIMALS)),
            "executed_total": str(executed_total.quantize(TWO_DECIMALS)),
            "remaining_total": str(total_remaining),
            "usage_pct_total": float(((executed_total / total_budget) * Decimal("100")).quantize(TWO_DECIMALS)) if total_budget > 0 else 0.0,
            "annual_budget": str(annual_budget.quantize(TWO_DECIMALS)) if annual_budget > 0 else None,
            "executed_annual": str(executed_annual.quantize(TWO_DECIMALS)),
            "remaining_annual": str((annual_budget - executed_annual).quantize(TWO_DECIMALS)) if annual_budget > 0 else None,
            "usage_pct_annual": float(((executed_annual / annual_budget) * Decimal("100")).quantize(TWO_DECIMALS)) if annual_budget > 0 else None,
            "monthly_budget": str(monthly_budget.quantize(TWO_DECIMALS)) if monthly_budget > 0 else None,
            "executed_monthly": str(executed_monthly.quantize(TWO_DECIMALS)),
            "remaining_monthly": str((monthly_budget - executed_monthly).quantize(TWO_DECIMALS)) if monthly_budget > 0 else None,
            "usage_pct_monthly": float(((executed_monthly / monthly_budget) * Decimal("100")).quantize(TWO_DECIMALS)) if monthly_budget > 0 else None,
            "effective_budget": str(effective_budget.quantize(TWO_DECIMALS)),
            "effective_executed": str(effective_executed.quantize(TWO_DECIMALS)),
            "effective_remaining": str(effective_remaining),
            "rules": ["legacy_monthly_if_defined", "legacy_annual_fallback", "legacy_total_fallback", "0_is_ok"],
        }

    return {
        "has_budget": False,
        "source": None,
        "budget_validation_applies": True,
        "zero_is_allowed": True,
        "can_post_if_exact_zero": True,
        "rules": ["no_budget_found"],
    }


def parse_breakdown_payload(raw_value: Any, field_name: str) -> Dict[str, Any]:
    if raw_value is None:
        return {}
    parsed = raw_value
    if isinstance(raw_value, str):
        txt = raw_value.strip()
        if txt == "":
            return {}
        try:
            parsed = json.loads(txt)
        except json.JSONDecodeError:
            raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_json", "message": f"{field_name} debe ser JSON válido"})
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_type", "message": f"{field_name} debe ser un objeto JSON"})
    return parsed


def parse_optional_budget_decimal(value: Optional[Any], field_name: str) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, str) and value.strip() == "":
        return Decimal("0")
    return decimal_from_value(value, field_name)


def normalize_budget_breakdown(payload: BudgetPlanInput):
    return normalize_budget_breakdown_values(payload.total, payload.annual_breakdown, payload.monthly_breakdown)


def normalize_budget_breakdown_values(total_amount: Optional[Any], annual_breakdown: Optional[Any], monthly_breakdown: Optional[Any]):
    total = parse_optional_budget_decimal(total_amount, "total_amount")
    ensure_non_negative(total, "total_amount")

    annual_raw = parse_breakdown_payload(annual_breakdown, "annual_breakdown")
    monthly_raw = parse_breakdown_payload(monthly_breakdown, "monthly_breakdown")

    annual: Dict[str, Decimal] = {}
    monthly: Dict[str, Decimal] = {}

    for y, amount in annual_raw.items():
        year_str = str(y)
        if not re.fullmatch(r"\d{4}", year_str):
            raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_key", "message": f"Clave inválida en annual_breakdown: {year_str}", "meta": {"field": "annual_breakdown", "key": year_str}})
        year_int = int(year_str)
        validate_year_in_range(year_int)
        try:
            dec = decimal_from_value(amount, f"annual_breakdown.{year_str}")
        except HTTPException:
            raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_value", "message": f"Monto inválido en annual_breakdown para {year_str}", "meta": {"field": "annual_breakdown", "key": year_str}})
        ensure_non_negative(dec, f"annual_breakdown.{year_str}")
        annual[year_str] = dec

    for ym, amount in monthly_raw.items():
        ym_str = str(ym)
        if not re.fullmatch(r"\d{4}-\d{2}", ym_str):
            raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_key", "message": f"Clave inválida en monthly_breakdown: {ym_str}", "meta": {"field": "monthly_breakdown", "key": ym_str}})
        year_str = ym_str[:4]
        month = int(ym_str[5:7])
        if month < 1 or month > 12:
            raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_key", "message": f"Mes inválido en monthly_breakdown: {ym_str}", "meta": {"field": "monthly_breakdown", "key": ym_str}})
        validate_year_in_range(int(year_str))
        try:
            dec = decimal_from_value(amount, f"monthly_breakdown.{ym_str}")
        except HTTPException:
            raise HTTPException(status_code=422, detail={"code": "invalid_breakdown_value", "message": f"Monto inválido en monthly_breakdown para {ym_str}", "meta": {"field": "monthly_breakdown", "key": ym_str}})
        ensure_non_negative(dec, f"monthly_breakdown.{ym_str}")
        monthly[ym_str] = dec

    annual_sum = sum(annual.values(), Decimal("0"))
    monthly_sum = sum(monthly.values(), Decimal("0"))

    if annual_sum > total:
        raise HTTPException(status_code=422, detail={"code": "annual_sum_exceeds_total", "message": "Annual sum exceeds total amount.", "meta": {"total": str(total), "annual_sum": str(annual_sum)}})

    if monthly_sum > total:
        raise HTTPException(status_code=422, detail={"code": "monthly_sum_exceeds_total", "message": "Monthly sum exceeds total amount.", "meta": {"total": str(total), "monthly_sum": str(monthly_sum)}})

    monthly_by_year: Dict[str, Decimal] = {}
    for ym, amount in monthly.items():
        y = ym[:4]
        monthly_by_year[y] = monthly_by_year.get(y, Decimal("0")) + amount

    for y, monthly_total in monthly_by_year.items():
        annual_total = annual.get(y)
        if annual_total is not None and monthly_total > annual_total:
            raise HTTPException(status_code=422, detail={"code": "monthly_sum_exceeds_annual", "message": f"Monthly sum for year {y} exceeds annual allocation.", "meta": {"year": int(y), "annual": str(annual_total), "monthly_sum": str(monthly_total)}})

    return total, annual, monthly


def compute_inventory_totals(payload: InventoryItemBase):
    m2_superficie = decimal_from_value(payload.m2_superficie, "m2_superficie")
    m2_construccion = decimal_from_value(payload.m2_construccion or 0, "m2_construccion")
    precio_m2_superficie = decimal_from_value(payload.precio_m2_superficie, "precio_m2_superficie")
    precio_m2_construccion = decimal_from_value(payload.precio_m2_construccion or 0, "precio_m2_construccion")
    descuento = decimal_from_value(payload.descuento_bonificacion or 0, "descuento_bonificacion")

    for name, val in [
        ("m2_superficie", m2_superficie),
        ("m2_construccion", m2_construccion),
        ("precio_m2_superficie", precio_m2_superficie),
        ("precio_m2_construccion", precio_m2_construccion),
        ("descuento_bonificacion", descuento),
    ]:
        ensure_non_negative(val, name)

    precio_venta = (m2_superficie * precio_m2_superficie) + (m2_construccion * precio_m2_construccion)
    precio_total = precio_venta - descuento
    return precio_venta, precio_total
def normalize_customer_name(name: Optional[str]) -> Optional[str]:
    if name is None:
        return None
    normalized = name.strip()
    return normalized or None


def to_optional_str(value: Optional[object]) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    normalized = str(value).strip()
    return normalized or None


ABONO_PARTIDAS = {"402", "403"}

def movement_counts_as_abono_doc(movement: dict) -> bool:
    if not movement:
        return False
    if movement.get("is_deleted") is True:
        return False
    if str(movement.get("partida_codigo")) not in ABONO_PARTIDAS:
        return False
    if movement.get("status") != MovementStatus.POSTED.value:
        return False
    return bool(movement.get("client_id"))


def get_inventory_clave(item: Optional[dict]) -> Optional[str]:
    if not item:
        return None
    lote = (item.get("lote_edificio") or "").strip()
    manzana = (item.get("manzana_departamento") or "").strip()
    if lote and manzana:
        return f"{lote}-{manzana}"
    return lote or manzana or None

def resolve_inventory_reference(item: Optional[dict]) -> Optional[str]:
    if not item:
        return None

    lote_edificio = to_optional_str(item.get("lote_edificio"))
    if lote_edificio and "-" in lote_edificio:
        return lote_edificio

    manzana_departamento = to_optional_str(item.get("manzana_departamento"))
    if lote_edificio and manzana_departamento:
        return f"{lote_edificio}-{manzana_departamento}"

    lote = to_optional_str(item.get("lote"))
    manzana = to_optional_str(item.get("manzana") or item.get("mz") or item.get("manzana_edificio") or item.get("edificio"))
    if lote and manzana:
        return f"{lote}-{manzana}"

    for key in ("code", "inventory_code", "ref", "reference"):
        value = to_optional_str(item.get(key))
        if value:
            return value

    return to_optional_str(item.get("id") or item.get("_id"))




async def get_client_abono_movements(client_doc: dict, exclude_movement_id: Optional[str] = None):
    if not client_doc or not client_doc.get("id"):
        return []

    fallback_name = normalize_customer_name(client_doc.get("nombre"))
    query = movement_active_query(extra={
        "status": MovementStatus.POSTED.value,
        "partida_codigo": {"$in": list(ABONO_PARTIDAS)},
    })
    candidate_movements = await db.movements.find(query, {"_id": 0}).to_list(5000)

    matched_movements = []
    for mov in candidate_movements:
        if exclude_movement_id and mov.get("id") == exclude_movement_id:
            continue
        if mov.get("client_id") == client_doc.get("id"):
            matched_movements.append(mov)
            continue
        if mov.get("client_id"):
            continue
        mov_name = normalize_customer_name(mov.get("customer_name"))
        if not mov_name or mov_name != fallback_name:
            continue
        if client_doc.get("project_id") and mov.get("project_id") != client_doc.get("project_id"):
            continue
        if client_doc.get("company_id"):
            project = await db.projects.find_one({"id": mov.get("project_id")}, {"_id": 0})
            if not project or project.get("empresa_id") != client_doc.get("company_id"):
                continue
        matched_movements.append(mov)
    return matched_movements


async def recalc_client_financials(client_id: str):
    client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client_doc:
        return None

    matched_movements = await get_client_abono_movements(client_doc)
    abonos_total = sum(decimal_from_value(m.get("amount_mxn", 0), "amount_mxn") for m in matched_movements)
    valor_total_raw = client_doc.get("precio_venta_snapshot")
    if valor_total_raw in (None, "", 0, 0.0):
        valor_total_raw = client_doc.get("saldo_restante", 0)
    valor_total = decimal_from_value(valor_total_raw, "precio_venta_snapshot")
    saldo = valor_total - abonos_total
    if saldo < Decimal("0"):
        saldo = Decimal("0")
    update = {
        "abonos_total_mxn": float(abonos_total),
        "saldo_restante": float(saldo),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.clients.update_one({"id": client_id}, {"$set": update})
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return updated


async def validate_client_abono_limit(client_id: str, delta_amount_mxn: Decimal, exclude_movement_id: Optional[str] = None):
    client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=422, detail={"code": "client_not_found", "message": "Cliente no válido"})
    valor_total_raw = client_doc.get("precio_venta_snapshot")
    if valor_total_raw in (None, "", 0, 0.0):
        valor_total_raw = client_doc.get("saldo_restante", 0)
    valor_total = decimal_from_value(valor_total_raw, "precio_venta_snapshot")
    movements = await get_client_abono_movements(client_doc, exclude_movement_id=exclude_movement_id)
    abonos_total = Decimal("0")
    for mov in movements:
        abonos_total += decimal_from_value(mov.get("amount_mxn", 0), "amount_mxn")
    projected = abonos_total + delta_amount_mxn
    if projected > valor_total:
        raise HTTPException(status_code=422, detail={"code": "payment_exceeds_balance", "message": "El abono excede el saldo restante"})


def render_basic_pdf(lines: List[str]) -> bytes:
    safe_lines = []
    for line in lines:
        safe = str(line).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        safe_lines.append(safe)
    content_lines = ["BT", "/F1 11 Tf", "50 780 Td"]
    first = True
    for line in safe_lines:
        if first:
            content_lines.append(f"({line}) Tj")
            first = False
        else:
            content_lines.append("0 -16 Td")
            content_lines.append(f"({line}) Tj")
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", errors="replace")

    objects = []
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")
    objects.append(b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n")
    objects.append(b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n")
    objects.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n")
    objects.append(f"5 0 obj << /Length {len(stream)} >> stream\n".encode("latin-1") + stream + b"\nendstream endobj\n")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_start = len(pdf)
    pdf.extend(f"xref\n0 {len(offsets)}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        pdf.extend(f"{off:010d} 00000 n \n".encode("latin-1"))
    pdf.extend(f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF".encode("latin-1"))
    return bytes(pdf)


def _pdf_escape(text: Any) -> str:
    return str(text or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_wrap(text: Any, max_chars: int) -> List[str]:
    raw = str(text or "").strip() or "N/A"
    words = raw.split()
    if not words:
        return ["N/A"]
    lines: List[str] = []
    line = ""
    for word in words:
        candidate = f"{line} {word}".strip()
        if len(candidate) <= max_chars:
            line = candidate
            continue
        if line:
            lines.append(line)
        while len(word) > max_chars:
            lines.append(word[:max_chars])
            word = word[max_chars:]
        line = word
    if line:
        lines.append(line)
    return lines or ["N/A"]


def _pdf_format_date_local(date_value: Any, fmt: str = "%d/%m/%Y") -> str:
    try:
        if isinstance(date_value, datetime):
            dt = date_value
        elif isinstance(date_value, str) and date_value.strip():
            dt = date_parser.parse(date_value)
        else:
            return "N/A"
        if dt.tzinfo is None:
            dt = TIMEZONE.localize(dt)
        else:
            dt = dt.astimezone(TIMEZONE)
        return dt.strftime(fmt)
    except Exception:
        return "N/A"


def _oc_numeric_display(folio: str) -> str:
    normalized = canonicalize_oc_folio(folio)
    digits = ''.join(ch for ch in normalized if ch.isdigit())
    if not digits:
        return normalized
    return f"{int(digits):04d}"


def format_money(amount: Any, currency: str = "MXN") -> str:
    if amount is None:
        return "—"
    try:
        value = money_dec(amount).quantize(TWO_DECIMALS)
    except Exception:
        return "—"
    code = str(currency or "MXN").upper()
    symbol = "US$" if code == "USD" else "$"
    return f"{symbol}{value:,.2f}"


def format_number(amount: Any) -> str:
    if amount is None:
        return "—"
    try:
        value = money_dec(amount).quantize(TWO_DECIMALS)
    except Exception:
        return "—"
    return f"{value:,.2f}"


def oc_pdf_filename(raw_folio: Optional[str]) -> str:
    base = canonicalize_oc_folio(raw_folio)
    if not base or base == "N/A":
        return "purchase-order.pdf"
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", base)
    safe = re.sub(r"_+", "_", safe).strip("._-")
    if not safe:
        return "purchase-order.pdf"
    return f"{safe}.pdf"


def derive_delivery_status(po: dict) -> Optional[str]:
    for key in ("sent_to_vendor", "submitted_to_vendor"):
        if po.get(key) is True:
            return "Enviado"
    for key in ("sent_at", "submitted_at", "vendor_sent_at"):
        if po.get(key):
            return "Enviado"
    return None


def derive_payment_status(po: dict) -> Optional[str]:
    if po.get("odoo_payment_id") or po.get("payment_odoo_id"):
        return "Enviado a Odoo"
    if po.get("paid_at"):
        return "Pagado"
    if po.get("payment_approved_at"):
        return "Aprobado"
    if po.get("approved_at") or po.get("status") == PurchaseOrderStatus.APPROVED_FOR_PAYMENT.value:
        return "Aprobado"
    return None


def resolve_pdf_logo_path() -> Optional[str]:
    env_logo = os.environ.get("QFINANCE_PDF_LOGO_PATH", "").strip()
    candidates: List[Path] = []
    if env_logo:
        candidates.append(Path(env_logo).expanduser())
    candidates.append(Path("/var/www/qfinance/brand/quantum_logo.png"))
    candidates.append(ROOT_DIR.parent / "frontend" / "public" / "brand" / "quantum_logo.png")

    for candidate in candidates:
        try:
            resolved = candidate.resolve(strict=True)
        except Exception:
            continue
        if resolved.is_file() and os.access(str(resolved), os.R_OK):
            return str(resolved)
    return None


def build_purchase_order_pdf_payload(po: dict) -> dict:
    folio = canonicalize_oc_folio(po.get("folio") or po.get("external_id"))
    order_date = _pdf_format_date_local(po.get("order_date"), "%Y-%m-%d")
    planned_date = _pdf_format_date_local(po.get("planned_date"), "%Y-%m-%d") if po.get("planned_date") else "N/A"

    buyer = {
        "name": po.get("company_name") or po.get("empresa_nombre") or po.get("company_id") or "S/I",
        "rfc": po.get("company_rfc") or "S/I",
        "address": po.get("company_address") or "S/I",
        "contact": po.get("company_contact") or "S/I",
    }
    vendor = {
        "name": po.get("vendor_name") or po.get("proveedor_nombre") or "S/I",
        "rfc": po.get("vendor_rfc") or po.get("proveedor_rfc") or "S/I",
        "address": po.get("vendor_address") or "S/I",
        "contact": po.get("vendor_email") or po.get("vendor_phone") or "S/I",
    }

    lines_payload = []
    for idx, line in enumerate(po.get("lines") or [], start=1):
        unit_value = (
            line.get("unit")
            or line.get("unidad")
            or line.get("uom")
            or line.get("uom_code")
            or line.get("unit_code")
            or "—"
        )
        lines_payload.append({
            "line_no": line.get("line_no") or idx,
            "code": line.get("partida_codigo") or "SERV",
            "description": line.get("description") or "S/I",
            "uom": unit_value,
            "qty": line.get("qty") or "0",
            "price_unit": line.get("price_unit") or "0",
            "amount": line.get("line_total") or "0",
            "iva_amount": line.get("iva_amount") or "0",
            "ret_isr": line.get("isr_withholding_amount") or "0",
        })

    subtotal = money_dec(po.get("subtotal_tax_base", 0)).quantize(TWO_DECIMALS)
    tax = money_dec(po.get("tax_total", 0)).quantize(TWO_DECIMALS)
    total = money_dec(po.get("total", 0)).quantize(TWO_DECIMALS)
    exchange_rate = money_dec(po.get("exchange_rate", 1)).quantize(TWO_DECIMALS)
    total_mxn = (total * exchange_rate).quantize(TWO_DECIMALS)
    iva_withholding_total = money_dec(po.get("iva_withholding_total", 0)).quantize(TWO_DECIMALS)
    generated_at = datetime.now(TIMEZONE)
    generated_label = generated_at.strftime("%d/%m/%Y, %I:%M %p").lower().replace("am", "a.m.").replace("pm", "p.m.")

    payload = {
        "folio": folio,
        "date": order_date,
        "planned_date": planned_date,
        "buyer": buyer,
        "vendor": vendor,
        "project": po.get("project_name") or po.get("proyecto_nombre") or po.get("project_id") or "S/I",
        "currency": po.get("currency") or "MXN",
        "exchange_rate": str(exchange_rate),
        "lines": lines_payload,
        "subtotal": str(subtotal),
        "tax": str(tax),
        "total": str(total),
        "total_mxn": str(total_mxn),
        "iva_withholding_total": str(iva_withholding_total),
        "bank": po.get("bank_details") or None,
        "notes": po.get("notes") or "S/I",
        "payment_terms": po.get("payment_terms") or "S/I",
        "status": po.get("status") or "draft",
        "approved_amount_total": str(money_dec(po.get("approved_amount_total", 0)).quantize(TWO_DECIMALS)),
        "pending_amount": str(money_dec(po.get("pending_amount", 0)).quantize(TWO_DECIMALS)),
        "generated_at": generated_label,
        "branding": {"company": "Quantum", "site": "quantumgrupo.mx"},
    }
    return payload


def render_purchase_order_pdf(po: dict) -> bytes:
    payload = build_purchase_order_pdf_payload(po)
    folio = payload["folio"]
    order_date = payload["date"]
    planned_date = payload["planned_date"]
    oc_number = _oc_numeric_display(folio)

    company_name = payload["buyer"]["name"]
    project_name = payload.get("project")
    vendor_name = payload["vendor"]["name"]
    vendor_rfc = payload["vendor"]["rfc"]

    delivery_status = derive_delivery_status(po)
    payment_status = derive_payment_status(po)

    lines = payload.get("lines") or []
    notes_text = payload.get("notes") or "S/I"
    bank = payload.get("bank") or {}

    logo_path = resolve_pdf_logo_path()
    logo_image: Optional[Dict[str, Any]] = None
    if logo_path:
        try:
            with Image.open(logo_path) as img:
                img_rgb = img.convert("RGB")
                width_px, height_px = img_rgb.size
                logo_image = {
                    "width_px": width_px,
                    "height_px": height_px,
                    "data": img_rgb.tobytes(),
                }
        except Exception:
            logo_image = None

    page_streams: List[bytes] = []

    def text_cmd(cmds: List[str], x: float, y: float, text: Any, size: int = 9):
        cmds += ["BT", f"/F1 {size} Tf", f"{x:.2f} {y:.2f} Td", f"({_pdf_escape(text)}) Tj", "ET"]

    def draw_wrapped(cmds: List[str], x: float, y: float, text: Any, size: int, width_chars: int, line_step: float = 12.0) -> float:
        lines_local = _pdf_wrap(text, width_chars)
        cursor = y
        for line in lines_local:
            text_cmd(cmds, x, cursor, line, size)
            cursor -= line_step
        return cursor

    def draw_header(cmds: List[str]):
        header_x = 40
        header_y = 748
        header_w = 532
        header_h = 92

        col1_w = 130
        col3_w = 180
        col2_w = header_w - col1_w - col3_w

        col1_x = header_x
        col2_x = col1_x + col1_w
        col3_x = col2_x + col2_w

        cmds += ["0.94 0.95 0.98 rg", f"{header_x} {header_y-header_h:.2f} {header_w} {header_h} re f", "0 0 0 rg"]
        cmds += ["0.82 0.85 0.92 RG", "0.5 w", f"{header_x} {header_y-header_h:.2f} {header_w} {header_h} re S"]
        cmds += ["0.87 0.89 0.94 RG", "0.4 w", f"{col2_x} {header_y-header_h:.2f} 0 {header_h} m {col2_x} {header_y:.2f} l S"]
        cmds += ["0.87 0.89 0.94 RG", "0.4 w", f"{col3_x} {header_y-header_h:.2f} 0 {header_h} m {col3_x} {header_y:.2f} l S", "0 0 0 rg"]

        # Left mini-table layout: [logo][Quantum | QFinance]
        left_cell_logo_x = col1_x + 8
        left_cell_logo_w = 46.0
        left_cell_text_x = left_cell_logo_x + left_cell_logo_w + 4
        logo_max_h = 30.0
        logo_max_w = left_cell_logo_w
        logo_drawn = False
        if logo_image:
            ratio = logo_image["width_px"] / max(1, logo_image["height_px"])
            logo_h = logo_max_h
            logo_w = logo_h * ratio
            if logo_w > logo_max_w:
                logo_w = logo_max_w
                logo_h = logo_w / max(ratio, 0.001)
            logo_x = left_cell_logo_x + ((left_cell_logo_w - logo_w) / 2)
            logo_y = header_y - 10 - logo_h
            cmds += ["1 1 1 rg", f"{logo_x-2:.2f} {logo_y-2:.2f} {logo_w+4:.2f} {logo_h+4:.2f} re f", "0 0 0 rg"]
            cmds += ["q", f"{logo_w:.2f} 0 0 {logo_h:.2f} {logo_x:.2f} {logo_y:.2f} cm", "/Im1 Do", "Q"]
            logo_drawn = True

        if not logo_drawn:
            text_cmd(cmds, col1_x + 8, header_y - 28, "Quantum", 13)
        text_cmd(cmds, left_cell_text_x, header_y - 30, "Quantum | QFinance", 8)

        text_cmd(cmds, col2_x + 10, header_y - 28, "ORDEN DE COMPRA", 20)
        text_cmd(cmds, col2_x + 10, header_y - 44, "DOCUMENTO COMERCIAL", 10)

        # right column content (wrapped in fixed width)
        y = header_y - 18
        text_cmd(cmds, col3_x + 8, y, f"Folio: {folio}", 10)
        y -= 14
        text_cmd(cmds, col3_x + 8, y, f"Fecha: {order_date}", 9)

        chips = []
        if delivery_status:
            chips.append(("Envío", delivery_status))
        if payment_status:
            chips.append(("Pago", payment_status))
        if chips:
            chip_y = header_y - header_h + 10
            chip_h = 14
            chip_w = (col3_w - 24) / max(1, len(chips))
            for idx_chip, (label, value) in enumerate(chips):
                cx = col3_x + 8 + idx_chip * chip_w
                cmds += ["0.92 0.92 0.94 rg", f"{cx:.2f} {chip_y:.2f} {chip_w-4:.2f} {chip_h} re f", "0 0 0 rg"]
                text_cmd(cmds, cx + 3, chip_y + 4, f"{label}: {value}"[:30], 7)

    def draw_party_cards(cmds: List[str]) -> float:
        left_x, right_x = 50, 315
        top_y = 644
        card_h = 86
        cmds += ["0.95 0.96 0.99 rg", f"{left_x} {top_y-card_h} 250 {card_h} re f", "0 0 0 rg"]
        cmds += ["0.95 0.96 0.99 rg", f"{right_x} {top_y-card_h} 257 {card_h} re f", "0 0 0 rg"]
        text_cmd(cmds, left_x+8, top_y-14, "COMPRADOR", 10)
        text_cmd(cmds, right_x+8, top_y-14, "VENDEDOR", 10)

        buyer = [
            f"Empresa: {company_name}",
            f"Proyecto: {project_name}",
            f"Moneda: {po.get('currency') or 'MXN'} | TC: {po.get('exchange_rate') or '1'}",
            f"Condiciones: {po.get('payment_terms') or 'S/I'}",
        ]
        vendor = [
            f"Proveedor: {vendor_name}",
            f"RFC: {vendor_rfc}",
            f"Factura proveedor: {po.get('invoice_folio') or 'S/I'}",
            f"Contacto: {po.get('vendor_email') or 'S/I'}",
        ]
        y = top_y - 30
        for t in buyer:
            text_cmd(cmds, left_x+8, y, t, 8)
            y -= 15
        y = top_y - 30
        for t in vendor:
            y = draw_wrapped(cmds, right_x+8, y, t, 8, 40, line_step=10) - 4
        return top_y - card_h - 10

    TABLE_COLS = {
        "num": (50, 72),
        "partida": (72, 108),
        "desc": (108, 236),
        "qty": (236, 274),
        "uom": (274, 304),
        "pu": (304, 378),
        "iva": (378, 446),
        "ret_isr": (446, 504),
        "total": (504, 572),
    }

    MONEY_CELL_PAD_LEFT = 5
    MONEY_CELL_PAD_RIGHT = 4
    MONEY_ROW_MIN_FONT = 7.6

    def format_table_money(amount: Any) -> str:
        if amount is None:
            return "$0.00"
        try:
            value = money_dec(amount).quantize(TWO_DECIMALS)
        except Exception:
            return "$0.00"
        return f"${value:,.2f}"

    def draw_table_header(cmds: List[str], y: float):
        cmds += ["0.05 0.18 0.56 rg", f"50 {y-16:.2f} 522 16 re f", "1 1 1 rg"]
        headers = [
            (TABLE_COLS["num"][0] + 4, "#"),
            (TABLE_COLS["partida"][0] + 4, "Partida"),
            (TABLE_COLS["desc"][0] + 4, "Descripción"),
            (TABLE_COLS["qty"][0] + 4, "Cant."),
            (TABLE_COLS["uom"][0] + 4, "Unidad"),
            (TABLE_COLS["pu"][0] + 2, "P. Unitario"),
            (TABLE_COLS["iva"][0] + 2, "IVA"),
            (TABLE_COLS["ret_isr"][0] + 2, "Ret ISR"),
            (TABLE_COLS["total"][0] + 2, "Total"),
        ]
        for x, h in headers:
            text_cmd(cmds, x, y-12, h, 8.5)
        cmds += ["0 0 0 rg"]

    def text_cmd_right(cmds: List[str], right_x: float, left_x: float, y: float, text: Any, size: float = 8.5, min_size: float = 7.8):
        raw = str(text or "")
        max_w = max(6.0, right_x - left_x)

        start_tenths = int(round(size * 10))
        min_tenths = int(round(min_size * 10))
        candidate_sizes = [t / 10 for t in range(start_tenths, min_tenths - 1, -2)]
        if min_size not in candidate_sizes:
            candidate_sizes.append(min_size)

        rendered = False
        for candidate in candidate_sizes:
            approx_w = len(raw) * (candidate * 0.54)
            if approx_w <= max_w:
                x = max(left_x, right_x - approx_w)
                cmds += ["BT", f"/F2 {candidate:.2f} Tf", f"{x:.2f} {y:.2f} Td", f"({_pdf_escape(raw)}) Tj", "ET"]
                rendered = True
                break

        if not rendered:
            approx_w = len(raw) * (min_size * 0.54)
            x = max(left_x, right_x - approx_w)
            cmds += ["BT", f"/F2 {min_size:.2f} Tf", f"{x:.2f} {y:.2f} Td", f"({_pdf_escape(raw)}) Tj", "ET"]

    def measure_cell_height(text: Any, width_chars: int, max_lines: Optional[int] = None, line_step: float = 10.0, min_height: float = 13.0) -> float:
        lines_local = _pdf_wrap(text, width_chars)
        if max_lines is not None:
            lines_local = lines_local[:max_lines]
        return max(min_height, len(lines_local) * line_step)

    def compute_row_height(line: dict) -> float:
        return max(
            measure_cell_height(line.get("description") or "-", 30, max_lines=2),
            measure_cell_height(line.get("uom") or "—", 6, max_lines=2),
            13.0,
        )

    def ensure_page_space(y_cursor: float, row_height: float, min_bottom: float = 170.0) -> bool:
        return (y_cursor - row_height) >= min_bottom

    def draw_row(cmds: List[str], line: dict, row_index: int, y_cursor: float, row_height: float, shaded: bool) -> None:
        if shaded:
            cmds += ["0.98 0.98 0.99 rg", f"50 {y_cursor-row_height+2:.2f} 522 {row_height:.2f} re f", "0 0 0 rg"]
        base_y = y_cursor - 9
        desc_lines = _pdf_wrap(line.get("description") or "-", 30)[:2]
        uom_lines = _pdf_wrap(line.get("uom") or "—", 6)[:2]
        text_cmd(cmds, TABLE_COLS["num"][0] + 4, base_y, line.get("line_no") or row_index + 1, 8.5)
        text_cmd(cmds, TABLE_COLS["partida"][0] + 2, base_y, line.get("code") or "", 8.5)
        for idx_desc, d in enumerate(desc_lines):
            text_cmd(cmds, TABLE_COLS["desc"][0] + 2, base_y - (idx_desc * 10), d, 8.0)
        text_cmd_right(cmds, TABLE_COLS["qty"][1] - 2, TABLE_COLS["qty"][0] + 2, base_y, str(line.get("qty") or "0"), 8.5)
        for idx_uom, uom in enumerate(uom_lines):
            text_cmd(cmds, TABLE_COLS["uom"][0] + 2, base_y - (idx_uom * 10), uom, 8.5)
        text_cmd_right(cmds, TABLE_COLS["pu"][1] - MONEY_CELL_PAD_RIGHT, TABLE_COLS["pu"][0] + MONEY_CELL_PAD_LEFT, base_y, format_table_money(line.get("price_unit")), 8.5, MONEY_ROW_MIN_FONT)
        text_cmd_right(cmds, TABLE_COLS["iva"][1] - MONEY_CELL_PAD_RIGHT, TABLE_COLS["iva"][0] + MONEY_CELL_PAD_LEFT, base_y, format_table_money(line.get("iva_amount")), 8.5, MONEY_ROW_MIN_FONT)
        text_cmd_right(cmds, TABLE_COLS["ret_isr"][1] - MONEY_CELL_PAD_RIGHT, TABLE_COLS["ret_isr"][0] + MONEY_CELL_PAD_LEFT, base_y, format_table_money(line.get("ret_isr")), 8.5, MONEY_ROW_MIN_FONT)
        text_cmd_right(cmds, TABLE_COLS["total"][1] - MONEY_CELL_PAD_RIGHT, TABLE_COLS["total"][0] + MONEY_CELL_PAD_LEFT, base_y, format_table_money(line.get("amount")), 8.5, MONEY_ROW_MIN_FONT)

    row_y_start_first = 518
    row_y_start_other = 706
    page_no = 1
    i = 0
    while i < len(lines) or page_no == 1:
        cmds: List[str] = []
        draw_header(cmds)
        if page_no == 1:
            y_table_top = draw_party_cards(cmds)
            y_cursor = min(y_table_top, row_y_start_first)
        else:
            y_cursor = row_y_start_other

        draw_table_header(cmds, y_cursor)
        y_cursor -= 20

        shaded = False
        while i < len(lines):
            line = lines[i] or {}
            row_height = compute_row_height(line)
            if not ensure_page_space(y_cursor, row_height):
                break
            draw_row(cmds, line, i, y_cursor, row_height, shaded)
            shaded = not shaded

            y_cursor -= row_height
            i += 1

        if i >= len(lines):
            box_y = max(120, y_cursor - 150)
            cmds += ["0.95 0.95 0.97 rg", f"340 {box_y:.2f} 232 130 re f", "0 0 0 rg"]
            currency_code = str(payload.get('currency') or 'MXN').upper()
            totals = [
                f"Subtotal ({currency_code}): {format_money(payload.get('subtotal'), currency_code)}",
                f"IVA ({currency_code}): {format_money(payload.get('tax'), currency_code)}",
                f"Ret IVA ({po.get('iva_withholding_rate') or '0'}%): {format_money(payload.get('iva_withholding_total'), currency_code)}",
                f"Ret ISR ({currency_code}): {format_money(po.get('withholding_isr_total'), currency_code)}",
                f"TOTAL ({currency_code}): {format_money(payload.get('total'), currency_code)}",
                f"TC: {format_number(payload.get('exchange_rate'))}",
                f"TOTAL (MXN): {format_money(payload.get('total_mxn'), 'MXN')}",
            ]
            ty = box_y + 112
            for t in totals:
                text_cmd(cmds, 350, ty, t, 10)
                ty -= 16

            notes_title_y = box_y + 26
            text_cmd(cmds, 50, notes_title_y, "NOTAS / COMENTARIOS ADICIONALES", 9)

            notes_lines = [] if notes_text == "S/I" else _pdf_wrap(notes_text, 95)
            bank_lines = []
            if bank and isinstance(bank, dict):
                mapping = [("banco", "Banco"), ("cuenta", "Cuenta"), ("clabe", "CLABE"), ("beneficiario", "Beneficiario")]
                for key, label in mapping:
                    value = str(bank.get(key) or "").strip()
                    if value:
                        bank_lines.append(f"• {label}: {value}")
            combined = notes_lines + bank_lines
            if not combined:
                combined = ["S/I"]

            notes_y = notes_title_y - 12
            for n in combined[:8]:
                text_cmd(cmds, 50, notes_y, n, 8)
                notes_y -= 11

        generated_human = f"Generado: {payload.get('generated_at')}"
        cmds += ["0.80 0.80 0.80 rg", "50 46 522 1 re f", "0 0 0 rg"]
        text_cmd(cmds, 50, 34, generated_human, 8)
        text_cmd(cmds, 250, 34, "Creado con QFinance", 8)
        text_cmd(cmds, 470, 34, "quantumgrupo.mx", 8)
        text_cmd(cmds, 548, 34, f"{page_no}", 8)
        page_no += 1
        page_streams.append("\n".join(cmds).encode("latin-1", errors="replace"))

    objects: List[bytes] = []
    info_obj_num = 3
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")

    page_count = len(page_streams)
    page_obj_start = 4
    content_obj_start = page_obj_start + page_count
    font_obj_num = content_obj_start + page_count
    mono_font_obj_num = font_obj_num + 1
    next_obj_num = mono_font_obj_num + 1

    logo_obj_num: Optional[int] = None
    logo_object: Optional[bytes] = None
    if logo_image:
        logo_obj_num = next_obj_num
        next_obj_num += 1
        img_data = logo_image["data"]
        compressed = zlib.compress(img_data)
        logo_object = (
            f"{logo_obj_num} 0 obj << /Type /XObject /Subtype /Image /Width {logo_image['width_px']} /Height {logo_image['height_px']} /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length {len(compressed)} >> stream\n".encode("latin-1")
            + compressed
            + b"\nendstream endobj\n"
        )

    kids = " ".join([f"{page_obj_start + idx} 0 R" for idx in range(page_count)])
    objects.append(f"2 0 obj << /Type /Pages /Kids [{kids}] /Count {page_count} >> endobj\n".encode("latin-1"))

    for idx in range(page_count):
        page_obj_num = page_obj_start + idx
        content_obj_num = content_obj_start + idx
        xobj_part = f" /XObject << /Im1 {logo_obj_num} 0 R >>" if logo_obj_num else ""
        objects.append(
            f"{page_obj_num} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 {font_obj_num} 0 R /F2 {mono_font_obj_num} 0 R >>{xobj_part} >> /Contents {content_obj_num} 0 R >> endobj\n".encode("latin-1")
        )

    for idx, stream in enumerate(page_streams):
        content_obj_num = content_obj_start + idx
        objects.append(
            f"{content_obj_num} 0 obj << /Length {len(stream)} >> stream\n".encode("latin-1") + stream + b"\nendstream endobj\n"
        )

    objects.append(f"{font_obj_num} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n".encode("latin-1"))
    objects.append(f"{mono_font_obj_num} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Courier >> endobj\n".encode("latin-1"))
    if logo_object:
        objects.append(logo_object)
    objects.insert(2, f"3 0 obj << /Title (Orden de Compra {folio}) /Creator (QFinance / quantumgrupo.mx) /Producer (QFinance / quantumgrupo.mx) >> endobj\n".encode("latin-1", errors="replace"))

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_start = len(pdf)
    pdf.extend(f"xref\n0 {len(offsets)}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        pdf.extend(f"{off:010d} 00000 n \n".encode("latin-1"))
    pdf.extend(f"trailer << /Size {len(offsets)} /Root 1 0 R /Info {info_obj_num} 0 R >>\nstartxref\n{xref_start}\n%%EOF".encode("latin-1"))
    return bytes(pdf)

# ========================= AUTH ROUTES =========================
@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate, current_user: dict = Depends(require_permission(Permission.MANAGE_USERS))):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=409, detail="Email ya registrado")
    existing_name = await db.users.find_one({"name": user_data.name}, {"_id": 0})
    if existing_name:
        raise HTTPException(status_code=409, detail="Nombre de usuario ya registrado")
    
    user = User(**user_data.model_dump(exclude={"password"}))
    doc = user.model_dump()
    doc['password_hash'] = hash_password(user_data.password)
    doc['must_change_password'] = False
    doc['created_at'] = doc['created_at'].isoformat()
    try:
        await db.users.insert_one(doc)
    except DuplicateKeyError as exc:
        msg = str(exc)
        if "email" in msg:
            raise HTTPException(status_code=409, detail="Email ya registrado")
        if "name" in msg or "username" in msg:
            raise HTTPException(status_code=409, detail="Nombre de usuario ya registrado")
        raise HTTPException(status_code=409, detail="Usuario duplicado")
    return user

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc:
        logger.info("Login failed: user not found for email=%s", credentials.email)
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    if not verify_password(credentials.password, user_doc.get('password_hash', '')):
        logger.info("Login failed: invalid password for email=%s", credentials.email)
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    if not user_doc.get('is_active', True):
        logger.info("Login failed: inactive user id=%s", user_doc.get('id'))
        raise HTTPException(status_code=401, detail="Usuario desactivado")
    
    must_change_password = bool(user_doc.get('must_change_password', False))
    empresa_ids = list(user_doc.get("empresa_ids") or ([] if not user_doc.get("empresa_id") else [user_doc.get("empresa_id")]))
    empresa_ids = [str(e) for e in empresa_ids if e]
    if is_operational_role(user_doc.get("role")) and not empresa_ids:
        raise HTTPException(status_code=422, detail={"code": "empresa_required", "message": "Asigna al menos una empresa al usuario en Consola Admin"})
    token = create_token(user_doc['id'], user_doc['email'], user_doc['role'], must_change_password=must_change_password, empresa_id=None, empresa_ids=empresa_ids)
    enriched_user = {k: v for k, v in user_doc.items() if k != 'password_hash'}
    enriched_user["empresa_ids"] = empresa_ids
    user = User(**enriched_user)
    
    # Log successful login
    await log_audit(
        {"user_id": user_doc['id'], "email": user_doc['email'], "role": user_doc['role']},
        "LOGIN",
        "auth",
        user_doc['id'],
        {"status": "success"}
    )
    
    return TokenResponse(access_token=token, user=user, must_change_password=must_change_password)

@api_router.post("/auth/logout")
async def logout(current_user: dict = Depends(get_current_user)):
    """Log user logout"""
    await log_audit(current_user, "LOGOUT", "auth", current_user["user_id"], {"status": "success"})
    return {"message": "Sesión cerrada"}

@api_router.post("/auth/change-password")
async def change_password(payload: ChangePasswordRequest, request: Request, current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if not verify_password(payload.current_password, user_doc.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="La contraseña actual es incorrecta")
    if payload.current_password == payload.new_password:
        raise HTTPException(status_code=422, detail="La nueva contraseña debe ser diferente a la actual")
    validate_password_policy(payload.new_password)

    await db.users.update_one({"id": user_doc["id"]}, {"$set": {
        "password_hash": hash_password(payload.new_password),
        "must_change_password": False,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    await log_audit(current_user, "PASSWORD_CHANGED", "users", user_doc["id"], {"source": "settings"}, ip_address=request.client.host if request.client else None)
    fresh = await db.users.find_one({"id": user_doc["id"]}, {"_id": 0})
    token = create_token(fresh["id"], fresh["email"], fresh["role"], must_change_password=False, empresa_id=None if is_operational_role(fresh.get("role")) else (fresh.get("empresa_id") or fresh.get("company_id")), empresa_ids=list(fresh.get("empresa_ids") or ([] if not fresh.get("empresa_id") else [fresh.get("empresa_id")])) )
    user = User(**{k: v for k, v in fresh.items() if k != "password_hash"})
    return TokenResponse(access_token=token, user=user, must_change_password=False)


@api_router.post("/auth/force-change-password")
async def force_change_password(payload: ForceChangePasswordRequest, request: Request, current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if not user_doc.get("must_change_password", False):
        raise HTTPException(status_code=403, detail="El usuario no requiere cambio forzado de contraseña")
    validate_password_policy(payload.new_password)
    if verify_password(payload.new_password, user_doc.get("password_hash", "")):
        raise HTTPException(status_code=422, detail="La nueva contraseña debe ser diferente a la actual")

    await db.users.update_one({"id": user_doc["id"]}, {"$set": {
        "password_hash": hash_password(payload.new_password),
        "must_change_password": False,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    await log_audit(current_user, "PASSWORD_CHANGED", "users", user_doc["id"], {"source": "force_change"}, ip_address=request.client.host if request.client else None)
    fresh = await db.users.find_one({"id": user_doc["id"]}, {"_id": 0})
    token = create_token(fresh["id"], fresh["email"], fresh["role"], must_change_password=False, empresa_id=None if is_operational_role(fresh.get("role")) else (fresh.get("empresa_id") or fresh.get("company_id")), empresa_ids=list(fresh.get("empresa_ids") or ([] if not fresh.get("empresa_id") else [fresh.get("empresa_id")])) )
    user = User(**{k: v for k, v in fresh.items() if k != "password_hash"})
    return TokenResponse(access_token=token, user=user, must_change_password=False)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password_hash": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    empresa_ids = list(user_doc.get("empresa_ids") or ([] if not user_doc.get("empresa_id") else [user_doc.get("empresa_id")]))
    user_doc["empresa_ids"] = [str(e) for e in empresa_ids if e]
    user_doc["empresa_id"] = current_user.get("empresa_id")
    return User(**user_doc)



@api_router.get("/auth/allowed-companies")
async def auth_allowed_companies(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") == UserRole.ADMIN.value:
        rows = await db.empresas.find(active_query(), {"_id": 0}).to_list(500)
        return [{"id": r.get("id"), "nombre": r.get("nombre")} for r in rows]
    empresa_ids = list(current_user.get("empresa_ids") or [])
    if not empresa_ids:
        return []
    rows = await db.empresas.find({"id": {"$in": empresa_ids}}, {"_id": 0}).to_list(500)
    return [{"id": r.get("id"), "nombre": r.get("nombre")} for r in rows]


class SelectCompanyRequest(BaseModel):
    empresa_id: str


@api_router.post("/auth/select-company")
async def auth_select_company(payload: SelectCompanyRequest, current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": current_user.get("user_id")}, {"_id": 0, "password_hash": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if current_user.get("role") == UserRole.ADMIN.value:
        token = create_token(user_doc["id"], user_doc["email"], user_doc["role"], must_change_password=False, empresa_id=payload.empresa_id, empresa_ids=list(user_doc.get("empresa_ids") or []))
        return {"access_token": token, "token_type": "bearer"}

    empresa_ids = list(user_doc.get("empresa_ids") or ([] if not user_doc.get("empresa_id") else [user_doc.get("empresa_id")]))
    empresa_ids = [str(e) for e in empresa_ids if e]
    if payload.empresa_id not in empresa_ids:
        raise HTTPException(status_code=403, detail={"code": "forbidden_company", "message": "La empresa seleccionada no está permitida para este usuario"})

    token = create_token(user_doc["id"], user_doc["email"], user_doc["role"], must_change_password=False, empresa_id=payload.empresa_id, empresa_ids=empresa_ids)
    return {"access_token": token, "token_type": "bearer"}

@api_router.get("/auth/permissions")
async def get_my_permissions(current_user: dict = Depends(get_current_user)):
    """Return current user's permissions for frontend RBAC"""
    role = current_user.get("role")
    permissions = ROLE_PERMISSIONS.get(role, [])
    return {
        "role": role,
        "permissions": permissions
    }

# ========================= USER ROUTES =========================
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: dict = Depends(require_permission(Permission.VIEW_USERS, Permission.MANAGE_USERS))):
    raise HTTPException(status_code=404, detail="La gestión de usuarios se movió a /api/admin/users")

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, updates: dict, current_user: dict = Depends(require_permission(Permission.MANAGE_USERS))):
    raise HTTPException(status_code=404, detail="La gestión de usuarios se movió a /api/admin/users")

# ========================= EMPRESA ROUTES =========================
@api_router.get("/empresas")
async def get_empresas(current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    empresas = await db.empresas.find({}, {"_id": 0}).to_list(1000)
    return empresas

@api_router.post("/empresas")
async def create_empresa(empresa_data: EmpresaBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    existing = await db.empresas.find_one({"nombre": empresa_data.nombre}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Empresa ya existe")
    
    empresa = Empresa(**empresa_data.model_dump())
    doc = empresa.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.empresas.insert_one(doc)
    await log_audit(current_user, "CREATE", "empresas", empresa.id, {"data": {"nombre": doc['nombre']}})
    doc.pop('_id', None)
    return doc

# ========================= CATALOGO PARTIDAS ROUTES =========================
@api_router.get("/catalogo-partidas")
async def get_catalogo_partidas(current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    partidas = await db.catalogo_partidas.find({}, {"_id": 0}).sort("codigo", 1).to_list(1000)
    return partidas

@api_router.get("/catalogo-partidas/{codigo}")
async def get_catalogo_partida(codigo: str, current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    partida = await db.catalogo_partidas.find_one({"codigo": codigo}, {"_id": 0})
    if not partida:
        raise HTTPException(status_code=404, detail=f"Partida {codigo} no encontrada en catálogo")
    return partida

# Helper: validar partida existe y está activa
async def validate_partida(codigo: str) -> dict:
    partida = await db.catalogo_partidas.find_one({"codigo": codigo}, {"_id": 0})
    if not partida:
        raise HTTPException(status_code=400, detail=f"Partida '{codigo}' no existe en el catálogo")
    if not partida.get('is_active', True):
        raise HTTPException(status_code=400, detail=f"Partida '{codigo}' está inactiva")
    return partida

# ========================= PROJECT ROUTES =========================
@api_router.get("/projects")
async def get_projects(
    empresa_id: Optional[str] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))
):
    query = {}
    if empresa_id:
        query["empresa_id"] = empresa_id
    projects = await db.projects.find(query, {"_id": 0}).to_list(1000)
    return projects

@api_router.post("/projects")
async def create_project(project_data: ProjectBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    # Validar que empresa existe
    empresa = await db.empresas.find_one({"id": project_data.empresa_id}, {"_id": 0})
    if not empresa:
        raise HTTPException(status_code=400, detail="Empresa no encontrada")
    
    project = Project(**project_data.model_dump())
    doc = project.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.projects.insert_one(doc)
    await log_audit(current_user, "CREATE", "projects", project.id, {"data": {"code": doc['code'], "name": doc['name'], "empresa_id": doc['empresa_id']}})
    doc.pop('_id', None)
    return doc

@api_router.put("/projects/{project_id}")
async def update_project(project_id: str, updates: ProjectBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    old_doc = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    # Validar empresa si se cambia
    if updates.empresa_id:
        empresa = await db.empresas.find_one({"id": updates.empresa_id}, {"_id": 0})
        if not empresa:
            raise HTTPException(status_code=400, detail="Empresa no encontrada")
    
    update_data = updates.model_dump()
    await db.projects.update_one({"id": project_id}, {"$set": update_data})
    await log_audit(current_user, "UPDATE", "projects", project_id, {
        "before": {"code": old_doc.get('code'), "name": old_doc.get('name')},
        "after": {"code": update_data.get('code'), "name": update_data.get('name')}
    })
    
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return Project(**updated)

# ========================= PARTIDA ROUTES =========================
@api_router.get("/partidas", response_model=List[Partida])
async def get_partidas(current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    partidas = await db.partidas.find({}, {"_id": 0}).to_list(1000)
    return [Partida(**p) for p in partidas]

@api_router.post("/partidas", response_model=Partida)
async def create_partida(partida_data: PartidaBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    partida = Partida(**partida_data.model_dump())
    doc = partida.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.partidas.insert_one(doc)
    await log_audit(current_user, "CREATE", "partidas", partida.id, {"data": {"code": doc['code'], "name": doc['name']}})
    return partida

@api_router.put("/partidas/{partida_id}", response_model=Partida)
async def update_partida(partida_id: str, updates: PartidaBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    old_doc = await db.partidas.find_one({"id": partida_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Partida no encontrada")
    
    update_data = updates.model_dump()
    await db.partidas.update_one({"id": partida_id}, {"$set": update_data})
    await log_audit(current_user, "UPDATE", "partidas", partida_id, {
        "before": {"code": old_doc.get('code'), "name": old_doc.get('name')},
        "after": {"code": update_data.get('code'), "name": update_data.get('name')}
    })
    
    updated = await db.partidas.find_one({"id": partida_id}, {"_id": 0})
    return Partida(**updated)

# ========================= PROVIDER ROUTES =========================
@api_router.get("/providers", response_model=List[Provider])
async def get_providers(
    include_inactive: bool = False,
    q: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS)),
):
    query = {} if include_inactive else {"is_active": {"$ne": False}}
    search_text = (q or "").strip()
    if search_text:
        safe_q = re.escape(search_text)
        query["$or"] = [
            {"name": {"$regex": safe_q, "$options": "i"}},
            {"rfc": {"$regex": safe_q, "$options": "i"}},
            {"code": {"$regex": safe_q, "$options": "i"}},
        ]
    max_limit = 100
    default_limit = 20
    effective_limit = min(max(int(limit or default_limit), 1), max_limit)
    safe_offset = max(int(offset or 0), 0)
    providers = await db.providers.find(query, {"_id": 0}).to_list(1000)
    providers.sort(key=lambda item: normalize_for_sort(item.get("name")))
    sliced = providers[safe_offset:safe_offset + effective_limit]
    return [Provider(**p) for p in sliced]

@api_router.post("/providers", response_model=Provider)
async def create_provider(provider_data: ProviderBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    existing = await db.providers.find_one({"code": provider_data.code}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Proveedor con este código ya existe")
    provider = Provider(**provider_data.model_dump())
    doc = provider.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.providers.insert_one(doc)
    await log_audit(current_user, "CREATE", "providers", provider.id, {"data": {"code": doc['code'], "name": doc['name']}})
    return provider

@api_router.put("/providers/{provider_id}", response_model=Provider)
async def update_provider(provider_id: str, updates: ProviderBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    old_doc = await db.providers.find_one({"id": provider_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    update_data = updates.model_dump()
    await db.providers.update_one({"id": provider_id}, {"$set": update_data})
    await log_audit(current_user, "UPDATE", "providers", provider_id, {
        "before": {"code": old_doc.get('code'), "name": old_doc.get('name')},
        "after": {"code": update_data.get('code'), "name": update_data.get('name')}
    })
    
    updated = await db.providers.find_one({"id": provider_id}, {"_id": 0})
    return Provider(**updated)


@api_router.put("/providers/{provider_id}/toggle")
async def toggle_provider(provider_id: str, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    provider = await db.providers.find_one({"id": provider_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    new_status = not provider.get("is_active", True)
    await db.providers.update_one({"id": provider_id}, {"$set": {"is_active": new_status}})
    await log_audit(current_user, "TOGGLE", "providers", provider_id, {"is_active": new_status})
    updated = await db.providers.find_one({"id": provider_id}, {"_id": 0})
    return Provider(**updated)

@api_router.get("/providers/export")
async def export_providers(format: str = Query("csv", pattern="^(csv|xlsx)$"), current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    providers = await db.providers.find({}, {"_id": 0}).sort("code", 1).to_list(5000)
    await log_audit(current_user, "EXPORT", "providers", "batch", {"format": format, "count": len(providers)})
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["code", "name", "rfc", "is_active"])
        writer.writeheader()
        for p in providers:
            writer.writerow({k: p.get(k) for k in ["code", "name", "rfc", "is_active"]})
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=providers.csv"})
    wb = Workbook()
    ws = wb.active
    ws.title = "providers"
    ws.append(["code", "name", "rfc", "is_active"])
    for p in providers:
        ws.append([p.get("code"), p.get("name"), p.get("rfc"), p.get("is_active", True)])
    b=io.BytesIO(); wb.save(b); b.seek(0)
    return StreamingResponse(b, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=providers.xlsx"})

@api_router.post("/providers/import")
async def import_providers(file: UploadFile = File(...), current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    filename=(file.filename or "").lower()
    content=await file.read()
    rows=[]
    if filename.endswith(".csv"):
        text=content.decode("utf-8-sig")
        rows=list(csv.DictReader(io.StringIO(text)))
    elif filename.endswith(".xlsx"):
        wb=load_workbook(io.BytesIO(content), read_only=True)
        ws=wb.active
        headers=[str(c.value).strip() if c.value is not None else "" for c in next(ws.rows)]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: ("" if v is None else str(v)) for i,v in enumerate(r)})
    else:
        raise HTTPException(status_code=400, detail="Formato soportado: csv/xlsx")
    required=["code","name"]
    errors=[]; duplicates=[]; created=0; updated=0
    seen=set()
    for idx,row in enumerate(rows, start=2):
        missing=[c for c in required if not str(row.get(c,"")).strip()]
        if missing:
            errors.append({"row": idx, "error": f"Columnas requeridas faltantes: {', '.join(missing)}"}); continue
        code=str(row.get("code")).strip().upper(); name=str(row.get("name")).strip(); rfc=str(row.get("rfc","")).strip().upper() or None
        is_active=str(row.get("is_active","true")).strip().lower() not in {"false","0","no","inactive"}
        if code in seen:
            duplicates.append({"row": idx, "code": code}); continue
        seen.add(code)
        existing=await db.providers.find_one({"code": code},{"_id":0})
        payload={"code":code,"name":name,"rfc":rfc,"is_active":is_active}
        if existing:
            await db.providers.update_one({"id":existing["id"]},{"$set":payload}); updated += 1
        else:
            provider=Provider(**payload); doc=provider.model_dump(); doc["created_at"]=doc["created_at"].isoformat(); await db.providers.insert_one(doc); created += 1
    await log_audit(current_user, "IMPORT", "providers", "batch", {"filename": file.filename, "rows": len(rows), "created": created, "updated": updated, "duplicates": len(duplicates), "errors": len(errors)})
    return {"total": len(rows), "created": created, "updated": updated, "duplicates": duplicates, "errors": errors}

# ========================= BUDGET ROUTES =========================
@api_router.get("/budgets")
async def get_budgets(
    project_id: Optional[str] = None,
    partida_codigo: Optional[str] = None,
    year: Optional[str] = None,
    month: Optional[str] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_BUDGETS))
):
    normalized_year = None if year in (None, "", "all", "TODO", "todo") else int(year)
    normalized_month = None if month in (None, "", "all", "TODO", "todo") else int(month)

    if normalized_year:
        validate_year_in_range(normalized_year)
    if normalized_month and not normalized_year:
        raise HTTPException(status_code=422, detail={"code": "month_requires_year", "message": "month requiere year"})

    query = {}
    if project_id:
        query["project_id"] = project_id
    if partida_codigo:
        query["partida_codigo"] = partida_codigo

    projects = await db.projects.find({}, {"_id": 0}).to_list(5000)
    project_company_map = {p.get("id"): p.get("empresa_id") for p in projects}

    plans = await db.budget_plans.find(query, {"_id": 0}).to_list(1000)
    plan_items = []
    planned_pairs = set()
    for plan in plans:
        company_id = plan.get("company_id") or project_company_map.get(plan.get("project_id"))
        enforce_company_access(current_user, company_id)

        total_dec = decimal_from_value(plan.get("total_amount", "0"), "total_amount")
        annual_map = plan.get("annual_breakdown") or {}
        monthly_map = plan.get("monthly_breakdown") or {}

        period_amount = total_dec
        period_label = "total"
        if normalized_year and not normalized_month:
            period_label = "annual"
            year_key = str(normalized_year)
            if year_key in annual_map:
                period_amount = decimal_from_value(annual_map.get(year_key), f"annual_breakdown.{year_key}")
            else:
                monthly_sum = sum(
                    decimal_from_value(v, f"monthly_breakdown.{k}")
                    for k, v in monthly_map.items()
                    if str(k).startswith(f"{year_key}-")
                )
                period_amount = monthly_sum if monthly_sum > Decimal("0") else total_dec
        elif normalized_year and normalized_month:
            period_label = "monthly"
            ym_key = f"{normalized_year:04d}-{normalized_month:02d}"
            period_amount = decimal_from_value(monthly_map.get(ym_key, "0"), f"monthly_breakdown.{ym_key}")

        normalized = normalize_plan_response(plan)
        normalized["total_amount"] = str(period_amount)
        normalized["source"] = "plan"
        normalized["period_mode"] = period_label
        normalized["year"] = normalized_year
        normalized["month"] = normalized_month
        availability = await compute_budget_availability(plan.get("project_id"), plan.get("partida_codigo"), datetime((normalized_year or to_tijuana(datetime.now(timezone.utc)).year), (normalized_month or 1), 1))
        normalized.update({
            "budget_total_amount": availability.get("budget_total_amount"),
            "executed_total": availability.get("executed_total"),
            "remaining_total": availability.get("remaining_total"),
            "usage_pct_total": availability.get("usage_pct_total"),
            "executed_annual": availability.get("executed_annual"),
            "remaining_annual": availability.get("remaining_annual"),
            "usage_pct_annual": availability.get("usage_pct_annual"),
            "executed_monthly": availability.get("executed_monthly"),
            "remaining_monthly": availability.get("remaining_monthly"),
            "usage_pct_monthly": availability.get("usage_pct_monthly"),
        })
        plan_items.append(normalized)
        planned_pairs.add((plan.get("project_id"), plan.get("partida_codigo")))

    legacy_query = query.copy()
    if normalized_year:
        legacy_query["year"] = normalized_year
    if normalized_month:
        legacy_query["month"] = normalized_month

    legacy_rows = await db.budgets.find(legacy_query, {"_id": 0}).to_list(5000)
    grouped = {}
    for row in legacy_rows:
        pair = (row.get("project_id"), row.get("partida_codigo"))
        if pair in planned_pairs:
            continue
        company_id = project_company_map.get(row.get("project_id"))
        enforce_company_access(current_user, company_id)
        key = pair if (not normalized_year or not normalized_month) else (pair[0], pair[1], normalized_year, normalized_month)
        entry = grouped.setdefault(key, {
            "id": row.get("id"),
            "project_id": row.get("project_id"),
            "partida_codigo": row.get("partida_codigo"),
            "notes": row.get("notes"),
            "approval_status": "legacy",
            "source": "legacy",
            "period_mode": "monthly" if (normalized_year and normalized_month) else ("annual" if normalized_year else "total"),
            "year": normalized_year,
            "month": normalized_month,
            "total_amount": Decimal("0"),
        })
        entry["total_amount"] += decimal_from_value(row.get("amount_mxn", "0"), "amount_mxn")

    legacy_items = []
    for item in grouped.values():
        final = dict(item)
        final["total_amount"] = str(final["total_amount"])
        availability = await compute_budget_availability(final.get("project_id"), final.get("partida_codigo"), datetime((normalized_year or to_tijuana(datetime.now(timezone.utc)).year), (normalized_month or 1), 1))
        final.update({
            "budget_total_amount": availability.get("budget_total_amount"),
            "executed_total": availability.get("executed_total"),
            "remaining_total": availability.get("remaining_total"),
            "usage_pct_total": availability.get("usage_pct_total"),
            "executed_annual": availability.get("executed_annual"),
            "remaining_annual": availability.get("remaining_annual"),
            "usage_pct_annual": availability.get("usage_pct_annual"),
            "executed_monthly": availability.get("executed_monthly"),
            "remaining_monthly": availability.get("remaining_monthly"),
            "usage_pct_monthly": availability.get("usage_pct_monthly"),
        })
        legacy_items.append(sanitize_mongo_document(final))

    return plan_items + legacy_items


@api_router.get('/budgets/{budget_id}')
async def get_budget_by_id(budget_id: str, current_user: dict = Depends(require_permission(Permission.VIEW_BUDGETS))):
    plan = await db.budget_plans.find_one({"id": budget_id}, {"_id": 0})
    if plan:
        enforce_company_access(current_user, plan.get("company_id"))
        return normalize_plan_response(plan)
    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    project = await db.projects.find_one({"id": budget.get("project_id")}, {"_id": 0})
    enforce_company_access(current_user, project.get("empresa_id") if project else None)
    return sanitize_mongo_document(budget)


@api_router.post("/budgets", status_code=201)
async def create_budget(budget_data: Dict[str, Any], current_user: dict = Depends(require_permission(Permission.MANAGE_BUDGETS, Permission.REQUEST_BUDGETS))):
    # Contrato nuevo (budget plan) detectado por total_amount/annual_breakdown/monthly_breakdown.
    if any(key in budget_data for key in ["total_amount", "annual_breakdown", "monthly_breakdown"]):
        payload = BudgetWriteInput(**budget_data)
        project = await db.projects.find_one({"id": payload.project_id}, {"_id": 0})
        if not project:
            raise HTTPException(status_code=404, detail="Proyecto no encontrado")
        company_id = project.get("empresa_id")
        enforce_company_access(current_user, company_id)
        await validate_partida(payload.partida_codigo)
        enforce_capture_budget_scope(current_user, payload.partida_codigo)
        total, annual, monthly = normalize_budget_breakdown_values(payload.total_amount, payload.annual_breakdown, payload.monthly_breakdown)

        now = datetime.now(timezone.utc).isoformat()
        doc = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "project_id": payload.project_id,
            "partida_codigo": payload.partida_codigo,
            "total_amount": str(total),
            "annual_breakdown": decimal_map_to_strings(annual),
            "monthly_breakdown": decimal_map_to_strings(monthly),
            "notes": payload.notes,
            "created_by": current_user["user_id"],
            "updated_by": current_user["user_id"],
            "created_at": now,
            "updated_at": now,
            "approval_status": BudgetApprovalStatus.NOT_REQUIRED.value,
        }

        if get_budget_approval_mode() == "soft" and total > Decimal("0"):
            if is_admin_or_bypass(current_user):
                doc["approval_status"] = BudgetApprovalStatus.APPROVED.value
            else:
                doc["approval_status"] = BudgetApprovalStatus.PENDING.value

        await db.budget_plans.insert_one(doc)

        if doc["approval_status"] == BudgetApprovalStatus.PENDING.value:
            await ensure_pending_approval(
                approval_type=ApprovalType.BUDGET_DEFINITION,
                company_id=company_id,
                project_id=payload.project_id,
                budget_id=doc["id"],
                requested_by=current_user["user_id"],
                dedupe_key=f"budget_definition:{doc['id']}",
                metadata={"partida_codigo": payload.partida_codigo, "total_amount": str(total)},
            )
        await log_audit(current_user, "CREATE", "budget_plans", doc["id"], {"data": doc})
        return normalize_plan_response(doc)

    # Legacy path
    payload = BudgetBase(**budget_data)
    validate_year_in_range(payload.year)
    if current_user.get("role") != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo admin puede crear presupuestos directamente")
    await validate_partida(payload.partida_codigo)
    enforce_capture_budget_scope(current_user, payload.partida_codigo)
    project = await db.projects.find_one({"id": payload.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=400, detail="Proyecto no encontrado")
    existing = await db.budgets.find_one({"project_id": payload.project_id, "partida_codigo": payload.partida_codigo, "year": payload.year, "month": payload.month}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe presupuesto para este proyecto/partida/mes")
    budget = Budget(**payload.model_dump(), created_by=current_user["user_id"])
    doc = budget.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['amount_mxn'] = str(doc['amount_mxn'])
    await db.budgets.insert_one(doc)
    await log_audit(current_user, "CREATE", "budgets", budget.id, {"data": doc})
    return sanitize_mongo_document(doc)


@api_router.put("/budgets/{budget_id}")
async def update_budget(budget_id: str, updates: Dict[str, Any], current_user: dict = Depends(require_permission(Permission.MANAGE_BUDGETS, Permission.REQUEST_BUDGETS))):
    plan = await db.budget_plans.find_one({"id": budget_id}, {"_id": 0})
    if plan:
        enforce_company_access(current_user, plan.get("company_id"))
        payload = BudgetWriteInput(**updates)
        if payload.project_id != plan.get("project_id"):
            raise HTTPException(status_code=422, detail={"code": "project_id_immutable", "message": "project_id no puede cambiar"})
        if payload.partida_codigo != plan.get("partida_codigo"):
            raise HTTPException(status_code=422, detail={"code": "partida_immutable", "message": "partida_codigo no puede cambiar"})

        enforce_capture_budget_scope(current_user, payload.partida_codigo)
        total, annual, monthly = normalize_budget_breakdown_values(payload.total_amount, payload.annual_breakdown, payload.monthly_breakdown)
        set_data = {
            "total_amount": str(total),
            "annual_breakdown": decimal_map_to_strings(annual),
            "monthly_breakdown": decimal_map_to_strings(monthly),
            "notes": payload.notes,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["user_id"],
        }

        if get_budget_approval_mode() == "soft":
            if total <= Decimal("0"):
                set_data["approval_status"] = BudgetApprovalStatus.NOT_REQUIRED.value
            elif is_admin_or_bypass(current_user):
                set_data["approval_status"] = BudgetApprovalStatus.APPROVED.value
            else:
                set_data["approval_status"] = BudgetApprovalStatus.PENDING.value

        await db.budget_plans.update_one({"id": budget_id}, {"$set": set_data})

        if set_data.get("approval_status") == BudgetApprovalStatus.PENDING.value:
            await ensure_pending_approval(
                approval_type=ApprovalType.BUDGET_DEFINITION,
                company_id=plan.get("company_id"),
                project_id=plan.get("project_id"),
                budget_id=budget_id,
                requested_by=current_user["user_id"],
                dedupe_key=f"budget_definition:{budget_id}",
                metadata={"partida_codigo": plan.get("partida_codigo"), "total_amount": str(total)},
            )

        updated = await db.budget_plans.find_one({"id": budget_id}, {"_id": 0})
        await log_audit(current_user, "UPDATE", "budget_plans", budget_id, {"after": updated})
        return normalize_plan_response(updated)

    payload = BudgetBase(**updates)
    validate_year_in_range(payload.year)
    await validate_partida(payload.partida_codigo)
    enforce_capture_budget_scope(current_user, payload.partida_codigo)
    old_doc = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    update_data = payload.model_dump()
    update_data["amount_mxn"] = str(update_data["amount_mxn"])
    await db.budgets.update_one({"id": budget_id}, {"$set": update_data})
    await log_audit(current_user, "UPDATE", "budgets", budget_id, {"before": old_doc, "after": update_data})
    updated = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    return sanitize_mongo_document(updated)


@api_router.delete("/budgets/{budget_id}")
async def delete_budget(budget_id: str, current_user: dict = Depends(require_permission(Permission.MANAGE_BUDGETS))):
    old_doc = await db.budget_plans.find_one({"id": budget_id}, {"_id": 0})
    if old_doc:
        enforce_company_access(current_user, old_doc.get("company_id"))
        await db.budget_plans.delete_one({"id": budget_id})
        await log_audit(current_user, "DELETE", "budget_plans", budget_id, {"deleted": old_doc})
        return {"message": "Presupuesto eliminado"}

    legacy = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not legacy:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    await db.budgets.delete_one({"id": budget_id})
    await log_audit(current_user, "DELETE", "budgets", budget_id, {"deleted": legacy})
    return {"message": "Presupuesto eliminado"}


@api_router.post("/budgets/{budget_id}/request-approval")
async def budget_request_approval(budget_id: str, current_user: dict = Depends(require_permission(Permission.MANAGE_BUDGETS, Permission.REQUEST_BUDGETS))):
    budget = await get_budget_plan_with_scope(budget_id, current_user)
    approval = await ensure_pending_approval(
        approval_type=ApprovalType.BUDGET_DEFINITION,
        company_id=budget.get("company_id"),
        project_id=budget.get("project_id"),
        budget_id=budget_id,
        requested_by=current_user["user_id"],
        dedupe_key=f"budget_definition:{budget_id}",
        metadata={"partida_codigo": budget.get("partida_codigo"), "total_amount": budget.get("total_amount")},
    )
    await db.budget_plans.update_one({"id": budget_id}, {"$set": {"approval_status": BudgetApprovalStatus.PENDING.value, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user["user_id"]}})
    return sanitize_mongo_document(approval)


@api_router.post("/budgets/{budget_id}/approve")
async def budget_approve(budget_id: str, payload: ApprovalDecisionInput, current_user: dict = Depends(require_permission(Permission.APPROVE_REJECT))):
    budget = await get_budget_plan_with_scope(budget_id, current_user)
    pending = await db.authorizations.find_one({"approval_type": ApprovalType.BUDGET_DEFINITION.value, "budget_id": budget_id, "status": AuthorizationStatus.PENDING.value}, {"_id": 0})
    await db.budget_plans.update_one({"id": budget_id}, {"$set": {"approval_status": BudgetApprovalStatus.APPROVED.value, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user["user_id"]}})
    if pending:
        await db.authorizations.update_one({"id": pending["id"]}, {"$set": {"status": AuthorizationStatus.APPROVED.value, "decision_by": current_user["user_id"], "decision_at": datetime.now(timezone.utc).isoformat(), "comment": payload.comment}})
    return {"message": "Presupuesto aprobado", "budget_id": budget_id}


@api_router.post("/budgets/{budget_id}/reject")
async def budget_reject(budget_id: str, payload: ApprovalDecisionInput, current_user: dict = Depends(require_permission(Permission.APPROVE_REJECT))):
    budget = await get_budget_plan_with_scope(budget_id, current_user)
    pending = await db.authorizations.find_one({"approval_type": ApprovalType.BUDGET_DEFINITION.value, "budget_id": budget_id, "status": AuthorizationStatus.PENDING.value}, {"_id": 0})
    await db.budget_plans.update_one({"id": budget_id}, {"$set": {"approval_status": BudgetApprovalStatus.REJECTED.value, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user["user_id"]}})
    if pending:
        await db.authorizations.update_one({"id": pending["id"]}, {"$set": {"status": AuthorizationStatus.REJECTED.value, "decision_by": current_user["user_id"], "decision_at": datetime.now(timezone.utc).isoformat(), "comment": payload.comment}})
    return {"message": "Presupuesto rechazado", "budget_id": budget_id}


@api_router.post("/budgets/plan", status_code=201)
async def create_budget_plan(payload: BudgetPlanInput, current_user: dict = Depends(require_permission(Permission.MANAGE_BUDGETS))):
    # compatibilidad: redirige al contrato nuevo.
    req = {
        "project_id": payload.project_id,
        "partida_codigo": payload.partida_codigo,
        "total_amount": payload.total,
        "annual_breakdown": payload.annual_breakdown,
        "monthly_breakdown": payload.monthly_breakdown,
        "notes": payload.notes,
    }
    return await create_budget(req, current_user)


@api_router.get("/budget-requests")
async def get_budget_requests(status: Optional[str] = None, current_user: dict = Depends(require_permission(Permission.VIEW_BUDGETS))):
    query = {}
    if status:
        query["status"] = status
    if current_user.get("role") == UserRole.FINANZAS.value:
        query["requested_by"] = current_user["user_id"]
    rows = await db.budget_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [sanitize_mongo_document(r) for r in rows]


@api_router.get("/budget-availability")
async def get_budget_availability(
    project_id: str,
    partida_codigo: str,
    date: Optional[str] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_BUDGETS, Permission.CREATE_MOVEMENT)),
):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail={"code": "project_not_found", "message": "Proyecto no encontrado"})
    enforce_company_access(current_user, project.get("empresa_id"))

    target_date = parse_date_tijuana(date) if date else to_tijuana(datetime.now(timezone.utc))
    availability = await compute_budget_availability(project_id, partida_codigo, target_date)
    is_income = is_ingresos_partida(partida_codigo)
    availability.update({
        "project_id": project_id,
        "partida_codigo": partida_codigo,
        "date": target_date.date().isoformat(),
        "is_income_partida": is_income,
        "budget_validation_applies": False if is_income else availability.get("budget_validation_applies", True),
        "admin_total_only_bypass": current_user.get("role") == UserRole.ADMIN.value and availability.get("has_budget") and availability.get("source") == "plan" and not availability.get("annual_budget") and not availability.get("monthly_budget"),
    })
    return availability


def _po_lock(po_id: str):
    lock_map = getattr(app.state, "po_locks", None)
    if lock_map is None:
        lock_map = {}
        app.state.po_locks = lock_map
    if po_id not in lock_map:
        lock_map[po_id] = asyncio.Lock()
    return lock_map[po_id]


async def _sync_po_odoo_stub(po: dict) -> dict:
    mode = (os.getenv("ODOO_MODE", "stub") or "stub").strip().lower()
    existing = await db.odoo_sync_purchase_orders.find_one({"purchase_order_id": po.get("id")}, {"_id": 0})
    if existing:
        return existing
    if mode == "live":
        raise HTTPException(status_code=501, detail={"code": "odoo_live_not_enabled", "message": "Odoo live not enabled yet"})
    seed = abs(zlib.crc32((po.get("external_id") or "").encode("utf-8"))) % 100000 + 1000
    doc = {
        "id": str(uuid.uuid4()),
        "purchase_order_id": po.get("id"),
        "external_id": po.get("external_id"),
        "odoo_purchase_order_id": seed,
        "odoo_name": f"RFQ{str(seed)[-5:]}",
        "state": "purchase",
        "payload_hash": str(abs(zlib.crc32(json.dumps(po, sort_keys=True, default=str).encode("utf-8")))),
        "stub_seed": seed,
        "last_sync_mode": "stub",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.odoo_sync_purchase_orders.insert_one(doc)
    return doc


async def ensure_partial_indexes_for_movements():
    index_name = "uq_movements_po_line_origin_event_exists"
    expected_keys = [("purchase_order_line_id", ASCENDING), ("origin_event", ASCENDING)]
    expected_partial = {
        "purchase_order_line_id": {"$type": "string"},
        "origin_event": {"$type": "string"},
    }
    try:
        existing = await db.movements.index_information()
    except Exception:
        existing = {}

    legacy_names = ["purchase_order_line_id_1_origin_event_1"]
    for legacy in legacy_names:
        if legacy in existing:
            try:
                await db.movements.drop_index(legacy)
            except Exception:
                pass

    current = existing.get(index_name)
    if current:
        same_key = list(current.get("key", [])) == expected_keys
        same_unique = bool(current.get("unique"))
        same_partial = (current.get("partialFilterExpression") or {}) == expected_partial
        if not (same_key and same_unique and same_partial):
            try:
                await db.movements.drop_index(index_name)
            except Exception:
                pass

    await db.movements.create_index(expected_keys, name=index_name, unique=True, partialFilterExpression=expected_partial)


async def ensure_purchase_order_indexes():
    await db.purchase_orders.create_index([("company_id", ASCENDING), ("external_id", ASCENDING)], unique=True)

    index_name = "uq_purchase_orders_folio_exists"
    expected_keys = [("folio", ASCENDING)]
    expected_partial = {"folio": {"$type": "string"}}
    try:
        existing = await db.purchase_orders.index_information()
    except Exception:
        existing = {}

    # Cleanup legacy/incorrect folio indexes that can break startup with legacy nulls
    for name in ["folio_1", index_name]:
        info = existing.get(name)
        if not info:
            continue
        same_key = list(info.get("key", [])) == expected_keys
        same_unique = bool(info.get("unique"))
        same_partial = (info.get("partialFilterExpression") or {}) == expected_partial
        if not (same_key and same_unique and same_partial):
            try:
                await db.purchase_orders.drop_index(name)
            except Exception:
                pass

    await db.purchase_orders.create_index(expected_keys, unique=True, name=index_name, partialFilterExpression=expected_partial)


def canonicalize_oc_folio(raw_folio: Optional[str]) -> str:
    folio = (raw_folio or "").strip().upper()
    if not folio:
        return "N/A"
    if folio.startswith("OC-"):
        folio = folio[3:]
    if folio.startswith("OC") and len(folio) > 2:
        body = folio[2:]
        if body.isdigit():
            return f"OC{int(body):06d}"
    if folio.isdigit():
        return f"OC{int(folio):06d}"
    return folio


def sanitize_sensitive_dict(payload: Optional[dict]) -> dict:
    data = dict(payload or {})
    for key in list(data.keys()):
        if key.lower() in {"odoo_api_key", "api_key", "token", "password", "access_token"}:
            data[key] = "***"
    return data


async def upsert_collection_doc(collection, query: dict, payload: dict):
    existing = await collection.find_one(query, {"_id": 0})
    if existing:
        await collection.update_one(query, {"$set": payload})
    else:
        doc = dict(query)
        doc.update(payload)
        await collection.insert_one(doc)


async def get_odoo_config() -> dict:
    config_col = getattr(db, "config", None)
    if config_col is None:
        return {
            "odoo_mode": "stub",
            "odoo_url": "",
            "odoo_db": "",
            "odoo_username": "",
            "odoo_api_key": "",
            "default_model": "purchase.order",
        }
    row = await config_col.find_one({"key": "odoo_integration"}, {"_id": 0})
    cfg = row.get("value", {}) if row else {}
    return {
        "odoo_mode": str(cfg.get("odoo_mode", "stub")).lower(),
        "odoo_url": str(cfg.get("odoo_url", "")).strip(),
        "odoo_db": str(cfg.get("odoo_db", "")).strip(),
        "odoo_username": str(cfg.get("odoo_username", "")).strip(),
        "odoo_api_key": str(cfg.get("odoo_api_key", "")).strip(),
        "default_model": str(cfg.get("default_model", "purchase.order")).strip() or "purchase.order",
    }


async def save_odoo_config(current_user: dict, cfg: dict):
    await upsert_collection_doc(
        db.config,
        {"key": "odoo_integration"},
        {"value": cfg, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.get("user_id")},
    )


async def _odoo_send_purchase_order(po: dict, actor: dict, manual_retry: bool = False) -> dict:
    cfg = await get_odoo_config()
    mode = cfg.get("odoo_mode", "stub")
    payload = {
        "purchase_order_id": po.get("id"),
        "folio": po.get("folio") or po.get("external_id"),
        "vendor": {"name": po.get("vendor_name"), "vat": po.get("vendor_rfc"), "email": po.get("vendor_email"), "phone": po.get("vendor_phone"), "street": po.get("vendor_address")},
        "lines": [
            {
                "name": line.get("description") or f"Partida {line.get('partida_codigo')}",
                "product_qty": float(money_dec(line.get("qty", 0))),
                "price_unit": float(money_dec(line.get("price_unit", 0))),
            }
            for line in (po.get("lines") or [])
        ],
    }
    event = "ODOO_RETRY" if manual_retry else "ODOO_AUTO_SEND"
    if mode != "live":
        doc = {
            "id": str(uuid.uuid4()),
            "purchase_order_id": po.get("id"),
            "odoo_status": "stubbed",
            "odoo_model": "purchase.order",
            "odoo_record_id": None,
            "payload_json": payload,
            "last_error": None,
            "last_sync_mode": "stub",
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await upsert_collection_doc(db.odoo_sync_purchase_orders, {"purchase_order_id": po.get("id")}, doc)
        await db.purchase_orders.update_one({"id": po.get("id")}, {"$set": {"odoo_status": "stubbed", "odoo_payload_json": payload, "updated_at": datetime.now(timezone.utc).isoformat()}})
        await log_audit(actor, event, "purchase_orders", po.get("id"), {"odoo_status": "stubbed", "mode": "stub"})
        return doc

    if not cfg.get("odoo_url") or not cfg.get("odoo_db") or not cfg.get("odoo_username") or not cfg.get("odoo_api_key"):
        raise HTTPException(status_code=422, detail=structured_error("odoo_config_incomplete", "Configuración Odoo incompleta para modo LIVE"))

    try:
        common = xmlrpc.client.ServerProxy(f"{cfg['odoo_url'].rstrip('/')}/xmlrpc/2/common")
        uid = common.authenticate(cfg["odoo_db"], cfg["odoo_username"], cfg["odoo_api_key"], {})
        if not uid:
            raise HTTPException(status_code=403, detail=structured_error("odoo_auth_failed", "Autenticación Odoo fallida"))
        models = xmlrpc.client.ServerProxy(f"{cfg['odoo_url'].rstrip('/')}/xmlrpc/2/object")
        vendor = payload["vendor"]
        partner_ids = []
        if vendor.get("vat"):
            partner_ids = models.execute_kw(cfg["odoo_db"], uid, cfg["odoo_api_key"], "res.partner", "search", [[("vat", "=", vendor["vat"])]] , {"limit": 1})
        if not partner_ids and vendor.get("name"):
            partner_ids = models.execute_kw(cfg["odoo_db"], uid, cfg["odoo_api_key"], "res.partner", "search", [[("name", "ilike", vendor["name"])]] , {"limit": 1})
        if partner_ids:
            partner_id = partner_ids[0]
        else:
            partner_id = models.execute_kw(cfg["odoo_db"], uid, cfg["odoo_api_key"], "res.partner", "create", [{
                "name": vendor.get("name") or "Proveedor OC",
                "vat": vendor.get("vat") or False,
                "email": vendor.get("email") or False,
                "phone": vendor.get("phone") or False,
                "street": vendor.get("street") or False,
                "supplier_rank": 1,
            }])

        odoo_lines = [[0, 0, line] for line in payload["lines"]]
        vals = {
            "partner_id": partner_id,
            "origin": payload.get("folio"),
            "date_order": po.get("order_date"),
            "order_line": odoo_lines,
        }
        model_name = cfg.get("default_model") or "purchase.order"
        record_id = models.execute_kw(cfg["odoo_db"], uid, cfg["odoo_api_key"], model_name, "create", [vals])
        url_record = f"{cfg['odoo_url'].rstrip('/')}/web#id={record_id}&model={model_name}&view_type=form"
        doc = {
            "id": str(uuid.uuid4()),
            "purchase_order_id": po.get("id"),
            "odoo_status": "sent",
            "odoo_model": model_name,
            "odoo_record_id": str(record_id),
            "odoo_url_to_record": url_record,
            "payload_json": payload,
            "last_error": None,
            "last_sync_mode": "live",
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await upsert_collection_doc(db.odoo_sync_purchase_orders, {"purchase_order_id": po.get("id")}, doc)
        await db.purchase_orders.update_one({"id": po.get("id")}, {"$set": {"odoo_status": "sent", "odoo_model": model_name, "odoo_record_id": str(record_id), "odoo_payload_json": payload, "odoo_url_to_record": url_record, "updated_at": datetime.now(timezone.utc).isoformat()}})
        await log_audit(actor, event, "purchase_orders", po.get("id"), {"odoo_status": "sent", "odoo_model": model_name, "odoo_record_id": str(record_id), "mode": "live"})
        return doc
    except HTTPException:
        raise
    except Exception as exc:
        msg = str(exc)
        await upsert_collection_doc(db.odoo_sync_purchase_orders, {"purchase_order_id": po.get("id")}, {
            "id": str(uuid.uuid4()),
            "purchase_order_id": po.get("id"),
            "odoo_status": "failed",
            "odoo_model": cfg.get("default_model") or "purchase.order",
            "odoo_record_id": None,
            "payload_json": payload,
            "last_error": msg,
            "last_sync_mode": "live",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        await db.purchase_orders.update_one({"id": po.get("id")}, {"$set": {"odoo_status": "failed", "odoo_last_error": msg, "odoo_payload_json": payload, "updated_at": datetime.now(timezone.utc).isoformat()}})
        await log_audit(actor, event, "purchase_orders", po.get("id"), {"odoo_status": "failed", "mode": "live", "error": msg})
        raise HTTPException(status_code=409, detail=structured_error("odoo_send_failed", "No se pudo enviar a Odoo", {"error": msg}))


def structured_error(code: str, message: str, details: Optional[Dict[str, Any]] = None) -> dict:
    payload = {"code": code, "message": message}
    if details:
        payload["details"] = to_json_safe(details)
    return payload


async def build_purchase_order_budget_summary(po: dict, approval_amount: Optional[Decimal] = None) -> dict:
    order_date = date_parser.parse(po.get("order_date")) if po.get("order_date") else datetime.now(timezone.utc)
    po_total = money_dec(po.get("total", 0)).quantize(TWO_DECIMALS)
    approved_total = money_dec(po.get("approved_amount_total", 0)).quantize(TWO_DECIMALS)
    pending_total = (po_total - approved_total).quantize(TWO_DECIMALS)
    if pending_total < 0:
        pending_total = Decimal("0.00")
    preview_partial = money_dec(approval_amount) if approval_amount is not None else pending_total
    if preview_partial < 0:
        preview_partial = Decimal("0.00")

    partida_docs = await db.catalogo_partidas.find({}, {"_id": 0}).to_list(2000)
    partida_map = {str(p.get("codigo")): p.get("nombre") for p in partida_docs}

    ratio_partial = (preview_partial / po_total) if po_total > 0 else Decimal("0")
    ratio_full = (pending_total / po_total) if po_total > 0 else Decimal("0")

    by_partida = []
    aggregate = {
        "budget_total": Decimal("0.00"),
        "executed_current": Decimal("0.00"),
        "available_current": Decimal("0.00"),
        "available_after_partial": Decimal("0.00"),
        "available_after_full": Decimal("0.00"),
    }

    lines_by_partida: Dict[str, Decimal] = {}
    for line in po.get("lines", []):
        partida = str(line.get("partida_codigo") or "")
        if not partida:
            continue
        lines_by_partida[partida] = (lines_by_partida.get(partida, Decimal("0.00")) + money_dec(line.get("line_total", 0))).quantize(TWO_DECIMALS)

    for partida, line_total in sorted(lines_by_partida.items()):
        approved_line = (line_total * (approved_total / po_total)).quantize(TWO_DECIMALS) if po_total > 0 else Decimal("0.00")
        pending_line = (line_total - approved_line).quantize(TWO_DECIMALS)
        partial_line = (line_total * ratio_partial).quantize(TWO_DECIMALS)
        full_line = (line_total * ratio_full).quantize(TWO_DECIMALS)

        if partida in {"400", "401", "402", "403", "404"}:
            by_partida.append({
                "partida_codigo": partida,
                "partida_nombre": partida_map.get(partida),
                "budget_validation_applies": False,
                "budget_total": None,
                "executed_current": None,
                "available_current": None,
                "po_committed_current": None,
                "oc_line_total": str(line_total),
                "approved_accumulated_line": str(approved_line),
                "pending_line_amount": str(pending_line if pending_line > 0 else Decimal("0.00")),
                "available_after_partial_line": None,
                "available_after_full_line": None,
                "scope": "income",
            })
            continue

        av = await compute_budget_availability(po.get("project_id"), partida, order_date)
        b_total = money_dec(av.get("budget_total_amount", 0)).quantize(TWO_DECIMALS) if av.get("budget_total_amount") is not None else Decimal("0.00")
        b_exec = money_dec(av.get("executed_total", 0)).quantize(TWO_DECIMALS) if av.get("executed_total") is not None else Decimal("0.00")
        b_avail = money_dec(av.get("remaining_total", 0)).quantize(TWO_DECIMALS) if av.get("remaining_total") is not None else Decimal("0.00")
        after_partial = (b_avail - partial_line).quantize(TWO_DECIMALS)
        after_full = (b_avail - pending_line).quantize(TWO_DECIMALS)

        aggregate["budget_total"] += b_total
        aggregate["executed_current"] += b_exec
        aggregate["available_current"] += b_avail
        aggregate["available_after_partial"] += after_partial
        aggregate["available_after_full"] += after_full

        by_partida.append({
            "partida_codigo": partida,
            "partida_nombre": partida_map.get(partida),
            "budget_validation_applies": bool(av.get("budget_validation_applies", True)),
            "scope": av.get("effective_scope") or "total",
            "budget_total": str(b_total),
            "executed_current": str(b_exec),
            "available_current": str(b_avail),
            "po_committed_current": str(line_total),
            "oc_line_total": str(line_total),
            "approved_accumulated_line": str(approved_line),
            "pending_line_amount": str(pending_line if pending_line > 0 else Decimal("0.00")),
            "available_after_partial_line": str(after_partial),
            "available_after_full_line": str(after_full),
        })

    summary = {
        "scope": "total",
        "budget_total": str(aggregate["budget_total"].quantize(TWO_DECIMALS)),
        "executed_current": str(aggregate["executed_current"].quantize(TWO_DECIMALS)),
        "available_current": str(aggregate["available_current"].quantize(TWO_DECIMALS)),
        "po_total": str(po_total),
        "approved_accumulated": str(approved_total),
        "pending_amount": str(pending_total),
        "preview_partial_amount": str(preview_partial.quantize(TWO_DECIMALS)),
        "available_after_partial": str(aggregate["available_after_partial"].quantize(TWO_DECIMALS)),
        "available_after_full": str(aggregate["available_after_full"].quantize(TWO_DECIMALS)),
        "by_partida": by_partida,
    }
    # backward-compatible aliases already used by frontend
    summary.update({
        "presupuesto_total": summary["budget_total"],
        "ejecutado_actual": summary["executed_current"],
        "disponible_actual": summary["available_current"],
        "monto_oc": summary["po_total"],
        "monto_aprobado_acumulado": summary["approved_accumulated"],
        "monto_pendiente_oc": summary["pending_amount"],
        "monto_a_aprobar": summary["preview_partial_amount"],
        "restante_proyectado_si_aprueba": summary["available_after_full"],
    })
    return summary


async def generate_purchase_order_folio() -> str:
    counter = await db.counters.find_one_and_update(
        {"_id": "purchase_order_folio"},
        {"$inc": {"seq": 1}, "$setOnInsert": {"created_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = int(counter.get("seq", 1))
    return f"OC{seq:06d}"


def normalize_purchase_order_response(doc: dict) -> dict:
    normalized = sanitize_mongo_document(doc)
    normalized.setdefault("folio", normalized.get("external_id"))
    normalized.setdefault("external_id", normalized.get("folio"))
    normalized["subtotal"] = normalized.get("subtotal_tax_base")
    normalized["iva"] = normalized.get("tax_total")
    normalized["iva_retenido"] = normalized.get("iva_withholding_total", "0.00")
    normalized["total_neto"] = normalized.get("total")
    normalized["total_original"] = normalized.get("total")
    normalized["projected_amount_original"] = normalized.get("total")
    normalized["projected_amount_mxn"] = normalized.get("total_mxn", normalized.get("total"))
    return normalized


@api_router.post("/budgets/availability/oc-preview")
async def oc_budget_preview(payload: OCBudgetPreviewInput, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS))):
    project = await db.projects.find_one({"id": payload.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail={"code": "project_not_found", "message": "Proyecto no encontrado"})
    enforce_company_access(current_user, project.get("empresa_id"))
    exchange_rate_dec = parse_amount_like(payload.exchange_rate if payload.exchange_rate is not None else 1, "exchange_rate")
    if payload.currency != Currency.MXN and exchange_rate_dec <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_exchange_rate", "message": "exchange_rate debe ser mayor a 0 para moneda distinta a MXN"})
    target_date = parse_date_tijuana(payload.order_date)
    items = []
    grouped_requested_original: Dict[str, Decimal] = {}
    grouped_requested_mxn: Dict[str, Decimal] = {}
    for line in payload.lines:
        partida = str(line.get("partida_codigo"))
        requested_original = parse_amount_like(line.get("requested_amount", "0"), "requested_amount")
        requested_mxn = requested_original if payload.currency == Currency.MXN else (requested_original * exchange_rate_dec).quantize(TWO_DECIMALS)
        grouped_requested_original[partida] = (grouped_requested_original.get(partida, Decimal("0.00")) + requested_original).quantize(TWO_DECIMALS)
        grouped_requested_mxn[partida] = (grouped_requested_mxn.get(partida, Decimal("0.00")) + requested_mxn).quantize(TWO_DECIMALS)

    for partida, requested in grouped_requested_mxn.items():
        requested_original = grouped_requested_original.get(partida, Decimal("0.00")).quantize(TWO_DECIMALS)
        if partida in {"400", "401", "402", "403", "404"}:
            items.append({
                "partida_codigo": partida,
                "requested_amount": str(requested),
                "requested_amount_original": str(requested_original),
                "projected_amount_mxn": str(requested),
                "projected_amount_original": str(requested_original),
                "budget_validation_applies": False,
            })
            continue
        av = await compute_budget_availability(payload.project_id, partida, target_date)
        remaining_total = money_dec(av.get("remaining_total", 0)) if av.get("remaining_total") is not None else Decimal("0")
        projected_remaining_total = (remaining_total - requested).quantize(TWO_DECIMALS)
        remaining_annual = money_dec(av.get("remaining_annual", 0)) if av.get("remaining_annual") is not None else None
        projected_remaining_annual = (remaining_annual - requested).quantize(TWO_DECIMALS) if remaining_annual is not None else None
        remaining_monthly = money_dec(av.get("remaining_monthly", 0)) if av.get("remaining_monthly") is not None else None
        projected_remaining_monthly = (remaining_monthly - requested).quantize(TWO_DECIMALS) if remaining_monthly is not None else None
        exceeded = []
        if projected_remaining_total < 0:
            exceeded.append("total")
        if projected_remaining_annual is not None and projected_remaining_annual < 0:
            exceeded.append("annual")
        if projected_remaining_monthly is not None and projected_remaining_monthly < 0:
            exceeded.append("monthly")
        items.append({
            "partida_codigo": partida,
            "requested_amount": str(requested),
            "requested_amount_original": str(requested_original),
            "projected_amount_mxn": str(requested),
            "projected_amount_original": str(requested_original),
            "period": f"{target_date.year:04d}-{target_date.month:02d}",
            "budget_total_defined": bool(av.get("budget_total_amount") is not None),
            "budget_annual_defined": bool(av.get("annual_budget") is not None),
            "budget_monthly_defined": bool(av.get("monthly_budget") is not None),
            "budget_total": av.get("budget_total_amount"),
            "executed_total": av.get("executed_total"),
            "remaining_total_current": str(remaining_total),
            "remaining_annual_current": str(remaining_annual) if remaining_annual is not None else None,
            "remaining_monthly_current": str(remaining_monthly) if remaining_monthly is not None else None,
            "projected_remaining_total": str(projected_remaining_total),
            "projected_remaining_annual": str(projected_remaining_annual) if projected_remaining_annual is not None else None,
            "projected_remaining_monthly": str(projected_remaining_monthly) if projected_remaining_monthly is not None else None,
            "buckets_exceeded": exceeded,
            "can_proceed_workflow": True,
            "can_post_payment": len(exceeded) == 0,
        })

    summary_budget = Decimal("0.00")
    summary_executed = Decimal("0.00")
    summary_available = Decimal("0.00")
    summary_requested = Decimal("0.00")
    summary_requested_original = Decimal("0.00")
    summary_projected = Decimal("0.00")
    for item in items:
        if item.get("budget_validation_applies") is False:
            continue
        summary_budget += money_dec(item.get("budget_total") or 0)
        summary_executed += money_dec(item.get("executed_total") or 0)
        summary_available += money_dec(item.get("remaining_total_current") or 0)
        summary_requested += money_dec(item.get("requested_amount") or 0)
        summary_requested_original += money_dec(item.get("requested_amount_original") or 0)
        summary_projected += money_dec(item.get("projected_remaining_total") or 0)
    summary = {
        "budget_total": str(summary_budget.quantize(TWO_DECIMALS)),
        "executed_current": str(summary_executed.quantize(TWO_DECIMALS)),
        "available_current": str(summary_available.quantize(TWO_DECIMALS)),
        "po_preview_amount": str(summary_requested.quantize(TWO_DECIMALS)),
        "projected_amount_mxn": str(summary_requested.quantize(TWO_DECIMALS)),
        "projected_amount_original": str(summary_requested_original.quantize(TWO_DECIMALS)),
        "available_after_preview": str(summary_projected.quantize(TWO_DECIMALS)),
        "presupuesto_total": str(summary_budget.quantize(TWO_DECIMALS)),
        "ejecutado_actual": str(summary_executed.quantize(TWO_DECIMALS)),
        "disponible_actual": str(summary_available.quantize(TWO_DECIMALS)),
        "monto_solicitado": str(summary_requested.quantize(TWO_DECIMALS)),
        "monto_solicitado_original": str(summary_requested_original.quantize(TWO_DECIMALS)),
        "restante_proyectado": str(summary_projected.quantize(TWO_DECIMALS)),
    }
    return {
        "project_id": payload.project_id,
        "currency": payload.currency.value,
        "exchange_rate": str(exchange_rate_dec),
        "by_partida": items,
        "lines": items,
        "summary": summary,
    }


@api_router.post("/purchase-orders")
async def create_purchase_order(payload: PurchaseOrderCreate, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS))):
    project = await db.projects.find_one({"id": payload.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail={"code": "project_not_found", "message": "Proyecto no encontrado"})
    enforce_company_access(current_user, project.get("empresa_id"))
    folio = (payload.external_id or "").strip()
    if folio:
        existing = await db.purchase_orders.find_one({"company_id": project.get("empresa_id"), "folio": folio}, {"_id": 0})
    else:
        existing = None
    lines = [calculate_oc_line(line) for line in payload.lines]
    exchange_rate_dec = parse_amount_like(payload.exchange_rate if payload.exchange_rate is not None else 1, "exchange_rate")
    if payload.currency != Currency.MXN and exchange_rate_dec <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_exchange_rate", "message": "exchange_rate debe ser mayor a 0 para moneda distinta a MXN"})
    totals = summarize_oc_lines(
        lines,
        apply_iva_withholding=payload.apply_iva_withholding,
        iva_withholding_rate=parse_amount_like(payload.iva_withholding_rate or 0, "iva_withholding_rate"),
        currency=payload.currency.value,
        exchange_rate=exchange_rate_dec,
    )
    po_base = {
        "folio": folio,
        "external_id": folio,
        "invoice_folio": (payload.supplier_invoice_folio or payload.invoice_folio or "").strip()[:100] or None,
        "company_id": project.get("empresa_id"),
        "project_id": payload.project_id,
        "vendor_name": payload.vendor_name,
        "vendor_rfc": payload.vendor_rfc,
        "vendor_email": payload.vendor_email,
        "vendor_phone": payload.vendor_phone,
        "vendor_address": payload.vendor_address,
        "currency": payload.currency.value,
        "exchange_rate": str(exchange_rate_dec),
        "apply_iva_withholding": payload.apply_iva_withholding,
        "iva_withholding_rate": str(parse_amount_like(payload.iva_withholding_rate or 0, "iva_withholding_rate")),
        "order_date": parse_date_tijuana(payload.order_date).isoformat(),
        "planned_date": parse_date_tijuana(payload.planned_date).isoformat() if payload.planned_date else None,
        "notes": payload.notes,
        "payment_terms": payload.payment_terms,
        "fob": payload.fob,
        "lines": lines,
        **totals,
    }
    payload_hash = str(abs(zlib.crc32(json.dumps(po_base, sort_keys=True, default=str).encode("utf-8"))))
    if existing:
        if existing.get("payload_hash") == payload_hash:
            return {"purchase_order": sanitize_mongo_document(existing), "idempotent": True}
        if existing.get("status") in {PurchaseOrderStatus.DRAFT.value, PurchaseOrderStatus.REJECTED.value}:
            raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "OC existente con payload distinto; use PUT"})
        raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "OC existente no editable"})

    if not folio:
        folio = await generate_purchase_order_folio()
        po_base["folio"] = folio
        po_base["external_id"] = folio

    po = {
        "id": str(uuid.uuid4()),
        **po_base,
        "payload_hash": payload_hash,
        "status": PurchaseOrderStatus.DRAFT.value,
        "budget_gate_status": BudgetGateStatus.NOT_CHECKED.value,
        "posting_status": PostingStatus.NOT_POSTED.value,
        "created_by_user_id": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.purchase_orders.insert_one(po)
    await log_audit(current_user, "CREATE", "purchase_orders", po["id"], {"folio": po["folio"], "external_id": po["external_id"]})
    return {"purchase_order": normalize_purchase_order_response(po)}


@api_router.put("/purchase-orders/{po_id}")
async def update_purchase_order(po_id: str, payload: PurchaseOrderCreate, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS))):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
    enforce_company_access(current_user, po.get("company_id"))
    if po.get("status") not in {PurchaseOrderStatus.DRAFT.value, PurchaseOrderStatus.REJECTED.value}:
        raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "Solo se puede editar DRAFT/REJECTED"})
    lines = [calculate_oc_line(line) for line in payload.lines]
    exchange_rate_dec = parse_amount_like(payload.exchange_rate if payload.exchange_rate is not None else 1, "exchange_rate")
    if payload.currency != Currency.MXN and exchange_rate_dec <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_exchange_rate", "message": "exchange_rate debe ser mayor a 0 para moneda distinta a MXN"})
    totals = summarize_oc_lines(
        lines,
        apply_iva_withholding=payload.apply_iva_withholding,
        iva_withholding_rate=parse_amount_like(payload.iva_withholding_rate or 0, "iva_withholding_rate"),
        currency=payload.currency.value,
        exchange_rate=exchange_rate_dec,
    )
    update_doc = {
        "invoice_folio": (payload.supplier_invoice_folio or payload.invoice_folio or "").strip()[:100] or None,
        "vendor_name": payload.vendor_name,
        "vendor_rfc": payload.vendor_rfc,
        "vendor_email": payload.vendor_email,
        "vendor_phone": payload.vendor_phone,
        "vendor_address": payload.vendor_address,
        "currency": payload.currency.value,
        "exchange_rate": str(exchange_rate_dec),
        "apply_iva_withholding": payload.apply_iva_withholding,
        "iva_withholding_rate": str(parse_amount_like(payload.iva_withholding_rate or 0, "iva_withholding_rate")),
        "order_date": parse_date_tijuana(payload.order_date).isoformat(),
        "planned_date": parse_date_tijuana(payload.planned_date).isoformat() if payload.planned_date else None,
        "lines": lines,
        **totals,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "budget_gate_status": BudgetGateStatus.NOT_CHECKED.value,
        "posting_status": PostingStatus.NOT_POSTED.value,
    }
    await db.purchase_orders.update_one({"id": po_id}, {"$set": update_doc})
    saved = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    return {"purchase_order": normalize_purchase_order_response(saved)}


@api_router.post("/purchase-orders/{po_id}/submit")
async def submit_purchase_order(po_id: str, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS))):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
    enforce_company_access(current_user, po.get("company_id"))
    if po.get("status") not in {PurchaseOrderStatus.DRAFT.value, PurchaseOrderStatus.REJECTED.value}:
        raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "Estado inválido para submit"})
    gate = await evaluate_oc_budget_gate(po)
    await db.purchase_orders.update_one({"id": po_id}, {"$set": {
        "status": PurchaseOrderStatus.PENDING_APPROVAL.value,
        "budget_gate_status": BudgetGateStatus.OK.value if gate.get("ok") else BudgetGateStatus.EXCEPTION_PENDING.value,
        "budget_check_snapshot_json": gate,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    await ensure_pending_approval(
        approval_type=ApprovalType.PURCHASE_ORDER_WORKFLOW,
        company_id=po.get("company_id"),
        project_id=po.get("project_id"),
        requested_by=current_user["user_id"],
        purchase_order_id=po_id,
        dedupe_key=f"purchase_order_workflow:{po_id}",
        metadata={"folio": po.get("folio") or po.get("external_id"), "purchase_order_id": po_id},
    )
    await log_audit(current_user, "SUBMIT", "purchase_orders", po_id, {"status": PurchaseOrderStatus.PENDING_APPROVAL.value})
    saved = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    return {"purchase_order": normalize_purchase_order_response(saved)}


@api_router.post("/purchase-orders/{po_id}/approve")
async def approve_purchase_order(po_id: str, current_user: dict = Depends(require_roles(UserRole.ADMIN))):
    lock = _po_lock(po_id)
    async with lock:
        po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        if not po:
            raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
        enforce_company_access(current_user, po.get("company_id"))
        if po.get("posting_status") == PostingStatus.POSTED.value:
            return {"purchase_order": normalize_purchase_order_response(po), "idempotent": True}
        if po.get("status") not in {PurchaseOrderStatus.PENDING_APPROVAL.value, PurchaseOrderStatus.DRAFT.value}:
            raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "Estado inválido para approve"})

        gate = await evaluate_oc_budget_gate(po)
        if not gate.get("ok"):
            dedupe = f"po_overbudget:{po['id']}:{','.join(sorted({e.get('partida_codigo','') for e in gate.get('exceeded', [])}))}"
            approval = await ensure_pending_approval(
                approval_type=ApprovalType.OVERBUDGET_EXCEPTION,
                company_id=po.get("company_id"),
                project_id=po.get("project_id"),
                requested_by=current_user["user_id"],
                dedupe_key=dedupe,
                metadata={"purchase_order_id": po["id"], "external_id": po.get("external_id"), "exceeded": gate.get("exceeded")},
            )
            await db.purchase_orders.update_one({"id": po_id}, {"$set": {
                "budget_gate_status": BudgetGateStatus.EXCEPTION_PENDING.value,
                "posting_status": PostingStatus.NOT_POSTED.value,
                "budget_exception_approval_id": approval.get("id"),
                "budget_check_snapshot_json": gate,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }})
            raise HTTPException(status_code=409, detail={"code": "purchase_order_budget_exception", "message": "La orden de compra excede el presupuesto disponible y no puede postearse.", "meta": {"purchase_order_id": po_id, "approval_request_id": approval.get("id"), "buckets_exceeded": gate.get("exceeded")}})

        movement_ids = []
        for line in po.get("lines", []):
            dedupe_key = f"{po_id}:{line.get('id')}:OC_APPROVE"
            existing_mv = await db.movements.find_one({"purchase_order_line_id": line.get("id"), "origin_event": "OC_APPROVE"}, {"_id": 0})
            if existing_mv:
                movement_ids.append(existing_mv.get("id"))
                continue
            movement_doc = {
                "id": str(uuid.uuid4()),
                "project_id": po.get("project_id"),
                "partida_codigo": line.get("partida_codigo"),
                "provider_id": None,
                "date": po.get("order_date"),
                "currency": po.get("currency"),
                "amount_original": float(money_dec(line.get("line_total", 0))),
                "exchange_rate": float(decimal_from_value(po.get("exchange_rate", 1), "exchange_rate")),
                "amount_mxn": float(money_dec(line.get("line_total", 0))),
                "reference": po.get("external_id"),
                "description": f"OC {po.get('external_id')} línea {line.get('line_no')}",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["user_id"],
                "status": MovementStatus.POSTED.value,
                "purchase_order_id": po_id,
                "purchase_order_line_id": line.get("id"),
                "origin_event": "OC_APPROVE",
                "idempotency_key": dedupe_key,
                "is_active": True,
            }
            await db.movements.insert_one(movement_doc)
            await log_audit(current_user, "CREATE", "movements", movement_doc["id"], {"origin_event": "OC_APPROVE", "purchase_order_id": po_id})
            movement_ids.append(movement_doc["id"])

        try:
            odoo_sync = await _odoo_send_purchase_order(po, current_user, manual_retry=False)
        except HTTPException as exc:
            odoo_sync = {"odoo_status": "failed", "last_error": str(exc.detail)}

        await db.purchase_orders.update_one({"id": po_id}, {"$set": {
            "status": PurchaseOrderStatus.APPROVED_FOR_PAYMENT.value,
            "approved_by_user_id": current_user["user_id"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "budget_gate_status": BudgetGateStatus.OK.value,
            "posting_status": PostingStatus.POSTED.value,
            "budget_check_snapshot_json": gate,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }})
        await db.authorizations.update_many(
            {"approval_type": ApprovalType.PURCHASE_ORDER_WORKFLOW.value, "purchase_order_id": po_id, "status": AuthorizationStatus.PENDING.value},
            {"$set": {"status": AuthorizationStatus.APPROVED.value, "decision_by": current_user["user_id"], "decision_at": datetime.now(timezone.utc).isoformat()}},
        )
        saved = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        await log_audit(current_user, "APPROVE", "purchase_orders", po_id, {"movement_ids": movement_ids, "odoo_purchase_order_id": odoo_sync.get("odoo_purchase_order_id")})
        return {"purchase_order": normalize_purchase_order_response(saved), "movement_ids": movement_ids, "odoo": sanitize_mongo_document(odoo_sync)}


@api_router.post("/purchase-orders/{po_id}/reject")
async def reject_purchase_order(po_id: str, payload: PurchaseOrderRejectInput, current_user: dict = Depends(require_roles(UserRole.ADMIN))):
    lock = _po_lock(po_id)
    async with lock:
        po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        if not po:
            raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
        enforce_company_access(current_user, po.get("company_id"))
        if po.get("status") != PurchaseOrderStatus.PENDING_APPROVAL.value:
            raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "Estado inválido para reject"})
        if not payload.reason.strip():
            raise HTTPException(status_code=422, detail={"code": "rejection_reason_required", "message": "reason es obligatorio"})
        await db.purchase_orders.update_one({"id": po_id}, {"$set": {"status": PurchaseOrderStatus.REJECTED.value, "rejected_by_user_id": current_user["user_id"], "rejected_at": datetime.now(timezone.utc).isoformat(), "rejection_reason": payload.reason.strip(), "updated_at": datetime.now(timezone.utc).isoformat()}})
        await db.authorizations.update_many(
            {"approval_type": ApprovalType.PURCHASE_ORDER_WORKFLOW.value, "purchase_order_id": po_id, "status": AuthorizationStatus.PENDING.value},
            {"$set": {"status": AuthorizationStatus.REJECTED.value, "decision_by": current_user["user_id"], "decision_at": datetime.now(timezone.utc).isoformat(), "comment": payload.reason.strip()}},
        )
        await log_audit(current_user, "REJECT", "purchase_orders", po_id, {"reason": payload.reason.strip()})
        saved = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        return {"purchase_order": normalize_purchase_order_response(saved)}


@api_router.delete("/purchase-orders/{po_id}")
async def cancel_purchase_order(po_id: str, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS))):
    lock = _po_lock(po_id)
    async with lock:
        po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
        if not po:
            raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
        enforce_company_access(current_user, po.get("company_id"))
        if po.get("status") == PurchaseOrderStatus.APPROVED_FOR_PAYMENT.value:
            raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "No se puede cancelar OC aprobada"})
        if po.get("status") not in {PurchaseOrderStatus.DRAFT.value, PurchaseOrderStatus.REJECTED.value, PurchaseOrderStatus.PENDING_APPROVAL.value}:
            raise HTTPException(status_code=409, detail={"code": "oc_state_conflict", "message": "Estado inválido para cancelar"})
        await db.purchase_orders.update_one({"id": po_id}, {"$set": {"status": PurchaseOrderStatus.CANCELLED.value, "cancelled_by_user_id": current_user["user_id"], "cancelled_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}})
        await log_audit(current_user, "CANCEL", "purchase_orders", po_id, {})
        return {"message": "OC cancelada"}


@api_router.get("/purchase-orders")
async def list_purchase_orders(status: Optional[str] = None, project_id: Optional[str] = None, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS, UserRole.DIRECTOR))):
    query = {}
    if status:
        query["status"] = status
    if project_id:
        query["project_id"] = project_id
    rows = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    out = []
    for row in rows:
        if has_company_access(current_user, row.get("company_id")):
            out.append(normalize_purchase_order_response(row))
    return out


@api_router.get("/purchase-orders/{po_id}")
async def get_purchase_order(po_id: str, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS, UserRole.DIRECTOR))):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
    enforce_company_access(current_user, po.get("company_id"))
    return normalize_purchase_order_response(po)


@api_router.get("/purchase-orders/{po_id}/pdf")
async def purchase_order_pdf(po_id: str, current_user: dict = Depends(require_roles(UserRole.ADMIN, UserRole.FINANZAS, UserRole.DIRECTOR))):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail={"code": "purchase_order_not_found", "message": "OC no encontrada"})
    enforce_company_access(current_user, po.get("company_id"))
    project = await db.projects.find_one({"id": po.get("project_id")}, {"_id": 0}) if po.get("project_id") else None
    company_id = po.get("company_id") or (project or {}).get("empresa_id")
    empresa = await db.empresas.find_one({"id": company_id}, {"_id": 0}) if company_id else None
    render_payload = {**po, "company_name": (empresa or {}).get("nombre"), "project_name": (project or {}).get("name"), "project_code": (project or {}).get("code")}
    logo_path = os.getenv("QF_OC_LOGO_PATH")
    if logo_path and not Path(logo_path).exists():
        logger.warning("QF_OC_LOGO_PATH no existe: %s", logo_path)
    pdf_bytes = render_purchase_order_pdf(render_payload)
    await log_audit(current_user, "PDF", "purchase_orders", po_id, {"folio": po.get("folio") or po.get("external_id")})
    filename = oc_pdf_filename(po.get("folio") or po.get("external_id") or po_id)
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers={"Content-Disposition": f"inline; filename={filename}"})


@api_router.post("/budget-requests", status_code=201)
async def create_budget_request(payload: BudgetRequestBase, current_user: dict = Depends(require_permission(Permission.REQUEST_BUDGETS, Permission.MANAGE_BUDGETS))):
    validate_year_in_range(payload.year)
    await validate_partida(payload.partida_codigo)
    enforce_capture_budget_scope(current_user, payload.partida_codigo)
    if current_user.get("role") not in [UserRole.FINANZAS.value, UserRole.ADMIN.value]:
        raise HTTPException(status_code=403, detail="Rol sin permisos para solicitar presupuesto")
    req = BudgetRequest(**payload.model_dump(), requested_by=current_user["user_id"])
    doc = req.model_dump(); doc["created_at"] = doc["created_at"].isoformat(); doc["amount_mxn"] = str(doc["amount_mxn"])
    await db.budget_requests.insert_one(doc)
    await log_audit(current_user, "CREATE", "budget_requests", req.id, {"data": payload.model_dump()})
    return sanitize_mongo_document(doc)


@api_router.put("/budget-requests/{request_id}/resolve")
async def resolve_budget_request(request_id: str, resolution: AuthorizationResolve, current_user: dict = Depends(require_permission(Permission.MANAGE_BUDGETS))):
    if current_user.get("role") != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo admin puede resolver solicitudes")
    req = await db.budget_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if req.get("status") != AuthorizationStatus.PENDING.value:
        raise HTTPException(status_code=400, detail="Solicitud ya resuelta")
    update = {"status": resolution.status.value, "resolved_at": datetime.now(timezone.utc).isoformat(), "resolved_by": current_user["user_id"]}
    await db.budget_requests.update_one({"id": request_id}, {"$set": update})
    if resolution.status == AuthorizationStatus.APPROVED:
        existing = await db.budgets.find_one({"project_id": req["project_id"], "partida_codigo": req["partida_codigo"], "year": req["year"], "month": req["month"]}, {"_id":0})
        if not existing:
            budget = Budget(project_id=req["project_id"], partida_codigo=req["partida_codigo"], year=req["year"], month=req["month"], amount_mxn=req["amount_mxn"], notes=req.get("notes"), created_by=current_user["user_id"])
            d=budget.model_dump(); d["created_at"]=d["created_at"].isoformat(); d["amount_mxn"] = str(d["amount_mxn"]); await db.budgets.insert_one(d)
    await log_audit(current_user, f"BUDGET_REQUEST_{resolution.status.value.upper()}", "budget_requests", request_id, {"notes": resolution.notes})
    return {"message": f"Solicitud {resolution.status.value}"}

# ========================= EXCHANGE RATE ROUTES =========================
@api_router.get("/exchange-rates")
async def get_exchange_rates(current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    rates = await db.exchange_rates.find({}, {"_id": 0}).to_list(1000)
    return rates

@api_router.post("/exchange-rates")
async def create_exchange_rate(date_str: str, rate: float, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    existing = await db.exchange_rates.find_one({"date": date_str}, {"_id": 0})
    if existing:
        await db.exchange_rates.update_one({"date": date_str}, {"$set": {"rate": rate}})
        await log_audit(current_user, "UPDATE", "exchange_rates", date_str, {"rate": rate})
        return {"message": "Tipo de cambio actualizado"}
    
    exchange = ExchangeRate(date=date_str, rate=rate)
    doc = exchange.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.exchange_rates.insert_one(doc)
    await log_audit(current_user, "CREATE", "exchange_rates", date_str, {"rate": rate})
    return {"message": "Tipo de cambio creado"}

# ========================= MOVEMENT ROUTES =========================
@api_router.get("/movements")
async def get_movements(
    project_id: Optional[str] = None,
    partida_codigo: Optional[str] = None,
    provider_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_MOVEMENTS))
):
    query = {}
    if project_id:
        query["project_id"] = project_id
    if partida_codigo:
        query["partida_codigo"] = partida_codigo
    if provider_id:
        query["provider_id"] = provider_id
    if status:
        query["status"] = status
    
    movements = await db.movements.find(movement_active_query(extra=query), {"_id": 0}).to_list(5000)
    if year:
        validate_year_in_range(year)
    if year or month:
        filtered = []
        for m in movements:
            mov_date = date_parser.parse(m['date']) if isinstance(m['date'], str) else m['date']
            if year and mov_date.year != year:
                continue
            if month and mov_date.month != month:
                continue
            filtered.append(m)
        movements = filtered

    for m in movements:
        if m.get("amount_mxn") is None:
            amount_original = money_dec(m.get("amount_original") or 0)
            rate = money_dec(m.get("exchange_rate") or 1)
            computed = (amount_original * rate).quantize(TWO_DECIMALS)
            m["amount_mxn"] = float(computed)

    return movements

@api_router.post("/movements")
async def create_movement(movement_data: MovementCreate, current_user: dict = Depends(require_permission(Permission.CREATE_MOVEMENT))):
    # VALIDACIÓN BLOQUEANTE: partida debe existir en catálogo
    await validate_partida(movement_data.partida_codigo)
    enforce_capture_budget_scope(current_user, movement_data.partida_codigo)
    if enforce_oc_for_finanzas_egress_enabled() and current_user.get("role") == UserRole.FINANZAS.value and not is_ingresos_partida(movement_data.partida_codigo):
        raise HTTPException(status_code=403, detail={"code": "egress_manual_forbidden", "message": "FINANZAS no puede capturar egresos manuales; use Órdenes de Compra"})
    if current_user.get("role") == UserRole.CAPTURA_INGRESOS.value and not is_ingresos_partida(movement_data.partida_codigo):
        raise HTTPException(status_code=403, detail={"code": "captura_ingresos_only_402_403", "message": "CAPTURA_INGRESOS solo puede registrar ingresos 402/403"})

    customer_name = normalize_customer_name(movement_data.customer_name)
    no_provider_flow = str(movement_data.partida_codigo) in NO_PROVIDER_BUDGET_CODES
    
    # Validate references
    project = await db.projects.find_one({"id": movement_data.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=400, detail="Proyecto no válido")
    
    movement_data.provider_id = to_optional_str(movement_data.provider_id)
    movement_data.client_id = to_optional_str(movement_data.client_id)
    movement_data.reference = (movement_data.reference or "").strip()

    client_doc = None
    if no_provider_flow:
        if movement_data.provider_id is not None:
            raise HTTPException(status_code=422, detail={"code": "provider_not_allowed_for_abono", "message": "Las partidas 402/403 no aceptan proveedor"})
        if not movement_data.client_id:
            raise HTTPException(status_code=422, detail={"code": "client_required_for_partida_402_403", "message": "client_id es obligatorio para partidas 402/403"})
        client_doc = await db.clients.find_one({"id": movement_data.client_id}, {"_id": 0})
        if not client_doc:
            raise HTTPException(status_code=422, detail={"code": "client_not_found", "message": "Cliente no válido"})
        if client_doc.get("project_id") and client_doc.get("project_id") != movement_data.project_id:
            raise HTTPException(status_code=422, detail={"code": "client_project_mismatch", "message": "El proyecto del movimiento no coincide con el proyecto del cliente"})
        inventory_item = None
        inventory_reference = None
        if client_doc.get("inventory_item_id"):
            inventory_item = await db.inventory_items.find_one({"id": client_doc.get("inventory_item_id")}, {"_id": 0})
            if inventory_item:
                inventory_reference = resolve_inventory_reference(inventory_item) or client_doc.get("inventory_item_id")
        enforce_company_access(current_user, client_doc.get("company_id"))
        customer_name = normalize_customer_name(client_doc.get("nombre"))
        if not customer_name:
            raise HTTPException(status_code=422, detail={"code": "client_name_required", "message": "Cliente sin nombre válido"})
        if inventory_reference:
            movement_data.reference = inventory_reference
        elif not movement_data.reference:
            movement_data.reference = f"ABONO-{client_doc.get('id')}"
        provider = None
    else:
        if not movement_data.provider_id:
            raise HTTPException(status_code=422, detail={"code": "provider_required", "message": "provider_id es obligatorio para partidas distintas a 402/403"})
        provider = await db.providers.find_one({"id": movement_data.provider_id}, {"_id": 0})
        if not provider:
            raise HTTPException(status_code=422, detail={"code": "provider_not_found", "message": "Proveedor no válido"})
    
    amount_original_dec = money_dec(movement_data.amount_original)
    exchange_rate_dec = decimal_from_value(movement_data.exchange_rate, "exchange_rate").quantize(Decimal("0.0001"))
    if amount_original_dec <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_amount", "message": "Monto debe ser mayor a 0"})
    if movement_data.currency != Currency.MXN and exchange_rate_dec <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_exchange_rate", "message": "exchange_rate debe ser mayor a 0 para moneda distinta a MXN"})
    if movement_data.currency == Currency.MXN:
        exchange_rate_dec = Decimal("1")

    if project:
        enforce_company_access(current_user, project.get("empresa_id"))
    
    # Parse date
    parsed_date = parse_date_tijuana(movement_data.date)
    validate_date_in_range(parsed_date)
    amount_mxn_dec = (amount_original_dec * exchange_rate_dec).quantize(TWO_DECIMALS)

    if no_provider_flow and movement_data.client_id:
        await validate_client_abono_limit(movement_data.client_id, amount_mxn_dec)
    
    # Check for duplicates
    dup_check = await db.movements.find_one({
        "date": parsed_date.isoformat(),
        "provider_id": movement_data.provider_id if not no_provider_flow else None,
        "amount_original": movement_data.amount_original,
        "reference": movement_data.reference
    }, {"_id": 0})
    
    if dup_check:
        raise HTTPException(status_code=422, detail={"code": "duplicate_movement", "message": "Movimiento duplicado detectado"})
    
    # Check budget status
    year = parsed_date.year
    month = parsed_date.month
    
    budget_enforced = requires_budget(movement_data.partida_codigo)
    requires_auth = False
    auth_reason = ""
    overbudget_context = None
    if budget_enforced:
        budget_plan = await db.budget_plans.find_one({"project_id": movement_data.project_id, "partida_codigo": movement_data.partida_codigo}, {"_id": 0})
        overbudget = await evaluate_overbudget(movement_data.project_id, movement_data.partida_codigo, parsed_date, amount_mxn_dec) if budget_plan else None
        if overbudget:
            overbudget_meta = overbudget.get("metadata", {})
            overbudget_meta.setdefault("project_id", movement_data.project_id)
            overbudget_meta.setdefault("partida_code", movement_data.partida_codigo)
            overbudget_meta.setdefault("requested", str(amount_mxn_dec))
            overbudget_meta.setdefault("year", parsed_date.year)
            overbudget_meta.setdefault("month", parsed_date.month)
            if overbudget.get("total_exceeded"):
                overbudget_meta["scope"] = "total"
            elif overbudget.get("scope_exceeded"):
                overbudget_meta["scope"] = overbudget_meta.get("scope") or overbudget_meta.get("effective_scope") or "total"

            is_admin = current_user.get("role") == UserRole.ADMIN.value
            admin_bypass_allowed = is_admin and get_overbudget_admin_bypass_enabled() and not overbudget.get("total_exceeded")

            if admin_bypass_allowed:
                await log_audit(current_user, "OVERBUDGET_ADMIN_BYPASS", "movements", "pending_create", {
                    "code": "overbudget_admin_bypass",
                    "project_id": movement_data.project_id,
                    "partida_code": movement_data.partida_codigo,
                    "meta": overbudget_meta,
                })
            else:
                mode = get_overbudget_approval_mode()
                if mode in {"reject", "reject_and_request"}:
                    if mode == "reject_and_request":
                        await ensure_pending_approval(
                            approval_type=ApprovalType.OVERBUDGET_EXCEPTION,
                            company_id=project.get("empresa_id"),
                            project_id=movement_data.project_id,
                            budget_id=overbudget["plan"].get("id"),
                            requested_by=current_user["user_id"],
                            dedupe_key=f"overbudget:{movement_data.project_id}:{movement_data.partida_codigo}:{parsed_date.date().isoformat()}:{str(amount_mxn_dec)}:{','.join(overbudget_meta.get('triggered_buckets', []))}",
                            metadata=overbudget_meta,
                        )
                    raise HTTPException(status_code=422, detail={
                        "code": "overbudget_rejected_and_requested",
                        "message": "Movement exceeds available budget and an exception approval request was generated.",
                        "meta": {
                            "scope": overbudget_meta.get("scope") or "total",
                            "project_id": movement_data.project_id,
                            "partida_code": movement_data.partida_codigo,
                            "requested": str(amount_mxn_dec),
                            "available": overbudget_meta.get("available", "0"),
                            "year": parsed_date.year,
                            "month": parsed_date.month,
                        },
                    })
                elif mode == "pending_movement":
                    requires_auth = True
                    auth_reason = "Movimiento excede presupuesto"
                    overbudget_context = overbudget
        elif not budget_plan:
            budget = await db.budgets.find_one({
                "project_id": movement_data.project_id,
                "partida_codigo": movement_data.partida_codigo,
                "year": year,
                "month": month
            }, {"_id": 0})
            if not budget:
                raise HTTPException(status_code=422, detail="Presupuesto requerido no definido para la partida y periodo")
            current_movements = await db.movements.find({
                "project_id": movement_data.project_id,
                "partida_codigo": movement_data.partida_codigo,
                "status": MovementStatus.POSTED.value
            }, {"_id": 0}).to_list(5000)
            current_spent = sum(
                decimal_from_value(m.get('amount_mxn', 0), 'amount_mxn') for m in current_movements
                if date_parser.parse(m['date']).year == year and date_parser.parse(m['date']).month == month
            )
            budget_amount = decimal_from_value(budget.get('amount_mxn', 0), 'amount_mxn')
            projected_available = (budget_amount - (current_spent + amount_mxn_dec)).quantize(TWO_DECIMALS)
            if projected_available < Decimal("0.00"):
                raise HTTPException(status_code=422, detail="Saldo de presupuesto insuficiente para la partida y periodo")
    
    movement = Movement(
        project_id=movement_data.project_id,
        partida_codigo=movement_data.partida_codigo,
        provider_id=movement_data.provider_id if not no_provider_flow else None,
        customer_name=customer_name,
        date=parsed_date,
        currency=movement_data.currency,
        amount_original=movement_data.amount_original,
        exchange_rate=movement_data.exchange_rate,
        amount_mxn=float(amount_mxn_dec),
        reference=movement_data.reference,
        client_id=movement_data.client_id if no_provider_flow else None,
        description=movement_data.description,
        created_by=current_user["user_id"],
        status=MovementStatus.PENDING_APPROVAL if requires_auth else MovementStatus.POSTED
    )
    
    doc = movement.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    # Create authorization record if needed
    if requires_auth:
        auth = Authorization(
            movement_id=movement.id,
            reason=auth_reason,
            requested_by=current_user["user_id"]
        )
        auth_doc = auth.model_dump()
        auth_doc['created_at'] = auth_doc['created_at'].isoformat()
        # Add budget context for UI display
        auth_doc['budget_context'] = {
            'partida_codigo': movement_data.partida_codigo,
            'presupuesto': budget_amount,
            'ejecutado_actual': current_spent,
            'monto_movimiento': float(amount_mxn_dec),
            'porcentaje_actual': (current_spent / budget_amount * 100) if budget_amount > 0 else 0,
            'porcentaje_si_aprueba': percentage_if_posted
        }
        await db.authorizations.insert_one(auth_doc)
        doc['authorization_id'] = auth.id
    
    await db.movements.insert_one(doc)

    receipt_url = None
    if no_provider_flow and client_doc and not client_doc.get("inventory_item_id"):
        await log_audit(current_user, "MOVEMENT_402_403_MANUAL_REFERENCE", "movements", movement.id, {
            "movement_id": movement.id,
            "client_id": client_doc.get("id"),
            "message": "Cliente sin inventario ligado; se conserva referencia manual para abono 402/403",
            "reference": movement_data.reference,
        })
    if no_provider_flow and client_doc and movement_counts_as_abono_doc(doc):
        updated_client = await recalc_client_financials(client_doc["id"])
        await log_audit(current_user, "MOVEMENT_402_403_BALANCE", "clients", client_doc["id"], {
            "movement_id": movement.id,
            "client_id": client_doc["id"],
            "inventory_item_id": client_doc.get("inventory_item_id"),
            "abonos_total_mxn": updated_client.get("abonos_total_mxn", 0) if updated_client else 0,
            "saldo_restante": updated_client.get("saldo_restante", 0) if updated_client else 0,
        })
        receipt_url = f"/api/movements/{movement.id}/receipt.pdf"

    # Remove MongoDB _id before returning
    doc.pop('_id', None)
    
    await log_audit(current_user, "CREATE", "movements", movement.id, {"data": doc, "requires_auth": requires_auth, "receipt_url": receipt_url})
    
    return {"movement": sanitize_mongo_document(doc), "requires_authorization": requires_auth, "reason": auth_reason if requires_auth else None, "receipt_url": receipt_url}

@api_router.post("/movements/import-csv")
async def import_movements_csv(file: UploadFile = File(...), current_user: dict = Depends(require_permission(Permission.IMPORT_MOVEMENTS))):
    """
    Import CSV con validaciones estrictas según especificación Entrega B.
    Columnas requeridas: fecha, empresa, proyecto, partida, proveedor, moneda, monto, tipo_cambio, referencia, descripcion
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos CSV")
    
    # Log inicio de import
    import_start = datetime.now(timezone.utc)
    
    content = await file.read()
    # Try different encodings
    try:
        decoded = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            decoded = content.decode('latin-1')
        except UnicodeDecodeError:
            decoded = content.decode('utf-8', errors='replace')
    
    reader = csv.DictReader(io.StringIO(decoded))
    
    results = CSVImportResult(
        total_filas=0,
        insertadas=0,
        rechazadas=0,
        duplicadas_omitidas=0,
        errores=[],
        movements_created=[],
        authorizations_required=[]
    )
    
    # Columnas requeridas exactas
    required_columns = ['fecha', 'empresa', 'proyecto', 'partida', 'proveedor', 'moneda', 'monto', 'referencia']
    
    # Get lookups - empresas, proyectos, partidas del catálogo, proveedores
    empresas = {e['nombre'].upper(): e for e in await db.empresas.find({}, {"_id": 0}).to_list(1000)}
    empresas_by_id = {e['id']: e for e in await db.empresas.find({}, {"_id": 0}).to_list(1000)}
    
    projects_list = await db.projects.find({}, {"_id": 0}).to_list(1000)
    # Map by code AND name (upper) for flexibility
    projects_by_code = {p['code'].upper(): p for p in projects_list}
    projects_by_name = {p['name'].upper(): p for p in projects_list}
    
    # Catálogo de partidas (source of truth)
    catalogo = await db.catalogo_partidas.find({}, {"_id": 0}).to_list(1000)
    partidas_by_codigo = {p['codigo']: p for p in catalogo}
    partidas_by_nombre = {p['nombre'].upper(): p for p in catalogo}
    
    providers_list = await db.providers.find({}, {"_id": 0}).to_list(1000)
    providers_by_code = {p['code'].upper(): p for p in providers_list}
    providers_by_name = {p['name'].upper(): p for p in providers_list}
    
    row_num = 1
    rows_processed = []
    
    for row in reader:
        row_num += 1
        results.total_filas += 1
        errors = []
        
        # Normalize column names (lowercase, strip)
        row = {k.lower().strip(): v.strip() if v else '' for k, v in row.items()}
        
        # Validate required columns exist
        for col in required_columns:
            if col not in row or not row.get(col, '').strip():
                errors.append({"columna": col, "motivo": f"Campo '{col}' requerido y vacío"})
        
        if errors:
            results.errores.append({"fila": row_num, "errores": errors})
            results.rechazadas += 1
            continue
        
        # 1) VALIDAR EMPRESA existe y activa
        empresa_val = row.get('empresa', '').strip().upper()
        empresa = empresas.get(empresa_val)
        if not empresa:
            errors.append({"columna": "empresa", "motivo": f"Empresa '{row.get('empresa')}' no existe"})
        elif not empresa.get('is_active', True):
            errors.append({"columna": "empresa", "motivo": f"Empresa '{row.get('empresa')}' está inactiva"})
        
        # 2) VALIDAR PROYECTO existe, activo y pertenece a empresa
        proyecto_val = row.get('proyecto', '').strip().upper()
        project = projects_by_code.get(proyecto_val) or projects_by_name.get(proyecto_val)
        if not project:
            errors.append({"columna": "proyecto", "motivo": f"Proyecto '{row.get('proyecto')}' no existe"})
        elif not project.get('is_active', True):
            errors.append({"columna": "proyecto", "motivo": f"Proyecto '{row.get('proyecto')}' está inactivo"})
        elif empresa and project.get('empresa_id') != empresa.get('id'):
            empresa_proyecto = empresas_by_id.get(project.get('empresa_id'), {})
            errors.append({"columna": "proyecto", "motivo": f"Proyecto '{row.get('proyecto')}' no pertenece a empresa '{row.get('empresa')}' (pertenece a '{empresa_proyecto.get('nombre', 'N/A')}')"})
        
        # 3) VALIDAR PARTIDA existe y activa en catalogo_partidas
        partida_val = row.get('partida', '').strip()
        partida = partidas_by_codigo.get(partida_val) or partidas_by_nombre.get(partida_val.upper())
        if not partida:
            errors.append({"columna": "partida", "motivo": f"Partida '{row.get('partida')}' no existe en catálogo"})
        elif not partida.get('is_active', True):
            errors.append({"columna": "partida", "motivo": f"Partida '{row.get('partida')}' está inactiva"})
        
        # 4) VALIDAR PROVEEDOR existe
        proveedor_val = row.get('proveedor', '').strip().upper()
        provider = providers_by_code.get(proveedor_val) or providers_by_name.get(proveedor_val)
        if not provider:
            errors.append({"columna": "proveedor", "motivo": f"Proveedor '{row.get('proveedor')}' no existe"})
        elif not provider.get('is_active', True):
            errors.append({"columna": "proveedor", "motivo": f"Proveedor '{row.get('proveedor')}' está inactivo"})
        
        # 5) VALIDAR MONTO > 0
        try:
            monto = float(row.get('monto', '0').replace(',', '').replace('$', '').strip())
            if monto <= 0:
                errors.append({"columna": "monto", "motivo": "Monto debe ser mayor a 0"})
        except ValueError:
            errors.append({"columna": "monto", "motivo": f"Monto '{row.get('monto')}' no es un número válido"})
            monto = 0
        
        # 6) VALIDAR MONEDA ∈ {MXN, USD}
        moneda = row.get('moneda', '').strip().upper()
        if moneda not in ['MXN', 'USD']:
            errors.append({"columna": "moneda", "motivo": f"Moneda '{row.get('moneda')}' no válida (solo MXN o USD)"})
        
        # 7) VALIDAR TIPO_CAMBIO si USD
        tipo_cambio_str = row.get('tipo_cambio', '').strip()
        tipo_cambio = 1.0
        if moneda == 'USD':
            if not tipo_cambio_str:
                errors.append({"columna": "tipo_cambio", "motivo": "USD requiere tipo_cambio obligatorio"})
            else:
                try:
                    tipo_cambio = float(tipo_cambio_str.replace(',', ''))
                    if tipo_cambio <= 0:
                        errors.append({"columna": "tipo_cambio", "motivo": "tipo_cambio debe ser mayor a 0"})
                except ValueError:
                    errors.append({"columna": "tipo_cambio", "motivo": f"tipo_cambio '{tipo_cambio_str}' no es válido"})
        elif moneda == 'MXN':
            tipo_cambio = 1.0
        
        # 8) VALIDAR FECHA - interpretar como America/Tijuana si no tiene hora
        fecha_str = row.get('fecha', '').strip()
        parsed_date = None
        try:
            parsed_date = parse_date_tijuana(fecha_str)
        except Exception as e:
            errors.append({"columna": "fecha", "motivo": f"Fecha '{fecha_str}' no válida (use formato YYYY-MM-DD o DD/MM/YYYY)"})
        
        # Si hay errores de validación, rechazar fila
        if errors:
            results.errores.append({"fila": row_num, "errores": errors})
            results.rechazadas += 1
            continue
        
        # 9) CHECK DUPLICADOS: fecha+empresa+proyecto+proveedor+monto+referencia
        referencia = row.get('referencia', '').strip()
        dup_query = {
            "date": parsed_date.isoformat(),
            "project_id": project['id'],
            "provider_id": provider['id'],
            "amount_original": monto,
            "reference": referencia
        }
        
        # También verificar si ya procesamos en este batch
        dup_key = f"{parsed_date.isoformat()}|{empresa['id']}|{project['id']}|{provider['id']}|{monto}|{referencia}"
        if dup_key in rows_processed:
            results.errores.append({"fila": row_num, "errores": [{"columna": "duplicado", "motivo": "Duplicado en el mismo archivo CSV"}]})
            results.duplicadas_omitidas += 1
            continue
        
        dup_check = await db.movements.find_one(dup_query, {"_id": 0})
        if dup_check:
            results.duplicadas_omitidas += 1
            continue  # Omitir silenciosamente como duplicado
        
        rows_processed.append(dup_key)
        
        # CALCULAR monto_mxn
        monto_mxn = monto * tipo_cambio
        
        # CHECK BUDGET para determinar si requiere autorización
        year = parsed_date.year
        month = parsed_date.month
        budget_enforced = requires_budget(partida['codigo'])
        current_spent = 0
        budget_amount = 0
        requires_auth = False
        auth_reason = ""
        percentage_if_posted = 0
        if budget_enforced:
            budget = await db.budgets.find_one({
                "project_id": project['id'],
                "partida_codigo": partida['codigo'],
                "year": year,
                "month": month
            }, {"_id": 0})

            current_movements = await db.movements.find({
                "project_id": project['id'],
                "partida_codigo": partida['codigo'],
                "status": MovementStatus.POSTED.value
            }, {"_id": 0}).to_list(5000)

            current_spent = sum(
                m['amount_mxn'] for m in current_movements
                if date_parser.parse(m['date']).year == year and date_parser.parse(m['date']).month == month
            )

            budget_amount = budget['amount_mxn'] if budget else 0
            new_total = current_spent + monto_mxn
            percentage_if_posted = (new_total / budget_amount * 100) if budget_amount > 0 else 0

            if budget_amount == 0:
                errors.append({"columna": "presupuesto", "motivo": "Presupuesto requerido no definido para la partida y periodo"})
            elif new_total > budget_amount:
                errors.append({"columna": "presupuesto", "motivo": "Saldo de presupuesto insuficiente para la partida y periodo"})

        if errors:
            results.errores.append({"fila": row_num, "errores": errors})
            results.rechazadas += 1
            continue
        
        # CREATE MOVEMENT
        descripcion = row.get('descripcion', '').strip() if row.get('descripcion') else None
        
        movement = Movement(
            project_id=project['id'],
            partida_codigo=partida['codigo'],
            provider_id=provider['id'],
            date=parsed_date,
            currency=Currency(moneda),
            amount_original=monto,
            exchange_rate=tipo_cambio,
            amount_mxn=monto_mxn,
            reference=referencia,
            description=descripcion,
            created_by=current_user["user_id"],
            status=MovementStatus.PENDING_APPROVAL if requires_auth else MovementStatus.POSTED
        )
        
        doc = movement.model_dump()
        doc['date'] = doc['date'].isoformat()
        doc['created_at'] = doc['created_at'].isoformat()
        
        if requires_auth:
            auth = Authorization(
                movement_id=movement.id,
                reason=auth_reason,
                requested_by=current_user["user_id"]
            )
            auth_doc = auth.model_dump()
            auth_doc['created_at'] = auth_doc['created_at'].isoformat()
            # Add budget context
            auth_doc['budget_context'] = {
                'partida_codigo': partida['codigo'],
                'presupuesto': budget_amount,
                'ejecutado_actual': current_spent,
                'monto_movimiento': monto_mxn,
                'porcentaje_actual': (current_spent / budget_amount * 100) if budget_amount > 0 else 0,
                'porcentaje_si_aprueba': percentage_if_posted
            }
            await db.authorizations.insert_one(auth_doc)
            doc['authorization_id'] = auth.id
            results.authorizations_required.append(movement.id)
        
        await db.movements.insert_one(doc)
        results.movements_created.append(movement.id)
        results.insertadas += 1
    
    # LOG AUDIT de import
    import_end = datetime.now(timezone.utc)
    errores_resumen = [f"Fila {e['fila']}: {'; '.join([err['motivo'] for err in e['errores']])}" for e in results.errores[:20]]
    
    import_audit = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["user_id"],
        "user_email": current_user["email"],
        "action": "IMPORT_CSV",
        "timestamp_inicio": import_start.isoformat(),
        "timestamp_fin": import_end.isoformat(),
        "filtros": {"filename": file.filename},
        "conteos": {
            "total_filas": results.total_filas,
            "insertadas": results.insertadas,
            "rechazadas": results.rechazadas,
            "duplicadas_omitidas": results.duplicadas_omitidas
        },
        "errores_resumen": errores_resumen
    }
    await db.import_export_logs.insert_one(import_audit)
    
    await log_audit(current_user, "IMPORT_CSV", "movements", "batch", {
        "total_filas": results.total_filas,
        "insertadas": results.insertadas,
        "rechazadas": results.rechazadas,
        "duplicadas_omitidas": results.duplicadas_omitidas,
        "filename": file.filename
    })
    
    return results

# Keep legacy endpoint for backwards compatibility
@api_router.post("/movements/import")
async def import_movements_legacy(file: UploadFile = File(...), current_user: dict = Depends(require_permission(Permission.IMPORT_MOVEMENTS))):
    """Legacy import - redirects to new import-csv endpoint"""
    return await import_movements_csv(file, current_user)

# ========================= AUTHORIZATION ROUTES =========================
@api_router.get("/authorizations")
async def get_authorizations(
    status: Optional[str] = None,
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_AUTHORIZATIONS))
):
    """Get authorizations with filters and enriched movement data"""
    query = {}
    if year:
        validate_year_in_range(year)
    if status:
        query["status"] = status
    
    auths = await db.authorizations.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Enrich with movement details
    project_docs = await db.projects.find({}, {"_id": 0}).to_list(1000)
    project_map = {p['id']: p for p in project_docs}
    
    empresa_docs = await db.empresas.find({}, {"_id": 0}).to_list(1000)
    empresa_map = {e['id']: e for e in empresa_docs}
    
    provider_docs = await db.providers.find({}, {"_id": 0}).to_list(1000)
    provider_map = {p['id']: p for p in provider_docs}
    
    partida_docs = await db.catalogo_partidas.find({}, {"_id": 0}).to_list(1000)
    partida_map = {p['codigo']: p for p in partida_docs}
    
    enriched_auths = []
    for auth in auths:
        if auth.get("approval_type") == ApprovalType.PURCHASE_ORDER_WORKFLOW.value and auth.get("purchase_order_id"):
            po = await db.purchase_orders.find_one({"id": auth.get("purchase_order_id")}, {"_id": 0})
            if not po:
                continue
            if empresa_id and po.get("company_id") != empresa_id:
                continue
            if project_id and po.get("project_id") != project_id:
                continue
            po_date = date_parser.parse(po.get("order_date")) if po.get("order_date") else None
            if year and po_date and po_date.year != year:
                continue
            if month and po_date and po_date.month != month:
                continue
            partidas = sorted({str((line or {}).get("partida_codigo") or "") for line in (po.get("lines") or []) if (line or {}).get("partida_codigo")})
            auth["movement_details"] = {
                "date": po.get("order_date"),
                "empresa_id": po.get("company_id"),
                "empresa_nombre": empresa_map.get(po.get("company_id"), {}).get("nombre", "N/A"),
                "project_id": po.get("project_id"),
                "project_code": project_map.get(po.get("project_id"), {}).get("code", "N/A"),
                "project_name": project_map.get(po.get("project_id"), {}).get("name", "N/A"),
                "partida_codigo": ", ".join(partidas) if partidas else None,
                "partida_nombre": "Partidas OC",
                "provider_name": po.get("vendor_name") or "N/A",
                "provider_rfc": po.get("vendor_rfc") or "N/A",
                "moneda": po.get("currency"),
                "monto_original": po.get("total"),
                "tipo_cambio": po.get("exchange_rate"),
                "monto_mxn": float(po.get("total", 0) or 0),
                "referencia": po.get("invoice_folio") or po.get("folio") or po.get("external_id"),
                "descripcion": f"Orden de Compra {po.get('folio') or po.get('external_id')}",
            }
            partial_value = auth.get("partial_amount") or po.get("pending_amount")
            budget_summary = await build_purchase_order_budget_summary(po, money_dec(partial_value) if partial_value is not None else None)
            auth["purchase_order_details"] = {
                "id": po.get("id"),
                "folio": canonicalize_oc_folio(po.get("folio") or po.get("external_id")),
                "invoice_folio": po.get("invoice_folio"),
                "status": po.get("status"),
                "empresa_id": po.get("company_id"),
                "empresa_nombre": empresa_map.get(po.get("company_id"), {}).get("nombre"),
                "proyecto_id": po.get("project_id"),
                "proyecto_nombre": project_map.get(po.get("project_id"), {}).get("name"),
                "project_id": po.get("project_id"),
                "project": project_map.get(po.get("project_id"), {}).get("name"),
                "project_code": project_map.get(po.get("project_id"), {}).get("code"),
                "company_id": po.get("company_id"),
                "company": empresa_map.get(po.get("company_id"), {}).get("nombre"),
                "proveedor_nombre": po.get("vendor_name"),
                "proveedor_rfc": po.get("vendor_rfc"),
                "vendor_name": po.get("vendor_name"),
                "vendor_rfc": po.get("vendor_rfc"),
                "subtotal": po.get("subtotal_tax_base"),
                "iva_total": po.get("tax_total"),
                "ret_isr_total": po.get("withholding_isr_total"),
                "total": po.get("total"),
                "totals": {
                    "subtotal": po.get("subtotal_tax_base"),
                    "iva": po.get("tax_total"),
                    "ret_isr": po.get("withholding_isr_total"),
                    "total": po.get("total"),
                },
                "approved_amount_total": po.get("approved_amount_total") or "0.00",
                "pending_amount": po.get("pending_amount") or po.get("total"),
            }
            auth["budget_gate_summary"] = budget_summary
            auth["budget_preview"] = budget_summary
            auth["partial_approved_accumulated"] = po.get("approved_amount_total") or "0.00"
            auth["pending_amount_oc"] = po.get("pending_amount") or po.get("total")
            auth["next_estimated_impact"] = budget_summary.get("restante_proyectado_si_aprueba")
            enriched_auths.append(sanitize_mongo_document(auth))
            continue

        movement = None
        if auth.get('movement_id'):
            movement = await db.movements.find_one({"id": auth['movement_id']}, {"_id": 0})
        
        if movement:
            project = project_map.get(movement.get('project_id'), {})
            empresa = empresa_map.get(project.get('empresa_id'), {})
            provider = provider_map.get(movement.get('provider_id'), {})
            partida = partida_map.get(movement.get('partida_codigo'), {})
            
            # Apply filters
            if empresa_id and empresa.get('id') != empresa_id:
                continue
            if project_id and project.get('id') != project_id:
                continue
            
            mov_date = date_parser.parse(movement['date'])
            if year and mov_date.year != year:
                continue
            if month and mov_date.month != month:
                continue
            
            auth['movement_details'] = {
                'date': movement.get('date'),
                'empresa_id': empresa.get('id'),
                'empresa_nombre': empresa.get('nombre', 'N/A'),
                'project_id': project.get('id'),
                'project_code': project.get('code', 'N/A'),
                'project_name': project.get('name', 'N/A'),
                'partida_codigo': movement.get('partida_codigo'),
                'partida_nombre': partida.get('nombre', 'N/A'),
                'provider_name': provider.get('name', 'N/A'),
                'moneda': movement.get('currency'),
                'monto_original': movement.get('amount_original'),
                'tipo_cambio': movement.get('exchange_rate'),
                'monto_mxn': movement.get('amount_mxn'),
                'referencia': movement.get('reference'),
                'descripcion': movement.get('description')
            }
        else:
            # Skip if no movement and filters are applied
            if empresa_id or project_id or year or month:
                continue
            auth['movement_details'] = None
        
        enriched_auths.append(auth)
    
    return enriched_auths

@api_router.get("/authorizations/pending-summary")
async def get_pending_summary(
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get summary of pending authorizations for dashboard KPI"""
    now = to_tijuana(datetime.now(timezone.utc))
    year = year or now.year
    validate_year_in_range(year)
    month = month or now.month
    
    # Get pending movements
    pending_query = {"status": MovementStatus.PENDING_APPROVAL.value}
    pending_movements = await db.movements.find(movement_active_query(extra=pending_query), {"_id": 0}).to_list(5000)
    
    # Filter by empresa/project/date
    project_docs = await db.projects.find({}, {"_id": 0}).to_list(1000)
    project_map = {p['id']: p for p in project_docs}
    
    filtered_pending = []
    for mov in pending_movements:
        project = project_map.get(mov.get('project_id'), {})
        
        if empresa_id and project.get('empresa_id') != empresa_id:
            continue
        if project_id and mov.get('project_id') != project_id:
            continue
        
        mov_date = date_parser.parse(mov['date'])
        if mov_date.year != year or mov_date.month != month:
            continue
        
        filtered_pending.append(mov)
    
    total_pending_mxn = sum(m.get('amount_mxn', 0) for m in filtered_pending)
    
    return {
        "pending_count": len(filtered_pending),
        "pending_total_mxn": total_pending_mxn,
        "year": year,
        "month": month
    }

@api_router.put("/authorizations/{auth_id}")
async def resolve_authorization(
    auth_id: str,
    resolution: AuthorizationResolve,
    current_user: dict = Depends(require_permission(Permission.APPROVE_REJECT))
):
    auth = await db.authorizations.find_one({"id": auth_id}, {"_id": 0})
    if not auth:
        raise HTTPException(status_code=404, detail=structured_error("authorization_not_found", "Autorización no encontrada"))

    if auth.get('status') != 'pending':
        raise HTTPException(status_code=409, detail=structured_error("authorization_already_resolved", "Autorización ya resuelta", {"status": auth.get("status")}))

    # Reject requires notes/motivo
    if resolution.status == AuthorizationStatus.REJECTED and not (resolution.notes or "").strip():
        raise HTTPException(status_code=422, detail=structured_error("notes_required_for_rejection", "Rechazo requiere motivo/notas"))

    if auth.get("approval_type") == ApprovalType.PURCHASE_ORDER_WORKFLOW.value and auth.get("purchase_order_id"):
        po = await db.purchase_orders.find_one({"id": auth.get("purchase_order_id")}, {"_id": 0})
        if not po:
            raise HTTPException(status_code=404, detail=structured_error("purchase_order_not_found", "OC no encontrada"))
        enforce_company_access(current_user, po.get("company_id"))

        po_total = money_dec(po.get("total", 0))
        already_approved = money_dec(po.get("approved_amount_total", 0))
        pending_amount = (po_total - already_approved).quantize(TWO_DECIMALS)

        if resolution.status == AuthorizationStatus.REJECTED:
            await db.authorizations.update_one({"id": auth_id}, {"$set": {
                "status": AuthorizationStatus.REJECTED.value,
                "resolved_at": datetime.now(timezone.utc).isoformat(),
                "resolved_by": current_user["user_id"],
                "resolved_by_email": current_user["email"],
                "notes": resolution.notes,
            }})
            await db.purchase_orders.update_one({"id": po.get("id")}, {"$set": {"status": PurchaseOrderStatus.REJECTED.value, "rejection_reason": resolution.notes or "", "updated_at": datetime.now(timezone.utc).isoformat()}})
            await log_audit(current_user, "AUTH_REJECTED", "purchase_orders", po.get("id"), {"reason": resolution.notes})
            return {"message": "Autorización rejected"}

        partial_amount = money_dec(resolution.partial_amount) if resolution.partial_amount is not None else pending_amount
        if partial_amount <= 0:
            raise HTTPException(status_code=422, detail=structured_error("invalid_partial_amount", "Monto parcial debe ser mayor a 0", {"pending_amount": str(pending_amount)}))
        if partial_amount > pending_amount:
            raise HTTPException(status_code=422, detail=structured_error("partial_amount_exceeds_pending", "Monto parcial excede saldo pendiente", {"pending_amount": str(pending_amount), "partial_amount": str(partial_amount)}))

        approve_event_id = str(uuid.uuid4())
        ratio = (partial_amount / po_total) if po_total > 0 else Decimal("0")
        created_movements = []
        for line in po.get("lines", []):
            line_total = money_dec(line.get("line_total", 0))
            line_partial = (line_total * ratio).quantize(TWO_DECIMALS)
            if line_partial <= 0:
                continue
            origin_event = f"OC_APPROVE_PARTIAL:{approve_event_id}"
            existing_mv = await db.movements.find_one({"purchase_order_line_id": line.get("id"), "origin_event": origin_event}, {"_id": 0})
            if existing_mv:
                created_movements.append(existing_mv.get("id"))
                continue
            movement_doc = {
                "id": str(uuid.uuid4()),
                "project_id": po.get("project_id"),
                "partida_codigo": line.get("partida_codigo"),
                "provider_id": None,
                "date": po.get("order_date"),
                "currency": po.get("currency"),
                "amount_original": float(line_partial),
                "exchange_rate": float(decimal_from_value(po.get("exchange_rate", 1), "exchange_rate")),
                "amount_mxn": float(line_partial),
                "reference": po.get("invoice_folio") or po.get("folio") or po.get("external_id"),
                "description": f"OC parcial {po.get('folio') or po.get('external_id')} línea {line.get('line_no')}",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["user_id"],
                "status": MovementStatus.POSTED.value,
                "purchase_order_id": po.get("id"),
                "purchase_order_line_id": line.get("id"),
                "origin_event": origin_event,
                "is_active": True,
            }
            await db.movements.insert_one(movement_doc)
            created_movements.append(movement_doc["id"])

        new_approved_total = (already_approved + partial_amount).quantize(TWO_DECIMALS)
        new_pending = (po_total - new_approved_total).quantize(TWO_DECIMALS)
        new_status = PurchaseOrderStatus.APPROVED_FOR_PAYMENT.value if new_pending <= 0 else "partially_approved"
        posting_status = PostingStatus.POSTED.value if new_pending <= 0 else "partially_posted"

        await db.purchase_orders.update_one({"id": po.get("id")}, {"$set": {
            "approved_amount_total": str(new_approved_total),
            "pending_amount": str(new_pending if new_pending > 0 else Decimal("0.00")),
            "status": new_status,
            "posting_status": posting_status,
            "approved_by_user_id": current_user["user_id"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }})
        await db.authorizations.update_one({"id": auth_id}, {"$set": {
            "status": AuthorizationStatus.APPROVED.value,
            "resolved_at": datetime.now(timezone.utc).isoformat(),
            "resolved_by": current_user["user_id"],
            "resolved_by_email": current_user["email"],
            "notes": resolution.notes,
            "partial_amount": str(partial_amount),
            "approval_event_id": approve_event_id,
            "movement_ids": created_movements,
        }})
        if new_pending > 0:
            await ensure_pending_approval(
                approval_type=ApprovalType.PURCHASE_ORDER_WORKFLOW,
                company_id=po.get("company_id"),
                project_id=po.get("project_id"),
                requested_by=current_user["user_id"],
                purchase_order_id=po.get("id"),
                dedupe_key=f"purchase_order_workflow:{po.get('id')}:{str(new_pending)}",
                metadata={"folio": po.get("folio") or po.get("external_id"), "purchase_order_id": po.get("id"), "pending_amount": str(new_pending)},
            )

        updated_po = await db.purchase_orders.find_one({"id": po.get("id")}, {"_id": 0})
        budget_summary = await build_purchase_order_budget_summary(updated_po, new_pending if new_pending > 0 else Decimal("0.00"))
        await log_audit(current_user, "AUTH_APPROVED", "purchase_orders", po.get("id"), {"partial_amount": str(partial_amount), "approved_amount_total": str(new_approved_total), "pending_amount": str(new_pending), "movement_ids": created_movements})
        return {
            "message": "Autorización approved",
            "purchase_order_id": po.get("id"),
            "approved_amount": str(partial_amount),
            "approved_amount_total": str(new_approved_total),
            "pending_amount": str(new_pending if new_pending > 0 else Decimal("0.00")),
            "movement_ids": created_movements,
            "budget_gate_summary": budget_summary,
        }
    
    update_data = {
        "status": resolution.status.value,
        "resolved_at": datetime.now(timezone.utc).isoformat(),
        "resolved_by": current_user["user_id"],
        "resolved_by_email": current_user["email"],
        "notes": resolution.notes
    }
    
    await db.authorizations.update_one({"id": auth_id}, {"$set": update_data})
    
    # Update movement status: approved -> posted, rejected -> rejected
    movement = None
    if auth.get('movement_id'):
        new_status = MovementStatus.POSTED if resolution.status == AuthorizationStatus.APPROVED else MovementStatus.REJECTED
        await db.movements.update_one(
            {"id": auth['movement_id']},
            {"$set": {"status": new_status.value}}
        )
        movement = await db.movements.find_one({"id": auth['movement_id']}, {"_id": 0})
        if movement and movement_counts_as_abono_doc(movement):
            await recalc_client_financials(movement.get("client_id"))
    
    # Detailed audit log for approve/reject
    await log_audit(current_user, f"AUTH_{resolution.status.value.upper()}", "authorizations", auth_id, {
        "movement_id": auth.get('movement_id'),
        "resolution": resolution.status.value,
        "notes": resolution.notes,
        "amount_mxn": movement.get('amount_mxn') if movement else None,
        "partida_codigo": movement.get('partida_codigo') if movement else None,
        "budget_context": auth.get('budget_context')
    })
    
    return {"message": f"Autorización {resolution.status.value}"}

# ========================= REPORTS ROUTES =========================


def _to_period_decimal(value: Any) -> Decimal:
    try:
        return decimal_from_value(value or 0, "amount").quantize(TWO_DECIMALS)
    except HTTPException:
        return Decimal("0.00")


def _dashboard_decimal_to_float(value: Decimal) -> float:
    return float(value.quantize(TWO_DECIMALS))


def _month_keys_for_period(period: str, year: int, month: Optional[int], quarter: Optional[int]) -> List[str]:
    if period == "month":
        if month is None:
            raise HTTPException(status_code=422, detail={"code": "month_required", "message": "month es requerido cuando period=month"})
        if month < 1 or month > 12:
            raise HTTPException(status_code=422, detail={"code": "month_out_of_range", "message": "month debe estar entre 1 y 12"})
        return [f"{year:04d}-{month:02d}"]
    if period == "quarter":
        if quarter is None:
            raise HTTPException(status_code=422, detail={"code": "quarter_required", "message": "quarter es requerido cuando period=quarter"})
        if quarter < 1 or quarter > 4:
            raise HTTPException(status_code=422, detail={"code": "quarter_out_of_range", "message": "quarter debe estar entre 1 y 4"})
        start = (quarter - 1) * 3 + 1
        return [f"{year:04d}-{m:02d}" for m in range(start, start + 3)]
    return [f"{year:04d}-{m:02d}" for m in range(1, 13)]


def _match_dashboard_period(dt: datetime, period: str, year: int, month: Optional[int], quarter: Optional[int]) -> bool:
    if dt.year != year:
        return False
    if period == "month":
        return dt.month == month
    if period == "quarter":
        q = ((dt.month - 1) // 3) + 1
        return q == quarter
    return True


def _fmt_period_label(period: str, year: int, month: Optional[int], quarter: Optional[int]) -> str:
    if period == "all":
        return f"TODO {year}"
    if period == "year":
        return f"Anual {year}"
    if period == "quarter":
        return f"Q{quarter} {year}"
    return f"{year:04d}-{(month or 1):02d}"


def _derive_partida_group(code: str, partida_catalog: dict) -> str:
    part = partida_catalog.get(str(code), {}) or {}
    if part.get("grupo"):
        return str(part.get("grupo"))
    code_int = int(''.join(ch for ch in str(code) if ch.isdigit())[:3] or 0)
    if 100 <= code_int <= 199:
        return "COSTOS DIRECTOS"
    if 200 <= code_int <= 299:
        return "GASTOS VTA/ADM"
    if 300 <= code_int <= 399:
        return "GASTOS FINANCIEROS"
    if 400 <= code_int <= 499:
        return "INGRESOS"
    return "OTROS"


PL_DIRECT_COST_CODES = ["101", "102", "103", "104", "105", "106", "107", "108", "109", "110", "111"]
PL_SELLING_ADMIN_CODES = ["201", "202", "203", "204", "205", "206", "207"]
PL_FINANCIAL_CODES = ["301", "302"]
PL_INCOME_CODE = "405"
BUDGET_CONTROL_CODES = PL_DIRECT_COST_CODES + PL_SELLING_ADMIN_CODES + PL_FINANCIAL_CODES


def _dashboard_safe_pct(numerator: Decimal, denominator: Decimal) -> Optional[Decimal]:
    if denominator <= 0:
        return None
    return ((numerator / denominator) * Decimal("100")).quantize(TWO_DECIMALS)


def _budget_control_signal(budget: Decimal, real: Decimal, committed: Decimal) -> dict:
    budget_q = money_dec(budget)
    real_q = money_dec(real)
    committed_q = money_dec(committed)
    advance_pct = ((real_q / budget_q) * Decimal("100")).quantize(TWO_DECIMALS) if budget_q > 0 else None
    if budget_q == 0 and real_q == 0 and committed_q == 0:
        return {"traffic_light": "neutral", "advance_pct": None}
    if budget_q == 0 and (real_q > 0 or committed_q > 0):
        return {"traffic_light": "red", "advance_pct": None}
    if advance_pct is None:
        return {"traffic_light": "neutral", "advance_pct": None}
    if advance_pct <= Decimal("90"):
        return {"traffic_light": "green", "advance_pct": advance_pct}
    if advance_pct <= Decimal("100"):
        return {"traffic_light": "yellow", "advance_pct": advance_pct}
    return {"traffic_light": "red", "advance_pct": advance_pct}


def _expense_traffic_light(budget: Decimal, real: Decimal) -> str:
    if budget <= 0 and real <= 0:
        return "neutral"
    if budget <= 0 and real > 0:
        return "red"
    ratio = (real / budget) * Decimal("100")
    if ratio <= Decimal("90"):
        return "green"
    if ratio <= Decimal("100"):
        return "yellow"
    return "red"


def _utility_traffic_light(budget: Decimal, real: Decimal) -> str:
    if budget == 0:
        return "neutral" if real == 0 else ("green" if real > 0 else "red")
    if real >= budget:
        return "green"
    if real >= (budget * Decimal("0.9")):
        return "yellow"
    return "red"


def _build_pl_rows(by_partida: Dict[str, Dict[str, Decimal]], partida_map: Dict[str, dict], ingreso_405: Decimal) -> List[dict]:
    rows: List[dict] = []

    def _build_partida_row(code: str) -> dict:
        source = by_partida.get(code, {"presupuesto": Decimal("0.00"), "ejecutado": Decimal("0.00")})
        budget = source["presupuesto"].quantize(TWO_DECIMALS)
        real = source["ejecutado"].quantize(TWO_DECIMALS)
        remaining = (budget - real).quantize(TWO_DECIMALS)
        part = partida_map.get(code, {})
        pct = _dashboard_safe_pct(real, ingreso_405)
        return {
            "code": code,
            "name": part.get("nombre", code),
            "budget": _dashboard_decimal_to_float(budget),
            "real": _dashboard_decimal_to_float(real),
            "remaining": _dashboard_decimal_to_float(remaining),
            "income_pct": _dashboard_decimal_to_float(pct) if pct is not None else None,
            "traffic_light": _expense_traffic_light(budget, real),
            "row_type": "partida",
        }

    for code in PL_DIRECT_COST_CODES:
        rows.append(_build_partida_row(code))

    direct_budget = sum((by_partida.get(c, {}).get("presupuesto", Decimal("0.00")) for c in PL_DIRECT_COST_CODES), Decimal("0.00")).quantize(TWO_DECIMALS)
    direct_real = sum((by_partida.get(c, {}).get("ejecutado", Decimal("0.00")) for c in PL_DIRECT_COST_CODES), Decimal("0.00")).quantize(TWO_DECIMALS)
    gross_budget = (ingreso_405 - direct_budget).quantize(TWO_DECIMALS)
    gross_real = (ingreso_405 - direct_real).quantize(TWO_DECIMALS)
    gross_remaining = (gross_budget - gross_real).quantize(TWO_DECIMALS)
    gross_pct = _dashboard_safe_pct(gross_real, ingreso_405)
    rows.append({
        "code": "SUBTOTAL_GROSS",
        "name": "UTILIDAD BRUTA",
        "budget": _dashboard_decimal_to_float(gross_budget),
        "real": _dashboard_decimal_to_float(gross_real),
        "remaining": _dashboard_decimal_to_float(gross_remaining),
        "income_pct": _dashboard_decimal_to_float(gross_pct) if gross_pct is not None else None,
        "traffic_light": _utility_traffic_light(gross_budget, gross_real),
        "row_type": "subtotal",
    })

    for code in PL_SELLING_ADMIN_CODES:
        rows.append(_build_partida_row(code))
    sell_admin_budget = sum((by_partida.get(c, {}).get("presupuesto", Decimal("0.00")) for c in PL_SELLING_ADMIN_CODES), Decimal("0.00")).quantize(TWO_DECIMALS)
    sell_admin_real = sum((by_partida.get(c, {}).get("ejecutado", Decimal("0.00")) for c in PL_SELLING_ADMIN_CODES), Decimal("0.00")).quantize(TWO_DECIMALS)
    op_budget = (gross_budget - sell_admin_budget).quantize(TWO_DECIMALS)
    op_real = (gross_real - sell_admin_real).quantize(TWO_DECIMALS)
    op_remaining = (op_budget - op_real).quantize(TWO_DECIMALS)
    op_pct = _dashboard_safe_pct(op_real, ingreso_405)
    rows.append({
        "code": "SUBTOTAL_OPERATING",
        "name": "UTILIDAD OPERATIVA",
        "budget": _dashboard_decimal_to_float(op_budget),
        "real": _dashboard_decimal_to_float(op_real),
        "remaining": _dashboard_decimal_to_float(op_remaining),
        "income_pct": _dashboard_decimal_to_float(op_pct) if op_pct is not None else None,
        "traffic_light": _utility_traffic_light(op_budget, op_real),
        "row_type": "subtotal",
    })

    for code in PL_FINANCIAL_CODES:
        rows.append(_build_partida_row(code))
    financial_budget = sum((by_partida.get(c, {}).get("presupuesto", Decimal("0.00")) for c in PL_FINANCIAL_CODES), Decimal("0.00")).quantize(TWO_DECIMALS)
    financial_real = sum((by_partida.get(c, {}).get("ejecutado", Decimal("0.00")) for c in PL_FINANCIAL_CODES), Decimal("0.00")).quantize(TWO_DECIMALS)
    pbt_budget = (op_budget - financial_budget).quantize(TWO_DECIMALS)
    pbt_real = (op_real - financial_real).quantize(TWO_DECIMALS)
    pbt_remaining = (pbt_budget - pbt_real).quantize(TWO_DECIMALS)
    pbt_pct = _dashboard_safe_pct(pbt_real, ingreso_405)
    rows.append({
        "code": "SUBTOTAL_PRE_TAX",
        "name": "UTILIDAD ANTES DE IMPUESTOS",
        "budget": _dashboard_decimal_to_float(pbt_budget),
        "real": _dashboard_decimal_to_float(pbt_real),
        "remaining": _dashboard_decimal_to_float(pbt_remaining),
        "income_pct": _dashboard_decimal_to_float(pbt_pct) if pbt_pct is not None else None,
        "traffic_light": _utility_traffic_light(pbt_budget, pbt_real),
        "row_type": "subtotal",
    })
    return rows


def _build_corrida_rows(by_partida: List[dict], total_income: Decimal) -> dict:
    rows = []
    grouped: Dict[str, Dict[str, Decimal]] = {}
    for row in by_partida:
        group = str(row.get("partida_grupo") or "OTROS").upper()
        g = grouped.setdefault(group, {"presupuesto": Decimal("0.00"), "real": Decimal("0.00")})
        g["presupuesto"] += money_dec(row.get("presupuesto", 0))
        g["real"] += money_dec(row.get("ejecutado", 0))

    for row in by_partida:
        presupuesto = money_dec(row.get("presupuesto", 0)).quantize(TWO_DECIMALS)
        real = money_dec(row.get("ejecutado", 0)).quantize(TWO_DECIMALS)
        pct_income = ((real / total_income) * Decimal("100")).quantize(TWO_DECIMALS) if total_income > 0 else None
        flujo = (presupuesto - real).quantize(TWO_DECIMALS) if presupuesto > 0 else None
        rows.append({
            "type": "partida",
            "concepto": f"{row.get('partida_codigo')} {row.get('partida_nombre')}",
            "%_sobre_ingreso": float(pct_income) if pct_income is not None else None,
            "presupuesto": float(presupuesto),
            "real": float(real),
            "flujo_por_ejercer": float(flujo) if flujo is not None else None,
            "semaforo": row.get("traffic_light"),
            "status_label": row.get("status_label"),
            "grupo": row.get("partida_grupo"),
        })

    ingresos = grouped.get("INGRESOS", {"presupuesto": Decimal("0.00"), "real": Decimal("0.00")})
    costos = sum((v["real"] for k,v in grouped.items() if "DIRECT" in k or k in {"OBRA", "COSTO DIRECTO", "COSTOS DIRECTOS"}), Decimal("0.00"))
    gya = sum((v["real"] for k,v in grouped.items() if "VTA" in k or "ADM" in k or k in {"GYA"}), Decimal("0.00"))
    fin = sum((v["real"] for k,v in grouped.items() if "FINAN" in k), Decimal("0.00"))
    utilidad_bruta = ingresos["real"] - costos
    utilidad_operativa = utilidad_bruta - gya
    utilidad_antes_impuestos = utilidad_operativa - fin

    subtotals = [
        ("UTILIDAD BRUTA", utilidad_bruta),
        ("UTILIDAD OPERATIVA", utilidad_operativa),
        ("UTILIDAD ANTES IMPUESTOS", utilidad_antes_impuestos),
    ]
    for label, real in subtotals:
        pct_income = ((real / total_income) * Decimal("100")).quantize(TWO_DECIMALS) if total_income > 0 else None
        rows.append({"type": "subtotal", "concepto": label, "%_sobre_ingreso": float(pct_income) if pct_income is not None else None, "presupuesto": None, "real": float(real.quantize(TWO_DECIMALS)), "flujo_por_ejercer": None, "semaforo": "green" if real >= 0 else "red"})
    return {"rows": rows}


async def _dashboard_summary_data(current_user: dict, empresa_id: Optional[str], project_id: Optional[str], period: str, year: Optional[int], month: Optional[int], quarter: Optional[int], include_pending: bool = False):
    now = to_tijuana(datetime.now(timezone.utc))
    normalized_period = (period or "month").strip().lower()
    if normalized_period not in {"all", "month", "quarter", "year"}:
        raise HTTPException(status_code=422, detail={"code": "invalid_period", "message": "period debe ser all|month|quarter|year"})

    selected_year = year or now.year
    validate_year_in_range(selected_year)
    if month is not None and (month < 1 or month > 12):
        raise HTTPException(status_code=422, detail={"code": "month_out_of_range", "message": "month debe estar entre 1 y 12"})
    if quarter is not None and (quarter < 1 or quarter > 4):
        raise HTTPException(status_code=422, detail={"code": "quarter_out_of_range", "message": "quarter debe estar entre 1 y 4"})

    company_selector = (empresa_id or "all").strip() or "all"
    project_selector = (project_id or "all").strip() or "all"
    if company_selector != "all":
        enforce_company_access(current_user, company_selector)

    role = current_user.get("role")
    project_query = {}
    if role not in {UserRole.ADMIN.value, UserRole.FINANZAS.value}:
        allowed_company_ids = list(current_user.get("empresa_ids") or [])
        active_company = get_user_company_id(current_user)
        if active_company and active_company not in allowed_company_ids:
            allowed_company_ids.append(active_company)
        if not allowed_company_ids:
            raise HTTPException(status_code=403, detail={"code": "forbidden_company", "message": "Usuario sin empresas permitidas"})
        project_query["empresa_id"] = {"$in": allowed_company_ids}

    if company_selector != "all":
        project_query["empresa_id"] = company_selector
    if project_selector != "all":
        project_query["id"] = project_selector

    projects = await db.projects.find(project_query, {"_id": 0}).to_list(5000)
    if project_selector != "all" and not projects:
        raise HTTPException(status_code=403, detail={"code": "forbidden_project", "message": "Proyecto fuera de alcance"})

    project_ids = [str(p.get("id")) for p in projects if p.get("id")]
    if not project_ids:
        y, r = await get_dashboard_thresholds()
        return {
            "filtros": {
                "empresa_id": company_selector,
                "empresa_nombre": "Todas" if company_selector == "all" else company_selector,
                "project_id": project_selector,
                "project_nombre": "Todos" if project_selector == "all" else project_selector,
                "period": normalized_period,
                "period_label": _fmt_period_label(normalized_period, selected_year, month, quarter),
                "year": selected_year,
                "month": month,
                "quarter": quarter,
            },
            "totals": {
                "presupuesto_total": 0.0,
                "ejecutado_total": 0.0,
                "variacion_total": 0.0,
                "porcentaje_avance": 0.0,
                "ingreso_proyectado_405": 0.0,
                "por_ejercer_total": 0.0,
                "ejecucion_vs_ingreso_pct": None,
                "porcentaje_label": "0.00",
                "traffic_light": get_traffic_light(0, float(y), float(r)),
                "status_label": "SIN PRESUPUESTO (sin gasto)",
                "variation_color": "neutral",
            },
            "shared_kpis": {
                "ingreso_proyectado_405": 0.0,
                "presupuesto_total": 0.0,
                "real_ejecutado": 0.0,
                "por_ejercer": 0.0,
                "ejecucion_vs_ingreso_pct": None,
            },
            "by_partida": [],
            "rows": [],
            "subtotals": {},
            "pnl": {
                "rows": [],
                "subtotals": {},
            },
            "budget_control": {
                "summary": {
                    "red_count": 0,
                    "yellow_count": 0,
                    "overrun_count": 0,
                    "committed_total": 0.0,
                    "available_total": 0.0,
                },
                "rows": [],
                "totals": {
                    "budget": 0.0,
                    "real": 0.0,
                    "committed": 0.0,
                    "available": 0.0,
                },
            },
            "pending": {"count": 0, "total_mxn": 0.0},
            "pending_authorizations": await db.authorizations.count_documents({"status": "pending"}),
            "movements_count": 0,
            "include_pending": include_pending,
            "meta": {
                "pending_budget_policy": "excluded_from_official_budget",
                "budget_control_committed_policy": "purchase_orders.approved_for_payment pending_amount distributed by line ratio; excludes posted movements already in real",
                "budget_control_available_policy": "budget - real - committed",
            },
        }

    company_ids = sorted({str(p.get("empresa_id")) for p in projects if p.get("empresa_id")})
    empresa_docs = await db.empresas.find({"id": {"$in": company_ids}}, {"_id": 0}).to_list(500)
    empresa_map = {e.get("id"): e.get("nombre") for e in empresa_docs}
    project_map = {p.get("id"): p for p in projects}

    posted_statuses = [MovementStatus.POSTED.value]
    if include_pending:
        posted_statuses.append(MovementStatus.PENDING_APPROVAL.value)
    movement_query = movement_active_query(extra={"status": {"$in": posted_statuses}, "project_id": {"$in": project_ids}})
    movement_rows = await db.movements.find(movement_query, {"_id": 0}).to_list(20000)

    pending_rows = await db.movements.find(movement_active_query(extra={"status": MovementStatus.PENDING_APPROVAL.value, "project_id": {"$in": project_ids}}), {"_id": 0}).to_list(20000)

    month_keys = _month_keys_for_period(normalized_period, selected_year, month, quarter)

    by_partida: Dict[str, Dict[str, Decimal]] = {}

    plan_rows = await db.budget_plans.find({"project_id": {"$in": project_ids}}, {"_id": 0}).to_list(8000)
    plan_pairs = set()
    for plan in plan_rows:
        if plan.get("approval_status") in {BudgetApprovalStatus.PENDING.value, BudgetApprovalStatus.REJECTED.value}:
            # Convención por seguridad: no contar presupuestos pendientes/rechazados en dashboard oficial.
            continue
        proj_id = str(plan.get("project_id"))
        partida = str(plan.get("partida_codigo") or "N/A")
        if not proj_id or proj_id not in project_map:
            continue
        plan_pairs.add((proj_id, partida))
        total_amount = _to_period_decimal(plan.get("total_amount", 0))
        annual_map = plan.get("annual_breakdown") or {}
        monthly_map = plan.get("monthly_breakdown") or {}

        amount = Decimal("0.00")
        if normalized_period == "month":
            mk = month_keys[0]
            if mk in monthly_map:
                amount = _to_period_decimal(monthly_map.get(mk))
        elif normalized_period == "quarter":
            for mk in month_keys:
                if mk in monthly_map:
                    amount += _to_period_decimal(monthly_map.get(mk))
        else:
            year_key = str(selected_year)
            monthly_sum = Decimal("0.00")
            has_monthly = False
            for mk in month_keys:
                if mk in monthly_map:
                    has_monthly = True
                    monthly_sum += _to_period_decimal(monthly_map.get(mk))
            if has_monthly:
                amount = monthly_sum
            elif year_key in annual_map:
                amount = _to_period_decimal(annual_map.get(year_key))
            else:
                amount = total_amount

        row = by_partida.setdefault(partida, {"presupuesto": Decimal("0.00"), "ejecutado": Decimal("0.00")})
        row["presupuesto"] += amount

    legacy_query = {"project_id": {"$in": project_ids}, "year": selected_year}
    if normalized_period == "month":
        legacy_query["month"] = month
    elif normalized_period == "quarter":
        legacy_query["month"] = {"$in": [int(k[-2:]) for k in month_keys]}
    else:
        legacy_query["month"] = {"$in": list(range(1, 13))}
    legacy_rows = await db.budgets.find(legacy_query, {"_id": 0}).to_list(12000)
    for row_doc in legacy_rows:
        proj_id = str(row_doc.get("project_id") or "")
        partida = str(row_doc.get("partida_codigo") or "N/A")
        if (proj_id, partida) in plan_pairs:
            continue
        row = by_partida.setdefault(partida, {"presupuesto": Decimal("0.00"), "ejecutado": Decimal("0.00")})
        row["presupuesto"] += _to_period_decimal(row_doc.get("amount_mxn", 0))

    movements_in_period = []
    for mv in movement_rows:
        mv_date = date_parser.parse(mv.get("date")) if isinstance(mv.get("date"), str) else mv.get("date")
        if not mv_date:
            continue
        if _match_dashboard_period(to_tijuana(mv_date), normalized_period, selected_year, month, quarter):
            movements_in_period.append(mv)
            partida = str(mv.get("partida_codigo") or mv.get("partida_id") or "N/A")
            row = by_partida.setdefault(partida, {"presupuesto": Decimal("0.00"), "ejecutado": Decimal("0.00")})
            row["ejecutado"] += _to_period_decimal(abs(float(mv.get("amount_mxn", 0))))

    pending_in_period = []
    for mv in pending_rows:
        mv_date = date_parser.parse(mv.get("date")) if isinstance(mv.get("date"), str) else mv.get("date")
        if mv_date and _match_dashboard_period(to_tijuana(mv_date), normalized_period, selected_year, month, quarter):
            pending_in_period.append(mv)

    partida_docs = await db.catalogo_partidas.find({}, {"_id": 0}).to_list(1000)
    partida_map = {str(p.get("codigo")): p for p in partida_docs}
    yellow, red = await get_dashboard_thresholds()

    detail = []
    total_budget = Decimal("0.00")
    total_exec = Decimal("0.00")
    for code, values in by_partida.items():
        presupuesto = values["presupuesto"].quantize(TWO_DECIMALS)
        ejecutado = values["ejecutado"].quantize(TWO_DECIMALS)
        signal = build_budget_signal(presupuesto, ejecutado, yellow, red)
        part = partida_map.get(code, {})
        detail.append({
            "partida_codigo": code,
            "partida_nombre": part.get("nombre", code),
            "partida_grupo": _derive_partida_group(code, partida_map),
            "presupuesto": _dashboard_decimal_to_float(presupuesto),
            "ejecutado": _dashboard_decimal_to_float(ejecutado),
            "variacion": _dashboard_decimal_to_float(signal["variacion"]),
            "porcentaje": _dashboard_decimal_to_float(signal["porcentaje"]) if signal["porcentaje"] is not None else None,
            "porcentaje_label": signal["porcentaje_label"],
            "disponible": _dashboard_decimal_to_float(signal["variacion"]),
            "traffic_light": signal["traffic_light"],
            "status_label": signal["status_label"],
            "variation_color": signal["variation_color"],
        })
        total_budget += presupuesto
        total_exec += ejecutado

    total_signal = build_budget_signal(total_budget, total_exec, yellow, red)

    inventory_items = await db.inventory_items.find({"project_id": {"$in": project_ids}}, {"_id": 0}).to_list(10000)
    ingreso_405 = Decimal("0.00")
    for item in inventory_items:
        ingreso_405 += _to_period_decimal(item.get("precio_total", 0))
    ingreso_405 = ingreso_405.quantize(TWO_DECIMALS)

    by_partida[PL_INCOME_CODE] = {
        "presupuesto": ingreso_405,
        "ejecutado": ingreso_405,
    }

    pending_total = sum((_to_period_decimal(m.get("amount_mxn", 0)) for m in pending_in_period), Decimal("0.00"))

    budget_control_codes = set(BUDGET_CONTROL_CODES)
    committed_by_partida: Dict[str, Decimal] = {code: Decimal("0.00") for code in BUDGET_CONTROL_CODES}
    po_query = {
        "project_id": {"$in": project_ids},
        "status": PurchaseOrderStatus.APPROVED_FOR_PAYMENT.value,
    }
    purchase_orders = await db.purchase_orders.find(po_query, {"_id": 0}).to_list(10000)
    for po in purchase_orders:
        po_total = money_dec(po.get("total_mxn") or po.get("total") or 0)
        approved_amount_total = money_dec(po.get("approved_amount_total") or 0)
        pending_amount = money_dec(po.get("pending_amount") or (po_total - approved_amount_total))
        if pending_amount <= 0:
            continue
        lines = po.get("lines") or []
        line_total_sum = Decimal("0.00")
        line_amounts = []
        for line in lines:
            partida = str(line.get("partida_codigo") or "")
            line_total = money_dec(line.get("line_total") or 0)
            line_amount_mxn = line_total if po.get("currency") == Currency.MXN.value else (line_total * money_dec(po.get("exchange_rate") or 1)).quantize(TWO_DECIMALS)
            line_amounts.append((partida, line_amount_mxn))
            line_total_sum += line_amount_mxn
        if line_total_sum <= 0:
            continue
        ratio = (pending_amount / line_total_sum)
        for partida, line_amount in line_amounts:
            if partida not in budget_control_codes:
                continue
            committed_by_partida[partida] = (committed_by_partida.get(partida, Decimal("0.00")) + (line_amount * ratio)).quantize(TWO_DECIMALS)

    budget_control_rows = []
    bc_totals = {
        "budget": Decimal("0.00"),
        "real": Decimal("0.00"),
        "committed": Decimal("0.00"),
        "available": Decimal("0.00"),
    }
    red_count = 0
    yellow_count = 0
    overrun_count = 0
    for code in BUDGET_CONTROL_CODES:
        values = by_partida.get(code, {"presupuesto": Decimal("0.00"), "ejecutado": Decimal("0.00")})
        budget = money_dec(values.get("presupuesto") or 0)
        real = money_dec(values.get("ejecutado") or 0)
        committed = money_dec(committed_by_partida.get(code) or 0)
        available = (budget - real - committed).quantize(TWO_DECIMALS)
        signal = _budget_control_signal(budget, real, committed)
        traffic_light = signal["traffic_light"]
        advance_pct = signal["advance_pct"]
        if traffic_light == "red":
            red_count += 1
        elif traffic_light == "yellow":
            yellow_count += 1
        if budget > 0 and real > budget:
            overrun_count += 1
        if budget == 0 and (real > 0 or committed > 0):
            overrun_count += 1
        part = partida_map.get(code, {})
        budget_control_rows.append({
            "code": code,
            "name": part.get("nombre", code),
            "group": _derive_partida_group(code, partida_map),
            "budget": _dashboard_decimal_to_float(budget),
            "real": _dashboard_decimal_to_float(real),
            "committed": _dashboard_decimal_to_float(committed),
            "available": _dashboard_decimal_to_float(available),
            "advance_pct": _dashboard_decimal_to_float(advance_pct) if advance_pct is not None else None,
            "traffic_light": traffic_light,
        })
        bc_totals["budget"] += budget
        bc_totals["real"] += real
        bc_totals["committed"] += committed
        bc_totals["available"] += available

    company_name = "Todas" if company_selector == "all" else empresa_map.get(company_selector, company_selector)
    if project_selector == "all":
        project_name = "Todos"
    else:
        project_name = (project_map.get(project_selector) or {}).get("name", project_selector)

    income_total = ingreso_405
    corrida = _build_corrida_rows(detail, income_total)
    pl_rows = _build_pl_rows(by_partida, partida_map, ingreso_405)
    subtotals = {row.get("name"): row for row in pl_rows if row.get("row_type") == "subtotal"}

    por_ejercer_total = (total_budget - total_exec).quantize(TWO_DECIMALS)
    ejecucion_vs_ingreso_pct = _dashboard_safe_pct(total_exec, ingreso_405)

    return {
        "filtros": {
            "empresa_id": company_selector,
            "empresa_nombre": company_name,
            "project_id": project_selector,
            "project_nombre": project_name,
            "period": normalized_period,
            "period_label": _fmt_period_label(normalized_period, selected_year, month, quarter),
            "year": selected_year,
            "month": month,
            "quarter": quarter,
        },
        "year": selected_year,
        "month": month,
        "quarter": quarter,
        "period": normalized_period,
        "totals": {
            "presupuesto_total": _dashboard_decimal_to_float(total_budget),
            "ejecutado_total": _dashboard_decimal_to_float(total_exec),
            "variacion_total": _dashboard_decimal_to_float(total_signal["variacion"]),
            "porcentaje_avance": _dashboard_decimal_to_float(total_signal["porcentaje"]) if total_signal["porcentaje"] is not None else None,
            "ingreso_proyectado_405": _dashboard_decimal_to_float(ingreso_405),
            "por_ejercer_total": _dashboard_decimal_to_float(por_ejercer_total),
            "ejecucion_vs_ingreso_pct": _dashboard_decimal_to_float(ejecucion_vs_ingreso_pct) if ejecucion_vs_ingreso_pct is not None else None,
            "porcentaje_label": total_signal["porcentaje_label"],
            "traffic_light": total_signal["traffic_light"],
            "status_label": total_signal["status_label"],
            "variation_color": total_signal["variation_color"],
            # Compat fields
            "budget": _dashboard_decimal_to_float(total_budget),
            "real": _dashboard_decimal_to_float(total_exec),
            "variation": _dashboard_decimal_to_float(total_signal["variacion"]),
            "percentage": _dashboard_decimal_to_float(total_signal["porcentaje"]) if total_signal["porcentaje"] is not None else None,
        },
        "shared_kpis": {
            "ingreso_proyectado_405": _dashboard_decimal_to_float(ingreso_405),
            "presupuesto_total": _dashboard_decimal_to_float(total_budget),
            "real_ejecutado": _dashboard_decimal_to_float(total_exec),
            "por_ejercer": _dashboard_decimal_to_float(por_ejercer_total),
            "ejecucion_vs_ingreso_pct": _dashboard_decimal_to_float(ejecucion_vs_ingreso_pct) if ejecucion_vs_ingreso_pct is not None else None,
        },
        "pending": {
            "count": len(pending_in_period),
            "total_mxn": _dashboard_decimal_to_float(pending_total.quantize(TWO_DECIMALS)),
        },
        "by_partida": sorted(detail, key=lambda x: x.get("porcentaje", 0), reverse=True),
        "rows": pl_rows,
        "subtotals": subtotals,
        "pnl": {
            "rows": pl_rows,
            "subtotals": subtotals,
        },
        "budget_control": {
            "summary": {
                "red_count": red_count,
                "yellow_count": yellow_count,
                "overrun_count": overrun_count,
                "committed_total": _dashboard_decimal_to_float(bc_totals["committed"]),
                "available_total": _dashboard_decimal_to_float(bc_totals["available"]),
            },
            "rows": budget_control_rows,
            "totals": {
                "budget": _dashboard_decimal_to_float(bc_totals["budget"]),
                "real": _dashboard_decimal_to_float(bc_totals["real"]),
                "committed": _dashboard_decimal_to_float(bc_totals["committed"]),
                "available": _dashboard_decimal_to_float(bc_totals["available"]),
            },
        },
        "by_project": [],
        "corrida": corrida,
        "pending_authorizations": await db.authorizations.count_documents({"status": "pending"}),
        "movements_count": len(movements_in_period),
        "include_pending": include_pending,
        "meta": {
            "threshold_yellow": _dashboard_decimal_to_float(yellow),
            "threshold_red": _dashboard_decimal_to_float(red),
            "pending_budget_policy": "excluded_from_official_budget",
            "income_source": "inventory_items.precio_total",
            "income_partida_code": PL_INCOME_CODE,
            "budget_control_committed_policy": "purchase_orders.approved_for_payment pending_amount distributed by line ratio; excludes posted movements already in real",
            "budget_control_available_policy": "budget - real - committed",
        }
    }
@api_router.get("/reports/corrida")
async def get_corrida(
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    period: str = "month",
    year: Optional[int] = None,
    month: Optional[int] = None,
    quarter: Optional[int] = None,
    include_pending: bool = False,
    current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD))
):
    data = await _dashboard_summary_data(
        current_user=current_user,
        empresa_id=empresa_id,
        project_id=project_id,
        period=period,
        year=year,
        month=month,
        quarter=quarter,
        include_pending=include_pending,
    )
    return {
        "filtros": data.get("filtros", {}),
        "periodo": {
            "period": data.get("period"),
            "year": data.get("year"),
            "month": data.get("month"),
            "quarter": data.get("quarter"),
        },
        "rows": (data.get("corrida") or {}).get("rows", []),
        "totals": data.get("totals", {}),
    }

@api_router.get("/reports/dashboard")
async def get_dashboard(
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    period: str = "month",
    year: Optional[int] = None,
    month: Optional[int] = None,
    quarter: Optional[int] = None,
    include_pending: bool = False,  # compat legacy response field
    current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD))
):
    return await _dashboard_summary_data(
        current_user=current_user,
        empresa_id=empresa_id,
        project_id=project_id,
        period=period,
        year=year,
        month=month,
        quarter=quarter,
        include_pending=include_pending,
    )

@api_router.get("/reports/partida-detail/{partida_codigo}")
async def get_partida_detail(
    partida_codigo: str,
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    now = to_tijuana(datetime.now(timezone.utc))
    year = year or now.year
    validate_year_in_range(year)
    month = month or now.month
    
    # Get partida from catalogo
    partida = await db.catalogo_partidas.find_one({"codigo": partida_codigo}, {"_id": 0})
    if not partida:
        raise HTTPException(status_code=404, detail=f"Partida {partida_codigo} no encontrada en catálogo")
    
    # Get projects filtered by empresa
    project_ids = None
    if empresa_id:
        projects = await db.projects.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
        project_ids = [p['id'] for p in projects]
    
    # Get budgets
    budget_query = {"partida_codigo": partida_codigo, "year": year, "month": month}
    if project_id:
        budget_query["project_id"] = project_id
    elif project_ids:
        budget_query["project_id"] = {"$in": project_ids}
    
    budgets = await db.budgets.find(budget_query, {"_id": 0}).to_list(1000)
    total_budget = sum(b['amount_mxn'] for b in budgets)
    
    # Get movements - SOLO posted
    movement_query = {"partida_codigo": partida_codigo, "status": MovementStatus.POSTED.value}
    if project_id:
        movement_query["project_id"] = project_id
    elif project_ids:
        movement_query["project_id"] = {"$in": project_ids}
    
    all_movements = await db.movements.find(movement_active_query(extra=movement_query), {"_id": 0}).to_list(5000)
    movements = [
        m for m in all_movements
        if date_parser.parse(m['date']).year == year and date_parser.parse(m['date']).month == month
    ]
    
    total_real = sum(abs(float(m.get('amount_mxn',0))) for m in movements)
    percentage = (total_real / total_budget * 100) if total_budget > 0 else (100 if total_real > 0 else 0)
    
    # Get provider names
    provider_docs = await db.providers.find({}, {"_id": 0}).to_list(1000)
    provider_map = {p['id']: p for p in provider_docs}
    
    # Get project names
    project_docs = await db.projects.find({}, {"_id": 0}).to_list(1000)
    project_map = {p['id']: p for p in project_docs}
    
    # Enrich movements
    for m in movements:
        prov = provider_map.get(m.get('provider_id'), {})
        proj = project_map.get(m['project_id'], {})
        m['provider_name'] = prov.get('name', 'N/A')
        m['project_name'] = proj.get('name', 'N/A')
    
    return {
        "partida": {"codigo": partida['codigo'], "nombre": partida['nombre'], "grupo": partida['grupo']},
        "year": year,
        "month": month,
        "budget": total_budget,
        "real": total_real,
        "variation": total_budget - total_real,
        "percentage": percentage,
        "traffic_light": get_traffic_light(percentage),
        "movements": sorted(movements, key=lambda x: x['date'], reverse=True)
    }

@api_router.get("/reports/export-data")
async def get_export_data(
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Returns data for Excel export with 2 sheets:
    - Hoja 1 "Resumen" (KPIs)
    - Hoja 2 "Detalle" (por partida + semáforo)
    All dates in America/Tijuana timezone.
    Logs export action to audit.
    """
    now = to_tijuana(datetime.now(timezone.utc))
    year = year or now.year
    validate_year_in_range(year)
    month = month or now.month
    
    # Get empresas for filter names
    empresas = await db.empresas.find({}, {"_id": 0}).to_list(1000)
    empresa_map = {e['id']: e for e in empresas}
    
    # Get projects filtered by empresa
    project_query = {}
    if empresa_id:
        project_query["empresa_id"] = empresa_id
    all_projects = await db.projects.find(project_query, {"_id": 0}).to_list(1000)
    project_ids = [p['id'] for p in all_projects]
    project_map = {p['id']: p for p in all_projects}
    
    # Get budgets
    budget_query = {"year": year, "month": month}
    if project_id:
        budget_query["project_id"] = project_id
    elif empresa_id:
        budget_query["project_id"] = {"$in": project_ids}
    
    budgets = await db.budgets.find(budget_query, {"_id": 0}).to_list(1000)
    
    # Get movements
    movement_query = {"status": MovementStatus.POSTED.value}
    if project_id:
        movement_query["project_id"] = project_id
    elif empresa_id:
        movement_query["project_id"] = {"$in": project_ids}
    
    all_movements = await db.movements.find(movement_active_query(extra=movement_query), {"_id": 0}).to_list(5000)
    
    # Filter by date
    movements = [
        m for m in all_movements
        if date_parser.parse(m['date']).year == year and date_parser.parse(m['date']).month == month
    ]
    
    # Calculate totals
    total_budget = sum(b['amount_mxn'] for b in budgets)
    total_real = sum(abs(float(m.get('amount_mxn',0))) for m in movements)
    variation = total_budget - total_real
    percentage = (total_real / total_budget * 100) if total_budget > 0 else 0
    
    # By partida
    partidas_data = {}
    for b in budgets:
        key = b.get('partida_codigo', 'N/A')
        if key not in partidas_data:
            partidas_data[key] = {"budget": 0, "real": 0}
        partidas_data[key]["budget"] += b['amount_mxn']
    
    for m in movements:
        key = m.get('partida_codigo', 'N/A')
        if key not in partidas_data:
            partidas_data[key] = {"budget": 0, "real": 0}
        partidas_data[key]["real"] += abs(float(m.get('amount_mxn',0)))
    
    # Get partida info from catalogo
    partida_docs = await db.catalogo_partidas.find({}, {"_id": 0}).to_list(1000)
    partida_map = {p['codigo']: p for p in partida_docs}
    
    # Get providers
    provider_docs = await db.providers.find({}, {"_id": 0}).to_list(1000)
    provider_map = {p['id']: p for p in provider_docs}
    
    partidas_detail = []
    for partida_codigo, data in partidas_data.items():
        pct = (data['real'] / data['budget'] * 100) if data['budget'] > 0 else (100 if data['real'] > 0 else 0)
        partida_info = partida_map.get(partida_codigo, {})
        
        # Get movements for this partida
        partida_movements = [m for m in movements if m.get('partida_codigo') == partida_codigo]
        
        # Enrich movements with provider and project names, convert dates to Tijuana
        enriched_movements = []
        for mov in partida_movements:
            proj = project_map.get(mov['project_id'], {})
            prov = provider_map.get(mov.get('provider_id'), {})
            mov_date = date_parser.parse(mov['date'])
            mov_date_tj = to_tijuana(mov_date)
            
            enriched_movements.append({
                "fecha": mov_date_tj.strftime('%Y-%m-%d %H:%M'),
                "proyecto": proj.get('name', 'N/A'),
                "proveedor": prov.get('name', 'N/A'),
                "referencia": mov.get('reference', ''),
                "moneda": mov.get('currency', 'MXN'),
                "monto_original": mov.get('amount_original', 0),
                "tipo_cambio": mov.get('exchange_rate', 1),
                "monto_mxn": mov.get('amount_mxn', 0),
                "descripcion": mov.get('description', '')
            })
        
        partidas_detail.append({
            "codigo": partida_codigo,
            "nombre": partida_info.get('nombre', partida_codigo),
            "grupo": partida_info.get('grupo', 'N/A'),
            "presupuesto": data['budget'],
            "ejecutado": data['real'],
            "variacion": data['budget'] - data['real'],
            "porcentaje": pct,
            "semaforo": "Normal" if pct <= 90 else ("Alerta" if pct <= 100 else "Exceso"),
            "movimientos": sorted(enriched_movements, key=lambda x: x['fecha'], reverse=True)
        })
    
    # Build filter description
    filter_desc = {
        "empresa": empresa_map.get(empresa_id, {}).get('nombre', 'Todas') if empresa_id else "Todas",
        "proyecto": project_map.get(project_id, {}).get('name', 'Todos') if project_id else "Todos",
        "año": year,
        "mes": month
    }
    
    # LOG EXPORT to audit
    export_audit = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["user_id"],
        "user_email": current_user["email"],
        "action": "EXPORT_EXCEL",
        "timestamp_inicio": datetime.now(timezone.utc).isoformat(),
        "timestamp_fin": datetime.now(timezone.utc).isoformat(),
        "filtros": filter_desc,
        "conteos": {
            "partidas": len(partidas_detail),
            "movimientos": len(movements),
            "total_presupuesto": total_budget,
            "total_ejecutado": total_real
        }
    }
    await db.import_export_logs.insert_one(export_audit)
    
    await log_audit(current_user, "EXPORT_EXCEL", "reports", "export", {"filtros": filter_desc})
    
    months_es = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    
    return {
        "filtros": filter_desc,
        "periodo": f"{months_es.get(month, month)} {year}",
        "resumen": {
            "presupuesto": total_budget,
            "ejecutado": total_real,
            "variacion": variation,
            "porcentaje": percentage,
            "semaforo": "Normal" if percentage <= 90 else ("Alerta" if percentage <= 100 else "Exceso")
        },
        "detalle_partidas": sorted(partidas_detail, key=lambda x: x['porcentaje'], reverse=True),
        "generated_at": to_tijuana(datetime.now(timezone.utc)).strftime('%Y-%m-%d %H:%M:%S'),
        "timezone": "America/Tijuana"
    }

@api_router.get("/import-export-logs")
async def get_import_export_logs(
    action: Optional[str] = None,
    limit: int = Query(50, le=500),
    current_user: dict = Depends(require_permission(Permission.VIEW_AUDIT))
):
    """Get import/export audit logs"""
    query = {}
    if action:
        query["action"] = action
    
    logs = await db.import_export_logs.find(query, {"_id": 0}).sort("timestamp_inicio", -1).to_list(limit)
    return logs

# ========================= AUDIT LOG ROUTES (ENHANCED P3.2) =========================
@api_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    user_id: Optional[str] = None,
    user_role: Optional[str] = None,
    action: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(100, le=1000),
    current_user: dict = Depends(require_permission(Permission.VIEW_AUDIT))
):
    """Enhanced audit log with advanced filters"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if user_id:
        query["user_id"] = user_id
    if user_role:
        query["user_role"] = user_role
    if action:
        query["action"] = {"$regex": action, "$options": "i"}
    
    # Date range filter
    if date_from or date_to:
        query["timestamp"] = {}
        if date_from:
            query["timestamp"]["$gte"] = date_from
        if date_to:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return [to_json_safe(sanitize_mongo_document(log)) for log in logs]

@api_router.get("/audit-logs/export-csv")
async def export_audit_logs_csv(
    entity_type: Optional[str] = None,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(5000, le=10000),
    current_user: dict = Depends(require_permission(Permission.EXPORT_AUDIT))
):
    """Export audit log to CSV (Admin only)"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if user_id:
        query["user_id"] = user_id
    if action:
        query["action"] = {"$regex": action, "$options": "i"}
    if date_from or date_to:
        query["timestamp"] = {}
        if date_from:
            query["timestamp"]["$gte"] = date_from
        if date_to:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    logs = [to_json_safe(sanitize_mongo_document(log)) for log in logs]
    
    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Timestamp", "Usuario", "Email", "Rol", "Acción", "Entidad", "Entity ID", "Cambios"])
    
    for log in logs:
        changes_str = str(log.get('changes', {}))[:200]  # Truncate changes for CSV
        writer.writerow([
            log.get('timestamp', ''),
            log.get('user_id', ''),
            log.get('user_email', ''),
            log.get('user_role', ''),
            log.get('action', ''),
            log.get('entity_type', ''),
            log.get('entity_id', ''),
            changes_str
        ])
    
    await log_audit(current_user, "EXPORT_AUDIT_CSV", "audit_logs", "export", {
        "filters": {"entity_type": entity_type, "user_id": user_id, "action": action, "date_from": date_from, "date_to": date_to},
        "count": len(logs)
    })
    
    return {
        "csv_content": output.getvalue(),
        "filename": f"audit_log_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv",
        "count": len(logs)
    }

@api_router.get("/audit-logs/summary")
async def get_audit_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_AUDIT))
):
    """Get audit log summary for dashboard"""
    query = {}
    if date_from or date_to:
        query["timestamp"] = {}
        if date_from:
            query["timestamp"]["$gte"] = date_from
        if date_to:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
    
    logs = await db.audit_logs.find(query, {"_id": 0}).to_list(5000)
    logs = [to_json_safe(sanitize_mongo_document(log)) for log in logs]
    
    # Count by action
    by_action = {}
    by_user = {}
    by_entity = {}
    
    for log in logs:
        action = log.get('action', 'UNKNOWN')
        user = log.get('user_email', 'N/A')
        entity = log.get('entity_type', 'N/A')
        
        by_action[action] = by_action.get(action, 0) + 1
        by_user[user] = by_user.get(user, 0) + 1
        by_entity[entity] = by_entity.get(entity, 0) + 1
    
    return {
        "total_logs": len(logs),
        "by_action": dict(sorted(by_action.items(), key=lambda x: x[1], reverse=True)[:20]),
        "by_user": dict(sorted(by_user.items(), key=lambda x: x[1], reverse=True)[:10]),
        "by_entity": dict(sorted(by_entity.items(), key=lambda x: x[1], reverse=True)[:10])
    }

@api_router.get("/rbac/permissions-matrix")
async def get_permissions_matrix(current_user: dict = Depends(require_permission(Permission.VIEW_AUDIT))):
    """Return the RBAC permissions matrix for documentation"""
    return {
        "roles": [r.value for r in UserRole],
        "permissions": [p.value for p in Permission],
        "matrix": {role: perms for role, perms in ROLE_PERMISSIONS.items()}
    }

# ========================= CONFIG ROUTES =========================
@api_router.get("/config")
async def get_config(current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    configs = await db.config.find({}, {"_id": 0}).to_list(100)
    return {c['key']: c['value'] for c in configs}

@api_router.put("/config/{key}")
async def update_config(key: str, value: Any, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    existing = await db.config.find_one({"key": key}, {"_id": 0})
    
    if existing:
        old_value = existing['value']
        await db.config.update_one(
            {"key": key},
            {"$set": {"value": value, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user["user_id"]}}
        )
    else:
        config = ConfigSetting(key=key, value=value, updated_by=current_user["user_id"])
        doc = config.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.config.insert_one(doc)
        old_value = None
    
    await log_audit(current_user, "UPDATE", "config", key, {"before": old_value, "after": value})
    return {"message": "Configuración actualizada"}



@api_router.get("/admin/integrations/odoo")
async def get_odoo_integration(current_user: dict = Depends(require_roles(UserRole.ADMIN))):
    cfg = await get_odoo_config()
    masked = dict(cfg)
    masked["odoo_api_key"] = "***" if masked.get("odoo_api_key") else ""
    return masked


@api_router.put("/admin/integrations/odoo")
async def put_odoo_integration(payload: OdooIntegrationConfigInput, current_user: dict = Depends(require_roles(UserRole.ADMIN))):
    mode = (payload.odoo_mode or "stub").strip().lower()
    if mode not in {"stub", "live"}:
        raise HTTPException(status_code=422, detail=structured_error("invalid_odoo_mode", "odoo_mode debe ser stub o live"))
    cfg = {
        "odoo_mode": mode,
        "odoo_url": (payload.odoo_url or "").strip(),
        "odoo_db": (payload.odoo_db or "").strip(),
        "odoo_username": (payload.odoo_username or "").strip(),
        "odoo_api_key": (payload.odoo_api_key or "").strip(),
        "default_model": (payload.default_model or "purchase.order").strip() or "purchase.order",
    }
    await save_odoo_config(current_user, cfg)
    await log_audit(current_user, "ODOO_MODE_CHANGE", "config", "odoo_integration", {"after": sanitize_sensitive_dict(cfg)})
    out = dict(cfg)
    out["odoo_api_key"] = "***" if out.get("odoo_api_key") else ""
    return out


@api_router.post("/admin/integrations/odoo/test-connection")
async def test_odoo_integration(current_user: dict = Depends(require_roles(UserRole.ADMIN))):
    cfg = await get_odoo_config()
    required = ["odoo_url", "odoo_db", "odoo_username", "odoo_api_key"]
    missing = [k for k in required if not cfg.get(k)]
    if missing:
        raise HTTPException(status_code=422, detail=structured_error("odoo_config_incomplete", "Faltan campos de configuración", {"missing": missing}))
    try:
        common = xmlrpc.client.ServerProxy(f"{cfg['odoo_url'].rstrip('/')}/xmlrpc/2/common")
        uid = common.authenticate(cfg["odoo_db"], cfg["odoo_username"], cfg["odoo_api_key"], {})
        if not uid:
            raise HTTPException(status_code=403, detail=structured_error("odoo_auth_failed", "Autenticación Odoo fallida"))
        await log_audit(current_user, "ODOO_TEST_CONNECTION", "config", "odoo_integration", {"result": "ok", "mode": cfg.get("odoo_mode", "stub")})
        return {"ok": True, "uid": uid, "mode": cfg.get("odoo_mode", "stub")}
    except HTTPException:
        raise
    except Exception as exc:
        await log_audit(current_user, "ODOO_TEST_CONNECTION", "config", "odoo_integration", {"result": "failed", "error": str(exc)})
        raise HTTPException(status_code=409, detail=structured_error("odoo_connection_failed", "Error conectando con Odoo", {"error": str(exc)}))


@api_router.post("/purchase-orders/{po_id}/odoo-sync")
async def manual_sync_purchase_order_odoo(po_id: str, current_user: dict = Depends(require_roles(UserRole.ADMIN))):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail=structured_error("purchase_order_not_found", "OC no encontrada"))
    if po.get("status") != PurchaseOrderStatus.APPROVED_FOR_PAYMENT.value:
        raise HTTPException(status_code=409, detail=structured_error("oc_state_conflict", "Solo se puede enviar a Odoo cuando está approved_for_payment"))
    result = await _odoo_send_purchase_order(po, current_user, manual_retry=True)
    return {"purchase_order_id": po_id, "odoo": sanitize_mongo_document(result)}

# ========================= DATA MIGRATION =========================
@api_router.post("/migrate-movement-status")
async def migrate_movement_status():
    """Migrate old movement statuses to new format: normal/authorized -> posted"""
    # Update normal -> posted
    result_normal = await db.movements.update_many(
        {"status": "normal"},
        {"$set": {"status": MovementStatus.POSTED.value}}
    )
    
    # Update authorized -> posted
    result_auth = await db.movements.update_many(
        {"status": "authorized"},
        {"$set": {"status": MovementStatus.POSTED.value}}
    )
    
    # Update pending_authorization -> pending_approval
    result_pending = await db.movements.update_many(
        {"status": "pending_authorization"},
        {"$set": {"status": MovementStatus.PENDING_APPROVAL.value}}
    )
    
    return {
        "migrated": {
            "normal_to_posted": result_normal.modified_count,
            "authorized_to_posted": result_auth.modified_count,
            "pending_authorization_to_pending_approval": result_pending.modified_count
        }
    }

# ========================= DEMO DATA =========================
ADMIN_ENTITY_COLLECTIONS = {
    "empresas": ("empresas", [("nombre", 1)]),
    "proyectos": ("projects", [("name", 1)]),
    "catalogo_partidas": ("catalogo_partidas", [("codigo", 1)]),
    "proveedores": ("providers", [("name", 1)]),
    "usuarios": ("users", [("name", 1)]),
}

ADMIN_CATALOGOS_ALIAS = {
    "e": "empresas",
    "p": "proyectos",
    "c": "catalogo_partidas",
    "d": "proveedores",
    "u": "usuarios",
}



@api_router.get("/admin/catalogs/{entity}")
async def admin_list_entity(
    entity: str,
    include_inactive: bool = False,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    if entity not in ADMIN_ENTITY_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Entidad admin no soportada")

    collection, sort = ADMIN_ENTITY_COLLECTIONS[entity]
    query = active_query(include_inactive)
    projection = {"_id": 0}
    if collection == "users":
        projection["password_hash"] = 0
    cursor = db[collection].find(query, projection)
    if sort:
        cursor = cursor.sort(sort)
    rows = await cursor.to_list(1000)
    return [sanitize_mongo_document(row) for row in rows]


@api_router.get("/admin/empresas")
async def admin_empresas(current_user: dict = Depends(require_permission(Permission.MANAGE_USERS))):
    ensure_admin(current_user)
    rows = await db.empresas.find(active_query(include_inactive=True), {"_id": 0}).sort("nombre", 1).to_list(1000)
    return [sanitize_mongo_document(r) for r in rows]


@api_router.get("/admin/users")
async def admin_list_users(
    include_inactive: bool = True,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    query = {} if include_inactive else {"is_active": {"$ne": False}}
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("name", 1).to_list(1000)
    return [sanitize_mongo_document(u) for u in users]


class AdminUserRoleUpdate(BaseModel):
    role: UserRole


class AdminUserUpdate(BaseModel):
    role: Optional[UserRole] = None
    is_active: Optional[bool] = None
    empresa_ids: Optional[List[str]] = None


@api_router.put("/admin/users/{user_id}")
async def admin_update_user(user_id: str, payload: AdminUserUpdate, request: Request, current_user: dict = Depends(require_permission(Permission.MANAGE_USERS))):
    ensure_admin(current_user)
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    update_data = {}
    if payload.role is not None:
        update_data["role"] = payload.role.value
    if payload.is_active is not None:
        update_data["is_active"] = payload.is_active
    if payload.empresa_ids is not None:
        update_data["empresa_ids"] = [str(e).strip() for e in payload.empresa_ids if str(e).strip()]

    if not update_data:
        return sanitize_mongo_document(user)

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    await log_admin_action(request, current_user, "ADMIN_UPDATE", "users", user_id, True, before=user, after=updated)
    return sanitize_mongo_document(updated)


@api_router.patch("/admin/users/{user_id}/role")
async def admin_update_user_role(
    user_id: str,
    payload: AdminUserRoleUpdate,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if user_id == current_user.get("user_id") and payload.role != UserRole.ADMIN:
        raise HTTPException(status_code=409, detail="No puedes quitarte rol admin a ti mismo")

    if user.get("role") == UserRole.ADMIN.value and payload.role != UserRole.ADMIN:
        admins = await db.users.count_documents({"role": UserRole.ADMIN.value, "is_active": {"$ne": False}})
        if admins <= 1:
            raise HTTPException(status_code=409, detail="No puedes quitar el último admin")

    await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": payload.role.value, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    await log_admin_action(request, current_user, "ADMIN_UPDATE_ROLE", "users", user_id, True, before=user, after=updated)
    return sanitize_mongo_document(updated)


@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(
    user_id: str,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    if user_id == current_user.get("user_id"):
        raise HTTPException(status_code=409, detail="No puedes eliminar tu propio usuario")

    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if user.get("role") == UserRole.ADMIN.value:
        admins = await db.users.count_documents({"role": UserRole.ADMIN.value, "is_active": {"$ne": False}})
        if admins <= 1:
            raise HTTPException(status_code=409, detail="No puedes eliminar el último admin")

    await db.users.delete_one({"id": user_id})
    await log_admin_action(request, current_user, "ADMIN_DELETE_USER", "users", user_id, True, before=user)
    return {"message": "Usuario eliminado"}


class AdminPasswordUpdate(BaseModel):
    password: str


@api_router.patch("/admin/users/{user_id}/password")
async def admin_update_user_password(
    user_id: str,
    payload: AdminPasswordUpdate,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "password_hash": hash_password(payload.password),
            "must_change_password": True,
            "is_active": True,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    await log_admin_action(request, current_user, "ADMIN_RESET_PASSWORD", "users", user_id, True, after=updated)
    return {"message": "Contraseña actualizada", "user": updated}


@api_router.get("/admin/catalogos/{tipo}")
async def admin_catalogos_alias(
    tipo: str,
    de_inactive: bool = False,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    mapped_entity = ADMIN_CATALOGOS_ALIAS.get(tipo, tipo)
    return await admin_list_entity(mapped_entity, include_inactive=de_inactive, current_user=current_user)


@api_router.get("/admin/catalogos/{tipo}_de_inactive={de_inactive}")
async def admin_catalogos_legacy_path(
    tipo: str,
    de_inactive: bool,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    mapped_entity = ADMIN_CATALOGOS_ALIAS.get(tipo, tipo)
    return await admin_list_entity(mapped_entity, include_inactive=de_inactive, current_user=current_user)


@api_router.post("/admin/catalogs/{entity}", status_code=201)
async def admin_create_entity(
    entity: str,
    payload: dict,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    now = datetime.now(timezone.utc).isoformat()
    if entity not in ADMIN_ENTITY_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Entidad admin no soportada")

    collection = ADMIN_ENTITY_COLLECTIONS[entity][0]
    doc = payload.copy()
    doc["id"] = doc.get("id") or str(uuid.uuid4())
    doc.setdefault("is_active", True)
    doc.setdefault("created_at", now)
    doc.setdefault("updated_at", now)

    if collection == "users":
        password = doc.pop("password", None)
        if not password:
            raise HTTPException(status_code=400, detail="password es requerido")
        if not doc.get("email"):
            raise HTTPException(status_code=400, detail="email es requerido")
        if not doc.get("name"):
            raise HTTPException(status_code=400, detail="name es requerido")
        role = doc.get("role")
        try:
            doc["role"] = UserRole(role).value
        except Exception:
            allowed = ", ".join([r.value for r in UserRole])
            raise HTTPException(status_code=400, detail=f"role inválido. Valores permitidos: {allowed}")

        is_active_raw = doc.get("is_active", True)
        if isinstance(is_active_raw, str):
            lower = is_active_raw.strip().lower()
            if lower in {"true", "1", "yes", "si", "sí"}:
                doc["is_active"] = True
            elif lower in {"false", "0", "no"}:
                doc["is_active"] = False
            else:
                raise HTTPException(status_code=400, detail="is_active debe ser booleano (true/false)")
        else:
            doc["is_active"] = bool(is_active_raw)

        doc["password_hash"] = hash_password(password)
        doc["must_change_password"] = True

        existing_email = await db.users.find_one({"email": doc.get("email")}, {"_id": 0})
        if existing_email:
            raise HTTPException(status_code=409, detail="Ya existe un usuario con ese email")
        existing_name = await db.users.find_one({"name": doc.get("name")}, {"_id": 0})
        if existing_name:
            raise HTTPException(status_code=409, detail="Ya existe un usuario con ese username")

    try:
        await db[collection].insert_one(doc)
    except DuplicateKeyError:
        if collection == "users":
            raise HTTPException(status_code=409, detail="Usuario duplicado (email o username)")
        raise HTTPException(status_code=409, detail="Registro duplicado")
    await log_admin_action(request, current_user, "ADMIN_CREATE", entity, doc["id"], True, after={k: v for k, v in doc.items() if k != "password_hash"})
    doc.pop("password_hash", None)
    return sanitize_mongo_document(doc)


@api_router.put("/admin/catalogs/{entity}/{entity_id}")
async def admin_update_entity(
    entity: str,
    entity_id: str,
    payload: dict,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    if entity not in ADMIN_ENTITY_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Entidad admin no soportada")

    collection = ADMIN_ENTITY_COLLECTIONS[entity][0]
    old_doc = await db[collection].find_one({"id": entity_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if collection == "users" and "role" in payload and old_doc.get("role") == UserRole.ADMIN.value and payload.get("role") != UserRole.ADMIN.value:
        admins = await db.users.count_documents({"role": UserRole.ADMIN.value, "is_active": {"$ne": False}})
        if admins <= 1:
            raise HTTPException(status_code=409, detail="No puedes quitar el último admin")

    if collection == "users" and "password" in payload:
        payload["password_hash"] = hash_password(payload.pop("password"))
        payload["must_change_password"] = True

    if collection == "users" and "role" in payload:
        try:
            payload["role"] = UserRole(payload.get("role")).value
        except Exception:
            allowed = ", ".join([r.value for r in UserRole])
            raise HTTPException(status_code=400, detail=f"role inválido. Valores permitidos: {allowed}")

    if collection == "users" and "is_active" in payload and isinstance(payload.get("is_active"), str):
        lower = payload.get("is_active").strip().lower()
        if lower in {"true", "1", "yes", "si", "sí"}:
            payload["is_active"] = True
        elif lower in {"false", "0", "no"}:
            payload["is_active"] = False
        else:
            raise HTTPException(status_code=400, detail="is_active debe ser booleano (true/false)")

    if entity == "catalogo_partidas" and old_doc.get("codigo") != payload.get("codigo", old_doc.get("codigo")):
        raise HTTPException(status_code=400, detail="No se permite cambiar código de partida")

    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    try:
        await db[collection].update_one({"id": entity_id}, {"$set": payload})
    except DuplicateKeyError:
        if collection == "users":
            raise HTTPException(status_code=409, detail="Ya existe un usuario con ese email")
        raise HTTPException(status_code=409, detail="Registro duplicado")
    updated = await db[collection].find_one({"id": entity_id}, {"_id": 0})

    await log_admin_action(
        request,
        current_user,
        "ADMIN_UPDATE",
        entity,
        entity_id,
        True,
        before={k: v for k, v in old_doc.items() if k != "password_hash"},
        after={k: v for k, v in updated.items() if k != "password_hash"},
    )
    updated.pop("password_hash", None)
    return sanitize_mongo_document(updated)


@api_router.delete("/admin/catalogs/{entity}/{entity_id}")
async def admin_delete_entity(
    entity: str,
    entity_id: str,
    hard_delete: bool = False,
    request: Request = None,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    collection_map = {
        "empresas": "empresas",
        "proyectos": "projects",
        "catalogo_partidas": "catalogo_partidas",
        "proveedores": "providers",
        "usuarios": "users",
    }
    if entity not in collection_map:
        raise HTTPException(status_code=404, detail="Entidad admin no soportada")
    collection = collection_map[entity]

    old_doc = await db[collection].find_one({"id": entity_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    if collection == "users":
        if entity_id == current_user.get("user_id"):
            raise HTTPException(status_code=409, detail="No puedes eliminar tu propio usuario")
        if old_doc.get("role") == UserRole.ADMIN.value:
            admins = await db.users.count_documents({"role": UserRole.ADMIN.value, "is_active": {"$ne": False}})
            if admins <= 1:
                raise HTTPException(status_code=409, detail="No puedes eliminar el último admin")

    if hard_delete:
        await assert_no_references(entity, entity_id)
        await db[collection].delete_one({"id": entity_id})
        if request:
            await log_admin_action(request, current_user, "ADMIN_HARD_DELETE", entity, entity_id, True, before={k: v for k, v in old_doc.items() if k != "password_hash"})
        return {"message": "Eliminado físicamente"}

    await db[collection].update_one({"id": entity_id}, {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if request:
        await log_admin_action(request, current_user, "ADMIN_SOFT_DELETE", entity, entity_id, True, before={k: v for k, v in old_doc.items() if k != "password_hash"}, after={"is_active": False})
    return {"message": "Desactivado"}


@api_router.post("/admin/catalogs/{entity}/{entity_id}/restore")
async def admin_restore_entity(
    entity: str,
    entity_id: str,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    collection_map = {
        "empresas": "empresas",
        "proyectos": "projects",
        "catalogo_partidas": "catalogo_partidas",
        "proveedores": "providers",
        "usuarios": "users",
    }
    if entity not in collection_map:
        raise HTTPException(status_code=404, detail="Entidad admin no soportada")
    collection = collection_map[entity]

    old_doc = await db[collection].find_one({"id": entity_id}, {"_id": 0})
    if not old_doc:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    await db[collection].update_one({"id": entity_id}, {"$set": {"is_active": True, "updated_at": datetime.now(timezone.utc).isoformat()}})
    await log_admin_action(request, current_user, "ADMIN_RESTORE", entity, entity_id, True, before={"is_active": old_doc.get("is_active", True)}, after={"is_active": True})
    return {"message": "Restaurado"}


@api_router.get("/admin/movimientos")
async def admin_get_movimientos(
    include_inactive: bool = False,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    return await db.movements.find(movement_active_query(include_deleted=include_inactive), {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.patch("/movements/{movement_id}")
@api_router.put("/admin/movimientos/{movement_id}")
async def admin_update_movimiento(
    movement_id: str,
    payload: MovementAdminUpdate,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    old_doc = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not old_doc or old_doc.get("is_deleted") is True:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")

    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="reason es obligatorio")

    updates = payload.model_dump(exclude_unset=True)
    updates.pop("reason", None)

    if "partida_codigo" in updates:
        await validate_partida(updates["partida_codigo"])
        enforce_capture_budget_scope(current_user, updates["partida_codigo"])

    no_provider_flow = str(updates.get("partida_codigo", old_doc.get("partida_codigo"))) in NO_PROVIDER_BUDGET_CODES
    target_project_id = updates.get("project_id", old_doc.get("project_id"))
    project = await db.projects.find_one({"id": target_project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=400, detail="Proyecto no válido")
    enforce_company_access(current_user, project.get("empresa_id"))

    if no_provider_flow:
        provider_id = updates.get("provider_id", old_doc.get("provider_id"))
        client_id = updates.get("client_id", old_doc.get("client_id"))
        if provider_id:
            raise HTTPException(status_code=422, detail="Las partidas 402/403 no aceptan proveedor")
        if not client_id:
            raise HTTPException(status_code=422, detail={"code": "client_required_for_partida_402_403", "message": "client_id es obligatorio para partidas 402/403"})
        client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
        if not client_doc:
            raise HTTPException(status_code=422, detail={"code": "client_not_found", "message": "Cliente no válido"})
        if client_doc.get("project_id") and client_doc.get("project_id") != target_project_id:
            raise HTTPException(status_code=422, detail={"code": "client_project_mismatch", "message": "El proyecto del movimiento no coincide con el proyecto del cliente"})
        enforce_company_access(current_user, client_doc.get("company_id"))
        inventory_reference = None
        if client_doc.get("inventory_item_id"):
            inventory_item = await db.inventory_items.find_one({"id": client_doc.get("inventory_item_id")}, {"_id": 0})
            if inventory_item:
                inventory_reference = resolve_inventory_reference(inventory_item) or client_doc.get("inventory_item_id")
        updates["provider_id"] = None
        updates["client_id"] = client_id
        updates["customer_name"] = normalize_customer_name(client_doc.get("nombre"))
        if inventory_reference:
            updates["reference"] = inventory_reference
    elif "provider_id" in updates:
        provider = await db.providers.find_one({"id": updates["provider_id"]}, {"_id": 0})
        if not provider:
            raise HTTPException(status_code=400, detail="Proveedor no válido")

    if "amount_original" in updates and updates["amount_original"] <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_amount", "message": "Monto debe ser mayor a 0"})
    if "exchange_rate" in updates and updates["exchange_rate"] <= 0:
        raise HTTPException(status_code=400, detail="Tipo de cambio debe ser mayor a 0")

    if "date" in updates:
        parsed_date = parse_date_tijuana(updates["date"])
        validate_date_in_range(parsed_date)
        updates["date"] = parsed_date.isoformat()

    final_amount_original = updates.get("amount_original", old_doc.get("amount_original"))
    final_exchange_rate = updates.get("exchange_rate", old_doc.get("exchange_rate"))
    if "amount_original" in updates or "exchange_rate" in updates:
        updates["amount_mxn"] = final_amount_original * final_exchange_rate

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    projected = dict(old_doc)
    projected.update(updates)
    if movement_counts_as_abono_doc(projected):
        await validate_client_abono_limit(projected.get("client_id"), decimal_from_value(projected.get("amount_mxn", 0), "amount_mxn"), exclude_movement_id=movement_id)

    await db.movements.update_one({"id": movement_id}, {"$set": updates})
    updated = await db.movements.find_one({"id": movement_id}, {"_id": 0})

    affected_clients = set()
    if old_doc.get("client_id"):
        affected_clients.add(old_doc.get("client_id"))
    if updated.get("client_id"):
        affected_clients.add(updated.get("client_id"))
    for cid in affected_clients:
        await recalc_client_financials(cid)

    await log_admin_action(
        request,
        current_user,
        "ADMIN_UPDATE",
        "movimientos",
        movement_id,
        True,
        before=old_doc,
        after=updated,
        message=reason,
    )
    return updated


@api_router.delete("/movements/{movement_id}")
@api_router.delete("/admin/movimientos/{movement_id}")
async def admin_delete_movimiento(
    movement_id: str,
    payload: MovementAdminAction,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    mov = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not mov or mov.get("is_deleted") is True:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")

    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="reason es obligatorio")

    now_iso = datetime.now(timezone.utc).isoformat()
    after = {
        "is_deleted": True,
        "deleted_at": now_iso,
        "deleted_by": current_user["user_id"],
        "delete_reason": reason,
        "updated_at": now_iso,
    }
    await db.movements.update_one({"id": movement_id}, {"$set": after})
    if mov.get("client_id"):
        await recalc_client_financials(mov.get("client_id"))
    await log_admin_action(request, current_user, "ADMIN_SOFT_DELETE", "movimientos", movement_id, True, before=mov, after=after, message=reason)
    return {"message": "Movimiento eliminado"}


@api_router.delete("/movements/{movement_id}/hard", status_code=204)
@api_router.delete("/admin/movimientos/{movement_id}/hard", status_code=204)
async def admin_hard_delete_movimiento(
    movement_id: str,
    payload: MovementAdminAction,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="reason es obligatorio")
    if request.headers.get("X-Confirm-Hard-Delete") != "HARD-DELETE-MOVEMENT":
        raise HTTPException(status_code=422, detail="Falta confirmación fuerte de hard delete")

    mov = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not mov:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")

    await db.movements.delete_one({"id": movement_id})
    await log_admin_action(request, current_user, "ADMIN_HARD_DELETE", "movimientos", movement_id, True, before=mov, after=None, message=reason)
    return None


@api_router.post("/admin/movements/purge")
async def admin_purge_movements(
    payload: MovementPurgeRequest,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)

    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=422, detail="reason es obligatorio")
    if request.headers.get("X-Confirm-Purge") != "PURGE-MOVEMENTS":
        raise HTTPException(status_code=422, detail="Falta confirmación de purga")

    if payload.hard and request.headers.get("X-Confirm-Hard-Purge") != "HARD-PURGE-MOVEMENTS":
        raise HTTPException(status_code=422, detail="Falta confirmación fuerte de hard purge")

    query = movement_active_query(extra={})
    if payload.project_id:
        query["project_id"] = payload.project_id

    candidates = await db.movements.find(query, {"_id": 0}).to_list(5000)
    if payload.year:
        validate_year_in_range(payload.year)
    if payload.year or payload.month:
        filtered = []
        for m in candidates:
            mov_date = date_parser.parse(m['date']) if isinstance(m['date'], str) else m['date']
            if payload.year and mov_date.year != payload.year:
                continue
            if payload.month and mov_date.month != payload.month:
                continue
            filtered.append(m)
        candidates = filtered

    if not payload.force and not payload.project_id and not payload.year and not payload.month:
        raise HTTPException(status_code=422, detail="Sin filtros: requiere force=true")

    movement_ids = [m.get("id") for m in candidates if m.get("id")]
    count = len(movement_ids)

    if payload.dry_run:
        await log_admin_action(
            request,
            current_user,
            "ADMIN_PURGE_DRY_RUN",
            "movimientos",
            "batch",
            True,
            message=reason,
            after={"filters": payload.model_dump(exclude={"reason"}), "count": count},
        )
        return {"dry_run": True, "count": count}

    if payload.hard:
        if movement_ids:
            await db.movements.delete_many({"id": {"$in": movement_ids}})
    else:
        if movement_ids:
            now_iso = datetime.now(timezone.utc).isoformat()
            await db.movements.update_many(
                {"id": {"$in": movement_ids}},
                {"$set": {
                    "is_deleted": True,
                    "deleted_at": now_iso,
                    "deleted_by": current_user["user_id"],
                    "delete_reason": reason,
                    "updated_at": now_iso,
                }},
            )

    await log_admin_action(
        request,
        current_user,
        "ADMIN_PURGE_HARD" if payload.hard else "ADMIN_PURGE_SOFT",
        "movimientos",
        "batch",
        True,
        message=reason,
        after={"filters": payload.model_dump(exclude={"reason"}), "count": count},
    )
    return {"dry_run": False, "hard": payload.hard, "count": count}


@api_router.post("/admin/movimientos/{movement_id}/reverse")
async def admin_reverse_movement(
    movement_id: str,
    request: Request,
    current_user: dict = Depends(require_permission(Permission.MANAGE_USERS)),
):
    ensure_admin(current_user)
    movement = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    if movement.get("status") != MovementStatus.POSTED.value:
        raise HTTPException(status_code=409, detail="Solo movimientos posted pueden revertirse")

    existing_reversal = await db.movements.find_one({"reversal_of_id": movement_id}, {"_id": 0})
    if existing_reversal:
        raise HTTPException(status_code=409, detail="Movimiento ya tiene reversa")

    reverse_doc = movement.copy()
    reverse_doc["id"] = str(uuid.uuid4())
    reverse_doc["amount_original"] = -abs(float(movement.get("amount_original", 0)))
    reverse_doc["amount_mxn"] = -abs(float(movement.get("amount_mxn", 0)))
    reverse_doc["reference"] = f"REV-{movement.get('reference', movement_id)}"
    reverse_doc["description"] = f"Reversa de {movement_id}. {movement.get('description','')}".strip()
    reverse_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    reverse_doc["created_by"] = current_user["user_id"]
    reverse_doc["reversal_of_id"] = movement_id
    reverse_doc["is_active"] = True

    await db.movements.insert_one(reverse_doc)
    await log_admin_action(request, current_user, "ADMIN_REVERSE", "movimientos", reverse_doc["id"], True, before=movement, after=reverse_doc)
    return reverse_doc


# CORS
raw_cors_origins = os.environ.get('CORS_ORIGINS', 'http://localhost:3000,http://127.0.0.1:3000')
cors_origins = [origin.strip() for origin in raw_cors_origins.split(',') if origin.strip()]
allow_all_origins = '*' in cors_origins

app.add_middleware(
    CORSMiddleware,
    allow_credentials=not allow_all_origins,
    allow_origins=['*'] if allow_all_origins else cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def setup_indexes():
    await db.users.create_index([("email", ASCENDING)], unique=True)
    await db.users.create_index([("name", ASCENDING)], unique=True)
    await db.budget_plans.create_index([("project_id", ASCENDING), ("partida_codigo", ASCENDING)], unique=True)
    await ensure_purchase_order_indexes()
    await ensure_partial_indexes_for_movements()
    await db.movements.create_index([("project_id", ASCENDING), ("date", ASCENDING), ("partida_codigo", ASCENDING), ("is_deleted", ASCENDING)])
    await db.odoo_sync_purchase_orders.create_index([("purchase_order_id", ASCENDING)], unique=True)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


def _period_match(dt: datetime, year: Optional[int], month: Optional[int], period: str):
    if period == "total":
        return True
    if period == "annual":
        return dt.year == year
    if period == "quarterly":
        if dt.year != year:
            return False
        q = ((month or 1) - 1) // 3
        return ((dt.month - 1) // 3) == q
    return dt.year == year and dt.month == month


async def _dashboard_period_data(current_user: dict, empresa_id: Optional[str], project_id: Optional[str], year: Optional[int], month: Optional[int], period: str):
    if empresa_id:
        enforce_company_access(current_user, empresa_id)
    project_query = {}
    if empresa_id:
        project_query["empresa_id"] = empresa_id
    if project_id:
        project_query["id"] = project_id
    projects = await db.projects.find(project_query, {"_id": 0}).to_list(2000)
    if project_id and not projects:
        raise HTTPException(status_code=403, detail="Proyecto fuera de alcance")
    project_ids = [p["id"] for p in projects]

    budget_query = {"project_id": {"$in": project_ids}}
    budgets = await db.budgets.find(budget_query, {"_id": 0}).to_list(5000)

    mov_query = {
        "status": MovementStatus.POSTED.value,
        "project_id": {"$in": project_ids},
    }
    movements = await db.movements.find(movement_active_query(extra=mov_query), {"_id": 0}).to_list(5000)

    by_partida = {}
    total_budget = 0.0
    total_real = 0.0

    for b in budgets:
        y = b.get("year")
        m = b.get("month")
        fake_dt = datetime(y, m, 1)
        if _period_match(fake_dt, year, month, period):
            key = b.get("partida_codigo")
            by_partida.setdefault(key, {"partida_codigo": key, "budget": 0.0, "real": 0.0})
            by_partida[key]["budget"] += float(b.get("amount_mxn", 0))
            total_budget += float(b.get("amount_mxn", 0))

    for mv in movements:
        dt = date_parser.parse(mv["date"])
        if _period_match(dt, year, month, period):
            key = mv.get("partida_codigo")
            by_partida.setdefault(key, {"partida_codigo": key, "budget": 0.0, "real": 0.0})
            real_amount = abs(float(mv.get("amount_mxn", 0)))
            by_partida[key]["real"] += real_amount
            total_real += real_amount

    out_partidas = []
    for _, item in by_partida.items():
        pct = (item["real"] / item["budget"] * 100) if item["budget"] else (100 if item["real"] else 0)
        item["percentage"] = pct
        item["traffic_light"] = get_traffic_light(pct)
        item["variation"] = item["budget"] - item["real"]
        out_partidas.append(item)

    pct_total = (total_real / total_budget * 100) if total_budget else 0
    return {
        "totals": {
            "budget": total_budget,
            "real": total_real,
            "variation": total_budget - total_real,
            "percentage": pct_total,
            "traffic_light": get_traffic_light(pct_total),
        },
        "by_partida": sorted(out_partidas, key=lambda x: x["partida_codigo"] or ""),
        "period": period,
    }


@api_router.get("/dashboard/total")
async def dashboard_total(empresa_id: Optional[str] = None, project_id: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None, current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD))):
    return await _dashboard_period_data(current_user, empresa_id, project_id, year, month, "total")


@api_router.get("/dashboard/monthly")
async def dashboard_monthly(empresa_id: Optional[str] = None, project_id: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None, current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD))):
    now = to_tijuana(datetime.now(timezone.utc))
    return await _dashboard_period_data(current_user, empresa_id, project_id, year or now.year, month or now.month, "monthly")


@api_router.get("/dashboard/quarterly")
async def dashboard_quarterly(empresa_id: Optional[str] = None, project_id: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None, current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD))):
    now = to_tijuana(datetime.now(timezone.utc))
    return await _dashboard_period_data(current_user, empresa_id, project_id, year or now.year, month or now.month, "quarterly")


@api_router.get("/dashboard/annual")
async def dashboard_annual(empresa_id: Optional[str] = None, project_id: Optional[str] = None, year: Optional[int] = None, current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD))):
    now = to_tijuana(datetime.now(timezone.utc))
    return await _dashboard_period_data(current_user, empresa_id, project_id, year or now.year, 1, "annual")


@api_router.get("/dashboard/summary")
async def dashboard_summary_compat(
    empresa_id: Optional[str] = None,
    project_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(require_permission(Permission.VIEW_DASHBOARD)),
):
    """Backward-compatible alias used by legacy tests/UI."""
    now = to_tijuana(datetime.now(timezone.utc))
    return await _dashboard_period_data(current_user, empresa_id, project_id, year or now.year, month or now.month, "monthly")


@api_router.post("/inventory", status_code=201)
async def create_inventory_item(payload: InventoryItemBase, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    enforce_company_access(current_user, payload.company_id)
    precio_venta, precio_total = compute_inventory_totals(payload)
    doc = InventoryItem(**payload.model_dump(), precio_venta=precio_venta, precio_total=precio_total).model_dump()
    doc["created_at"] = doc["created_at"].isoformat(); doc["updated_at"] = doc["updated_at"].isoformat()
    for k in ["m2_superficie", "m2_construccion", "precio_m2_superficie", "precio_m2_construccion", "descuento_bonificacion", "precio_venta", "precio_total"]:
        doc[k] = float(doc[k])
    await db.inventory_items.insert_one(doc)
    await log_audit(current_user, "CREATE", "inventory", doc["id"], {"data": doc})
    return sanitize_mongo_document(doc)




def normalize_invoice_response(doc: dict) -> dict:
    clean = sanitize_mongo_document(doc)
    clean.setdefault("id", clean.get("_id"))
    return clean


@api_router.post("/invoices")
async def create_invoice(payload: InvoiceCreate, current_user: dict = Depends(require_permission(Permission.CREATE_MOVEMENT))):
    project = await db.projects.find_one({"id": payload.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail={"code": "project_not_found", "message": "Proyecto no encontrado"})
    enforce_company_access(current_user, payload.empresa_id)
    if project.get("empresa_id") != payload.empresa_id:
        raise HTTPException(status_code=422, detail={"code": "project_company_mismatch", "message": "El proyecto no corresponde a la empresa"})
    provider_name = payload.provider_name
    if payload.provider_id:
        provider = await db.providers.find_one({"id": payload.provider_id}, {"_id": 0})
        if not provider:
            raise HTTPException(status_code=422, detail={"code": "provider_not_found", "message": "Proveedor no válido"})
        provider_name = provider.get("name")
    exchange_rate_dec = parse_amount_like(payload.exchange_rate if payload.exchange_rate is not None else 1, "exchange_rate")
    if payload.currency != Currency.MXN and exchange_rate_dec <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_exchange_rate", "message": "exchange_rate debe ser mayor a 0 para moneda distinta a MXN"})
    if payload.currency == Currency.MXN:
        exchange_rate_dec = Decimal("1")
    total_original = money_dec(payload.invoice_total_original)
    total_mxn = (total_original * exchange_rate_dec).quantize(TWO_DECIMALS)
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "empresa_id": payload.empresa_id,
        "project_id": payload.project_id,
        "provider_id": payload.provider_id,
        "provider_name": provider_name,
        "invoice_folio": (payload.invoice_folio or "").strip(),
        "currency": payload.currency.value,
        "exchange_rate": str(exchange_rate_dec),
        "invoice_total_original": str(total_original),
        "invoice_total_mxn": str(total_mxn),
        "paid_mxn": "0.00",
        "balance_mxn": str(total_mxn),
        "status": (payload.status or "OPEN").upper(),
        "created_at": now,
        "updated_at": now,
        "created_by": current_user.get("user_id"),
    }
    await db.invoices.insert_one(doc)
    await log_audit(current_user, "CREATE", "invoices", doc["id"], {"folio": doc["invoice_folio"], "provider_id": doc.get("provider_id")})
    return normalize_invoice_response(doc)


@api_router.get("/invoices")
async def list_invoices(empresa_id: Optional[str] = None, project_id: Optional[str] = None, provider_id: Optional[str] = None, q: Optional[str] = None, current_user: dict = Depends(require_permission(Permission.VIEW_MOVEMENTS))):
    query = {}
    if empresa_id:
        enforce_company_access(current_user, empresa_id)
        query["empresa_id"] = empresa_id
    if project_id:
        query["project_id"] = project_id
    if provider_id:
        query["provider_id"] = provider_id
    if q and q.strip():
        safe_q = re.escape(q.strip())
        query["$or"] = [
            {"invoice_folio": {"$regex": safe_q, "$options": "i"}},
            {"provider_name": {"$regex": safe_q, "$options": "i"}},
        ]
    rows = await db.invoices.find(query, {"_id": 0}).to_list(2000)
    for row in rows:
        if row.get("currency") != Currency.MXN.value and not row.get("invoice_total_mxn"):
            er = money_dec(row.get("exchange_rate") or 0)
            original = money_dec(row.get("invoice_total_original") or 0)
            row["invoice_total_mxn"] = str((original * er).quantize(TWO_DECIMALS))
        if not row.get("balance_mxn"):
            total_mxn = money_dec(row.get("invoice_total_mxn") or 0)
            paid = money_dec(row.get("paid_mxn") or 0)
            row["balance_mxn"] = str((total_mxn - paid).quantize(TWO_DECIMALS))
    return [normalize_invoice_response(row) for row in rows]


@api_router.post("/invoices/{invoice_id}/pay")
async def pay_invoice(invoice_id: str, payload: InvoicePayInput, current_user: dict = Depends(require_permission(Permission.CREATE_MOVEMENT))):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail={"code": "invoice_not_found", "message": "Factura no encontrada"})
    enforce_company_access(current_user, invoice.get("empresa_id"))
    mode = (payload.mode or "").upper().strip()
    if mode not in {"ANTICIPO", "LIQUIDACION", "PAGO"}:
        raise HTTPException(status_code=422, detail={"code": "invalid_mode", "message": "mode inválido"})
    total_original = money_dec(invoice.get("invoice_total_original") or 0)
    total_mxn = money_dec(invoice.get("invoice_total_mxn") or 0)
    paid_mxn = money_dec(invoice.get("paid_mxn") or 0)
    balance_mxn = (total_mxn - paid_mxn).quantize(TWO_DECIMALS)
    exchange_rate_dec = money_dec(invoice.get("exchange_rate") or 1)

    if mode == "ANTICIPO":
        if payload.advance_pct is None:
            raise HTTPException(status_code=422, detail={"code": "advance_pct_required", "message": "advance_pct es obligatorio en ANTICIPO"})
        pct = parse_amount_like(payload.advance_pct, "advance_pct")
        if pct <= 0 or pct > 100:
            raise HTTPException(status_code=422, detail={"code": "invalid_advance_pct", "message": "advance_pct debe ser >0 y <=100"})
        amount_original = money_dec(payload.amount_original) if payload.amount_original is not None else (total_original * pct / Decimal("100")).quantize(TWO_DECIMALS)
    elif mode == "LIQUIDACION":
        amount_original = money_dec(payload.amount_original) if payload.amount_original is not None else (balance_mxn / max(exchange_rate_dec, Decimal("0.01"))).quantize(TWO_DECIMALS)
    else:
        amount_original = money_dec(payload.amount_original or 0)

    if amount_original <= 0:
        raise HTTPException(status_code=422, detail={"code": "invalid_amount", "message": "Monto inválido"})
    amount_mxn = (amount_original * exchange_rate_dec).quantize(TWO_DECIMALS)
    if amount_mxn > balance_mxn:
        raise HTTPException(status_code=422, detail={"code": "payment_exceeds_balance", "message": "El pago excede el saldo"})

    movement_doc = {
        "id": str(uuid.uuid4()),
        "project_id": invoice.get("project_id"),
        "partida_codigo": payload.partida_codigo,
        "provider_id": invoice.get("provider_id"),
        "date": parse_date_tijuana(payload.date).isoformat(),
        "currency": invoice.get("currency") or Currency.MXN.value,
        "amount_original": float(amount_original),
        "exchange_rate": float(exchange_rate_dec),
        "amount_mxn": float(amount_mxn),
        "reference": payload.reference,
        "description": payload.description or f"Pago {mode} factura {invoice.get('invoice_folio')}",
        "status": MovementStatus.POSTED.value,
        "created_by": current_user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "invoice_id": invoice_id,
        "payment_mode": mode,
    }
    await db.movements.insert_one(movement_doc)

    new_paid = (paid_mxn + amount_mxn).quantize(TWO_DECIMALS)
    new_balance = (total_mxn - new_paid).quantize(TWO_DECIMALS)
    status = "PAID" if new_balance <= 0 else ("PARTIAL" if new_paid > 0 else "OPEN")
    await db.invoices.update_one({"id": invoice_id}, {"$set": {"paid_mxn": str(new_paid), "balance_mxn": str(new_balance), "status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
    updated = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    await log_audit(current_user, "PAY", "invoices", invoice_id, {"mode": mode, "amount_mxn": str(amount_mxn), "movement_id": movement_doc["id"]})
    return {"invoice": normalize_invoice_response(updated), "movement": sanitize_mongo_document(movement_doc)}



@api_router.post("/inventory/import-csv")
async def import_inventory_csv(file: UploadFile = File(...), dry_run: bool = Query(False), current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    content = (await file.read()).decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(content))
    created = updated = skipped = 0
    errors = []
    samples = []
    row_num = 1
    for row in reader:
        row_num += 1
        try:
            code = (row.get("code") or row.get("sku") or "").strip()
            company_id = (row.get("company_id") or "").strip()
            project_id = (row.get("project_id") or "").strip()
            lote = (row.get("lote_edificio") or row.get("lote") or "").strip()
            manzana = (row.get("manzana_departamento") or row.get("manzana") or "").strip()
            if not company_id or not project_id or not lote or not manzana:
                raise ValueError("company_id, project_id, lote_edificio y manzana_departamento son obligatorios")
            payload = {
                "company_id": company_id,
                "project_id": project_id,
                "m2_superficie": float(row.get("m2_superficie") or 0),
                "m2_construccion": float(row.get("m2_construccion") or 0),
                "lote_edificio": lote,
                "manzana_departamento": manzana,
                "precio_m2_superficie": float(row.get("precio_m2_superficie") or 0),
                "precio_m2_construccion": float(row.get("precio_m2_construccion") or 0),
                "descuento_bonificacion": float(row.get("descuento_bonificacion") or 0),
                "code": code or None,
            }
            samples.append(payload)
            query = {"code": code} if code else {"company_id": company_id, "project_id": project_id, "lote_edificio": lote, "manzana_departamento": manzana}
            existing = await db.inventory_items.find_one(query, {"_id": 0})
            if dry_run:
                if existing: updated += 1
                else: created += 1
                continue
            if existing:
                item = InventoryItem(**{k: payload[k] for k in ["company_id","project_id","m2_superficie","m2_construccion","lote_edificio","manzana_departamento","precio_m2_superficie","precio_m2_construccion","descuento_bonificacion"]},
                    precio_venta=(Decimal(str(payload["m2_superficie"])) * Decimal(str(payload["precio_m2_superficie"]))),
                    precio_total=(Decimal(str(payload["m2_superficie"])) * Decimal(str(payload["precio_m2_superficie"])) + Decimal(str(payload["m2_construccion"])) * Decimal(str(payload["precio_m2_construccion"])) - Decimal(str(payload["descuento_bonificacion"]))))
                set_doc = item.model_dump()
                set_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
                set_doc["created_at"] = existing.get("created_at") or datetime.now(timezone.utc).isoformat()
                set_doc["precio_venta"] = float(set_doc["precio_venta"])
                set_doc["precio_total"] = float(set_doc["precio_total"])
                if code:
                    set_doc["code"] = code
                await db.inventory_items.update_one({"id": existing.get("id")}, {"$set": set_doc})
                updated += 1
            else:
                item = InventoryItem(**{k: payload[k] for k in ["company_id","project_id","m2_superficie","m2_construccion","lote_edificio","manzana_departamento","precio_m2_superficie","precio_m2_construccion","descuento_bonificacion"]},
                    precio_venta=(Decimal(str(payload["m2_superficie"])) * Decimal(str(payload["precio_m2_superficie"]))),
                    precio_total=(Decimal(str(payload["m2_superficie"])) * Decimal(str(payload["precio_m2_superficie"])) + Decimal(str(payload["m2_construccion"])) * Decimal(str(payload["precio_m2_construccion"])) - Decimal(str(payload["descuento_bonificacion"]))))
                doc = item.model_dump()
                doc["created_at"] = doc["created_at"].isoformat(); doc["updated_at"] = doc["updated_at"].isoformat(); doc["precio_venta"] = float(doc["precio_venta"]); doc["precio_total"] = float(doc["precio_total"])
                if code:
                    doc["code"] = code
                await db.inventory_items.insert_one(doc)
                created += 1
        except Exception as ex:
            skipped += 1
            errors.append({"row_number": row_num, "message": str(ex)})
    await log_audit(current_user, "IMPORT", "inventory", "batch", {"filename": file.filename, "created": created, "updated": updated, "skipped": skipped, "errors": len(errors), "dry_run": dry_run})
    return {"created_count": created, "updated_count": updated, "skipped_count": skipped, "errors": errors, "sample_rows": samples[:20] if dry_run else []}


@api_router.post("/clients/import-csv")
async def import_clients_csv(file: UploadFile = File(...), dry_run: bool = Query(False), current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    content = (await file.read()).decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(content))
    created = updated = skipped = 0
    errors = []
    samples = []
    row_num = 1
    for row in reader:
        row_num += 1
        try:
            company_id = (row.get("company_id") or "").strip()
            project_id = (row.get("project_id") or "").strip()
            nombre = (row.get("nombre") or row.get("name") or "").strip()
            telefono = (row.get("telefono") or row.get("phone") or "").strip() or None
            domicilio = (row.get("domicilio") or row.get("address") or "").strip() or None
            code = (row.get("code") or "").strip()
            if not company_id or not project_id or not nombre:
                raise ValueError("company_id, project_id y nombre son obligatorios")
            payload = {"company_id": company_id, "project_id": project_id, "nombre": nombre, "telefono": telefono, "domicilio": domicilio, "code": code or None}
            samples.append(payload)
            query = {"code": code} if code else {"company_id": company_id, "project_id": project_id, "nombre": nombre, "telefono": telefono}
            existing = await db.clients.find_one(query, {"_id": 0})
            if dry_run:
                if existing: updated += 1
                else: created += 1
                continue
            if existing:
                await db.clients.update_one({"id": existing.get("id")}, {"$set": {"nombre": nombre, "telefono": telefono, "domicilio": domicilio, "updated_at": datetime.now(timezone.utc).isoformat(), "code": code or existing.get("code")}})
                updated += 1
            else:
                doc = Client(company_id=company_id, project_id=project_id, nombre=nombre, telefono=telefono, domicilio=domicilio, inventory_item_id=None).model_dump()
                doc["created_at"] = doc["created_at"].isoformat(); doc["updated_at"] = doc["updated_at"].isoformat(); doc["precio_venta_snapshot"] = 0.0; doc["abonos_total_mxn"] = 0.0; doc["saldo_restante"] = 0.0; doc["code"] = code or None
                await db.clients.insert_one(doc)
                created += 1
        except Exception as ex:
            skipped += 1
            errors.append({"row_number": row_num, "message": str(ex)})
    await log_audit(current_user, "IMPORT", "clients", "batch", {"filename": file.filename, "created": created, "updated": updated, "skipped": skipped, "errors": len(errors), "dry_run": dry_run})
    return {"created_count": created, "updated_count": updated, "skipped_count": skipped, "errors": errors, "sample_rows": samples[:20] if dry_run else []}

@api_router.get("/inventory")
async def list_inventory(company_id: Optional[str] = None, project_id: Optional[str] = None, current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    query = {}
    scope = user_company_scope_query(current_user)
    query.update(scope)
    if company_id:
        enforce_company_access(current_user, company_id)
        query["company_id"] = company_id
    if project_id:
        query["project_id"] = project_id
    return await db.inventory_items.find(query, {"_id": 0}).to_list(5000)


async def _resolve_company_id(raw_company: Optional[str]) -> str:
    company_value = (raw_company or "").strip()
    if not company_value:
        raise HTTPException(status_code=422, detail={"code": "company_required", "message": "empresa/company_id es obligatorio"})
    empresa = await db.empresas.find_one({"id": company_value}, {"_id": 0})
    if empresa:
        return empresa.get("id")
    empresa = await db.empresas.find_one({"nombre": company_value}, {"_id": 0})
    if empresa:
        return empresa.get("id")
    raise HTTPException(status_code=422, detail={"code": "invalid_company", "message": "empresa/company_id inválido"})


async def _resolve_project_id(raw_project: Optional[str], company_id: str) -> str:
    project_value = (raw_project or "").strip()
    if not project_value:
        raise HTTPException(status_code=422, detail={"code": "project_required", "message": "proyecto/project_id es obligatorio"})
    project = await db.projects.find_one({"id": project_value}, {"_id": 0})
    if not project:
        project = await db.projects.find_one({"code": project_value}, {"_id": 0})
    if not project and "-" in project_value:
        code = project_value.split("-", 1)[0].strip()
        project = await db.projects.find_one({"code": code}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=422, detail={"code": "invalid_project", "message": "project_id/proyecto inválido"})
    if project.get("empresa_id") != company_id:
        raise HTTPException(status_code=422, detail={"code": "project_company_mismatch", "message": "project_id no pertenece a company_id"})
    return project.get("id")


async def _resolve_inventory_item_id(raw_inventory: Optional[str], company_id: str, project_id: str) -> Optional[str]:
    value = (raw_inventory or "").strip()
    if not value:
        return None
    if value.lower() in {"none", "null", "sin asignar"}:
        return None

    item = await db.inventory_items.find_one({"id": value}, {"_id": 0})
    if not item:
        if "-" in value:
            lote, manzana = [part.strip() for part in value.split("-", 1)]
            item = await db.inventory_items.find_one({"lote_edificio": lote, "manzana_departamento": manzana, "company_id": company_id, "project_id": project_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=422, detail={"code": "inventory_not_found", "message": "Inventario no encontrado"})
    if item.get("company_id") != company_id:
        raise HTTPException(status_code=422, detail={"code": "inventory_company_mismatch", "message": "inventory_item_id no pertenece a la empresa seleccionada"})
    if item.get("project_id") != project_id:
        raise HTTPException(status_code=422, detail={"code": "inventory_project_mismatch", "message": "inventory_item_id no pertenece al proyecto seleccionado"})
    return item.get("id")


async def _normalize_client_create_payload(payload: ClientCreateRequest) -> ClientBase:
    company_id = await _resolve_company_id(payload.company_id or payload.empresa)
    project_id = await _resolve_project_id(payload.project_id or payload.proyecto, company_id)
    inventory_item_id = await _resolve_inventory_item_id(payload.inventory_item_id or payload.inventario, company_id, project_id)

    nombre = (payload.nombre or "").strip().upper()
    if not nombre:
        raise HTTPException(status_code=422, detail={"code": "name_required", "message": "nombre es obligatorio"})

    return ClientBase(
        company_id=company_id,
        project_id=project_id,
        nombre=nombre,
        telefono=(payload.telefono or "").strip() or None,
        domicilio=(payload.domicilio or "").strip() or None,
        inventory_item_id=inventory_item_id,
    )


@api_router.post("/clients", status_code=201)
async def create_client(payload: ClientCreateRequest, current_user: dict = Depends(require_client_write_access())):
    logger.info("POST /api/clients payload=%s", {
        "company_id": payload.company_id,
        "empresa": payload.empresa,
        "project_id": payload.project_id,
        "proyecto": payload.proyecto,
        "nombre": (payload.nombre or "").strip(),
        "telefono": (payload.telefono or "").strip(),
        "domicilio": (payload.domicilio or "").strip(),
        "inventory_item_id": payload.inventory_item_id,
        "inventario": payload.inventario,
    })

    normalized = await _normalize_client_create_payload(payload)
    enforce_company_access(current_user, normalized.company_id)

    snapshot = Decimal("0")
    if normalized.inventory_item_id:
        inventory_item = await db.inventory_items.find_one({"id": normalized.inventory_item_id}, {"_id": 0})
        if not inventory_item:
            raise HTTPException(status_code=422, detail={"code": "inventory_not_found", "message": "Inventario no encontrado"})
        existing_for_inventory = await db.clients.find_one({"inventory_item_id": normalized.inventory_item_id}, {"_id": 0})
        if existing_for_inventory:
            raise HTTPException(status_code=409, detail={"code": "inventory_already_linked", "message": "El inventario seleccionado ya está ligado a otro cliente"})
        snapshot = decimal_from_value(inventory_item.get("precio_total", 0), "precio_total")

    duplicate = await db.clients.find_one({
        "company_id": normalized.company_id,
        "project_id": normalized.project_id,
        "nombre": normalized.nombre,
        "inventory_item_id": normalized.inventory_item_id,
    }, {"_id": 0})
    if duplicate:
        raise HTTPException(status_code=422, detail={"code": "duplicate_client", "message": "Cliente duplicado para la combinación empresa/proyecto/inventario"})

    doc = Client(**normalized.model_dump(), precio_venta_snapshot=snapshot, saldo_restante=snapshot).model_dump()
    doc["created_at"] = doc["created_at"].isoformat(); doc["updated_at"] = doc["updated_at"].isoformat()
    doc["precio_venta_snapshot"] = float(doc["precio_venta_snapshot"])
    doc["abonos_total_mxn"] = 0.0
    doc["saldo_restante"] = float(doc["saldo_restante"])
    try:
        await db.clients.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=422, detail={"code": "client_create_conflict", "message": "Conflicto al crear cliente (revisa inventario ligado o datos duplicados)"})
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error creating client")
        raise HTTPException(status_code=422, detail={"code": "client_create_failed", "message": "No se pudo crear el cliente con los datos enviados"})
    await log_audit(current_user, "CREATE", "clients", doc["id"], {"data": doc})
    return sanitize_mongo_document(doc)


@api_router.get("/clients")
async def list_clients(company_id: Optional[str] = None, project_id: Optional[str] = None, current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    query = {}
    query.update(user_company_scope_query(current_user))
    if company_id:
        enforce_company_access(current_user, company_id)
        query["company_id"] = company_id
    if project_id:
        query["project_id"] = project_id

    clients = await db.clients.find(query, {"_id": 0}).to_list(5000)
    inventory_ids = [c.get("inventory_item_id") for c in clients if c.get("inventory_item_id")]
    inventory_map = {}
    if inventory_ids:
        inventory_rows = await db.inventory_items.find({"id": {"$in": inventory_ids}}, {"_id": 0}).to_list(5000)
        inventory_map = {it.get("id"): it for it in inventory_rows}

    enriched = []
    for client_doc in clients:
        cid = client_doc.get("id")
        updated = await recalc_client_financials(cid) if cid else client_doc
        row = updated or client_doc
        valor_total = float(row.get("precio_venta_snapshot", 0) or 0)
        abonos = float(row.get("abonos_total_mxn", 0) or 0)
        saldo = valor_total - abonos
        if saldo < 0:
            saldo = 0.0
        item = inventory_map.get(row.get("inventory_item_id"))
        row["inventory_clave"] = get_inventory_clave(item)
        row["valor_total_mxn"] = valor_total
        row["abonos_total_mxn"] = abonos
        row["saldo_restante_mxn"] = saldo
        enriched.append(sanitize_mongo_document(row))
    return enriched




@api_router.put("/inventory/{item_id}")
async def update_inventory_item(item_id: str, payload: InventoryItemUpdate, current_user: dict = Depends(require_permission(Permission.MANAGE_CATALOGS))):
    item = await db.inventory_items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
    enforce_company_access(current_user, item.get("company_id"))

    merged = {
        "company_id": item.get("company_id"),
        "project_id": payload.project_id or item.get("project_id"),
        "m2_superficie": payload.m2_superficie if payload.m2_superficie is not None else item.get("m2_superficie"),
        "m2_construccion": payload.m2_construccion if payload.m2_construccion is not None else item.get("m2_construccion", 0),
        "lote_edificio": payload.lote_edificio if payload.lote_edificio is not None else item.get("lote_edificio"),
        "manzana_departamento": payload.manzana_departamento if payload.manzana_departamento is not None else item.get("manzana_departamento"),
        "precio_m2_superficie": payload.precio_m2_superficie if payload.precio_m2_superficie is not None else item.get("precio_m2_superficie"),
        "precio_m2_construccion": payload.precio_m2_construccion if payload.precio_m2_construccion is not None else item.get("precio_m2_construccion", 0),
        "descuento_bonificacion": payload.descuento_bonificacion if payload.descuento_bonificacion is not None else item.get("descuento_bonificacion", 0),
    }
    base = InventoryItemBase(**merged)
    precio_venta, precio_total = compute_inventory_totals(base)

    update_data = base.model_dump()
    update_data["precio_venta"] = float(precio_venta)
    update_data["precio_total"] = float(precio_total)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    for k in ["m2_superficie", "m2_construccion", "precio_m2_superficie", "precio_m2_construccion", "descuento_bonificacion"]:
        update_data[k] = float(update_data[k])

    await db.inventory_items.update_one({"id": item_id}, {"$set": update_data})
    updated = await db.inventory_items.find_one({"id": item_id}, {"_id": 0})
    await log_audit(current_user, "UPDATE", "inventory", item_id, {"before": item, "after": updated})
    return sanitize_mongo_document(updated)


@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, payload: ClientUpdate, current_user: dict = Depends(require_client_write_access())):
    client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    enforce_company_access(current_user, client_doc.get("company_id"))

    update_data = {}
    if payload.nombre is not None:
        nombre = payload.nombre.strip().upper()
        if not nombre:
            raise HTTPException(status_code=422, detail={"code": "name_required", "message": "nombre es obligatorio"})
        update_data["nombre"] = nombre
    if payload.telefono is not None:
        update_data["telefono"] = payload.telefono
    if payload.domicilio is not None:
        update_data["domicilio"] = payload.domicilio
    if payload.inventory_item_id is not None:
        if payload.inventory_item_id:
            inventory_item = await db.inventory_items.find_one({"id": payload.inventory_item_id}, {"_id": 0})
            if not inventory_item:
                raise HTTPException(status_code=422, detail={"code": "inventory_not_found", "message": "inventory_item_id inválido"})
            if inventory_item.get("company_id") != client_doc.get("company_id") or inventory_item.get("project_id") != client_doc.get("project_id"):
                raise HTTPException(status_code=422, detail={"code": "inventory_scope_mismatch", "message": "El inventario no pertenece a la misma empresa/proyecto del cliente"})
            existing_for_inventory = await db.clients.find_one({"inventory_item_id": payload.inventory_item_id}, {"_id": 0})
            if existing_for_inventory and existing_for_inventory.get("id") != client_id:
                raise HTTPException(status_code=422, detail={"code": "inventory_already_linked", "message": "El inventario seleccionado ya está ligado a otro cliente"})
            update_data["inventory_item_id"] = payload.inventory_item_id
            snapshot = decimal_from_value(inventory_item.get("precio_total", 0), "precio_total")
            update_data["precio_venta_snapshot"] = float(snapshot)
            if decimal_from_value(client_doc.get("saldo_restante", 0), "saldo_restante") <= Decimal("0"):
                update_data["saldo_restante"] = float(snapshot)
        else:
            update_data["inventory_item_id"] = None

    if not update_data:
        return sanitize_mongo_document(client_doc)

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    updated = await recalc_client_financials(client_id) or updated
    await log_audit(current_user, "UPDATE", "clients", client_id, {"before": client_doc, "after": updated})
    return sanitize_mongo_document(updated)


@api_router.get("/inventory/summary")
async def inventory_summary(company_id: Optional[str] = None, project_id: Optional[str] = None, current_user: dict = Depends(require_permission(Permission.VIEW_CATALOGS))):
    query = {}
    query.update(user_company_scope_query(current_user))
    if company_id:
        enforce_company_access(current_user, company_id)
        query["company_id"] = company_id
    if project_id:
        query["project_id"] = project_id

    items = await db.inventory_items.find(query, {"_id": 0}).to_list(5000)
    item_ids = [it.get("id") for it in items if it.get("id")]
    valor_total = sum(float(it.get("precio_total", 0) or 0) for it in items)

    clients_query = {}
    if item_ids:
        clients_query["inventory_item_id"] = {"$in": item_ids}
    else:
        clients_query["inventory_item_id"] = {"$in": []}
    clients = await db.clients.find(clients_query, {"_id": 0}).to_list(5000)

    cobrado = 0.0
    for c in clients:
        if c.get("id"):
            updated = await recalc_client_financials(c["id"])
            cobrado += float((updated or c).get("abonos_total_mxn", 0) or 0)
    restante = valor_total - cobrado
    if restante < 0:
        restante = 0.0

    return {
        "valor_total_inventario_mxn": valor_total,
        "cobrado_mxn": cobrado,
        "restante_por_cobrar_mxn": restante,
        "counts": {
            "items_count": len(items),
            "clients_count": len(clients),
        },
    }


@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, request: Request, current_user: dict = Depends(require_permission(Permission.MANAGE_USERS))):
    ensure_admin(current_user)
    client_doc = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client_doc:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    enforce_company_access(current_user, client_doc.get("company_id"))
    movement_ref = await db.movements.find_one(movement_active_query(extra={"client_id": client_id}), {"_id": 0})
    if movement_ref:
        raise HTTPException(status_code=409, detail="No se puede eliminar cliente con movimientos relacionados")
    await db.clients.delete_one({"id": client_id})
    await log_admin_action(request, current_user, "ADMIN_HARD_DELETE", "clients", client_id, True, before=client_doc, after=None, message="delete client")
    return {"message": "Cliente eliminado"}


@api_router.delete("/inventory/{item_id}")
async def delete_inventory(item_id: str, request: Request, current_user: dict = Depends(require_permission(Permission.MANAGE_USERS))):
    ensure_admin(current_user)
    item_doc = await db.inventory_items.find_one({"id": item_id}, {"_id": 0})
    if not item_doc:
        raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
    enforce_company_access(current_user, item_doc.get("company_id"))
    linked_client = await db.clients.find_one({"inventory_item_id": item_id}, {"_id": 0})
    if linked_client:
        raise HTTPException(status_code=409, detail="No se puede eliminar inventario ligado a clientes")
    await db.inventory_items.delete_one({"id": item_id})
    await log_admin_action(request, current_user, "ADMIN_HARD_DELETE", "inventory", item_id, True, before=item_doc, after=None, message="delete inventory")
    return {"message": "Inventario eliminado"}


async def find_movement_for_receipt(movement_id: str) -> Optional[dict]:
    lookup_id = to_optional_str(movement_id)
    if not lookup_id:
        return None

    exact = await db.movements.find_one({"id": lookup_id}, {"_id": 0})
    if exact:
        return exact

    prefix = lookup_id.split("_", 1)[0].strip()
    if prefix and prefix != lookup_id:
        prefixed = await db.movements.find_one({"id": prefix}, {"_id": 0})
        if prefixed:
            return prefixed

    candidates = await db.movements.find({}, {"_id": 0}).to_list(5000)
    for mov in candidates:
        mov_id = to_optional_str(mov.get("id"))
        if not mov_id:
            continue
        if mov_id == lookup_id:
            return mov
        if mov_id.startswith(f"{prefix}_") and prefix:
            return mov
        if lookup_id.startswith(f"{mov_id}_"):
            return mov
    return None

@api_router.get("/movements/{movement_id}/receipt.pdf")
async def movement_receipt_pdf(movement_id: str, current_user: dict = Depends(require_permission(Permission.VIEW_MOVEMENTS))):
    movement = await find_movement_for_receipt(movement_id)
    if not movement:
        raise HTTPException(status_code=404, detail={"code": "movement_not_found", "message": "Movimiento no encontrado", "movement_id": movement_id})
    project = await db.projects.find_one({"id": movement.get("project_id")}, {"_id": 0})
    if project:
        enforce_company_access(current_user, project.get("empresa_id"))
    empresa = await db.empresas.find_one({"id": project.get("empresa_id")}, {"_id": 0}) if project else None

    partida = str(movement.get("partida_codigo") or "")
    is_ingreso = partida.startswith("4")
    if current_user.get("role") == UserRole.CAPTURA_INGRESOS.value and not is_ingreso:
        raise HTTPException(status_code=403, detail="Rol captura_ingresos solo puede imprimir recibos de ingresos (4xx)")

    client_doc = await db.clients.find_one({"id": movement.get("client_id")}, {"_id": 0}) if movement.get("client_id") else None
    inventory_item = None

    if not client_doc and partida in ABONO_PARTIDAS and movement.get("reference") and project:
        ref_value = movement.get("reference")
        inv_queries = [
            {"id": ref_value, "company_id": project.get("empresa_id"), "project_id": movement.get("project_id")},
            {"reference": ref_value, "company_id": project.get("empresa_id"), "project_id": movement.get("project_id")},
            {"ref": ref_value, "company_id": project.get("empresa_id"), "project_id": movement.get("project_id")},
            {"lote_edificio": ref_value, "company_id": project.get("empresa_id"), "project_id": movement.get("project_id")},
        ]
        for inv_query in inv_queries:
            inventory_item = await db.inventory_items.find_one(inv_query, {"_id": 0})
            if inventory_item:
                client_doc = await db.clients.find_one({"inventory_item_id": inventory_item.get("id")}, {"_id": 0})
                if client_doc:
                    break

    if not client_doc:
        fallback_name = normalize_customer_name(movement.get("customer_name"))
        if fallback_name and movement.get("project_id"):
            legacy_query = {"project_id": movement.get("project_id"), "nombre": fallback_name.upper()}
            if project and project.get("empresa_id"):
                legacy_query["company_id"] = project.get("empresa_id")
            client_doc = await db.clients.find_one(legacy_query, {"_id": 0})

    if client_doc and client_doc.get("id") and movement.get("client_id") != client_doc.get("id"):
        movement["client_id"] = client_doc.get("id")
        await db.movements.update_one({"id": movement.get("id")}, {"$set": {"client_id": client_doc.get("id"), "updated_at": datetime.now(timezone.utc).isoformat()}})

    if client_doc:
        client_doc = await recalc_client_financials(client_doc.get("id")) or client_doc
    if not inventory_item and client_doc and client_doc.get("inventory_item_id"):
        inventory_item = await db.inventory_items.find_one({"id": client_doc.get("inventory_item_id")}, {"_id": 0})

    lines = [
        "RECIBO DE ABONO - QUANTUM",
        f"Folio: {movement.get('id')}",
        f"Fecha emisión: {to_tijuana(datetime.now(timezone.utc)).strftime('%Y-%m-%d %H:%M:%S')}",
        f"Cliente: {client_doc.get('nombre') if client_doc else movement.get('customer_name') or 'S/I'}",
        f"Empresa: {empresa.get('nombre') if empresa else ''}",
        f"Proyecto: {project.get('code') if project else ''} - {project.get('name') if project else ''}",
        f"Inventario: {get_inventory_clave(inventory_item) or (movement.get('reference') or 'N/A')}",
        f"Partida: {movement.get('partida_codigo')}",
        f"Referencia: {movement.get('reference')}",
        f"Moneda: {movement.get('currency')}",
        f"Monto original: {movement.get('amount_original')}",
        f"Tipo cambio: {movement.get('exchange_rate')}",
        f"Monto MXN: {movement.get('amount_mxn')}",
        f"Descripción: {movement.get('description') or ''}",
        f"Valor total MXN: {client_doc.get('precio_venta_snapshot') if client_doc else ''}",
        f"Abonos acumulados MXN: {client_doc.get('abonos_total_mxn') if client_doc else ''}",
        f"Saldo restante MXN: {client_doc.get('saldo_restante') if client_doc else ''}",
    ]
    pdf_bytes = render_basic_pdf(lines)
    await log_audit(current_user, "RECEIPT_PDF", "movements", movement_id, {"reference": movement.get("reference")})
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers={"Content-Disposition": f"inline; filename=recibo_{movement_id}.pdf"})


# Include router
app.include_router(api_router)
